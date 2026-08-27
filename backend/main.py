import os
import time
import secrets
import hashlib
import hmac
import secrets


from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo

import httpx
import psycopg

from psycopg.rows import dict_row

from fastapi import (
    FastAPI,
    HTTPException,
    Response,
    File,
    UploadFile,
    Header,
    Depends,
)

from fastapi.security import (
    HTTPAuthorizationCredentials,
    HTTPBearer,
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

from openpyxl import load_workbook
from pydantic import BaseModel


# =========================
# FastAPI
# =========================

app = FastAPI()


# =========================
# CORS
# =========================

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://127.0.0.1:5500",
        "http://localhost:5500",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


# =========================
# 파일 경로
# =========================

BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent

FRONTEND_DIR = PROJECT_DIR / "frontend"

EXCEL_PATH = BASE_DIR / "data" / "matches.xlsx"
RESULTS_PATH = BASE_DIR / "data" / "results.xlsx"
PLAYER_RANKINGS_PATH = BASE_DIR / "data" / "player_rankings.xlsx"
PLAYOFFS_PATH = BASE_DIR / "data" / "playoffs.xlsx"

# =========================
# NEXON Open API
# =========================

NEXON_API_KEY = os.getenv(
    "NEXON_API_KEY"
)

DATABASE_URL = os.getenv(
    "DATABASE_URL"
)

ADMIN_PASSWORD = os.getenv(
    "ADMIN_PASSWORD"
)


ADMIN_SESSION_SECONDS = (
    8 * 60 * 60
)


ADMIN_SESSIONS = {}

NEXON_API_BASE_URL = (
    "https://open.api.nexon.com/fconline/v1"
)

def get_community_post_attachments(
    cursor,
    post_id: int,
):

    cursor.execute(
        """
        SELECT
            id,
            original_file_name,
            content_type,
            sort_order

        FROM community_attachments

        WHERE post_id = %s

        ORDER BY
            sort_order ASC,
            id ASC
        """,
        (
            post_id,
        ),
    )

    rows = cursor.fetchall()

    return [
        {
            "id": row["id"],
            "original_file_name": row["original_file_name"],
            "content_type": row["content_type"],
            "sort_order": row["sort_order"],
            "image_url": (
                "/api/community/"
                f"attachments/{row['id']}"
            ),
        }
        for row in rows
    ]

# =========================
# FC Online 선수 메타데이터
# =========================

SPID_METADATA_CACHE = None
SEASON_METADATA_CACHE = None


def get_spid_metadata():

    global SPID_METADATA_CACHE


    if SPID_METADATA_CACHE is not None:
        return SPID_METADATA_CACHE


    url = (
        "https://open.api.nexon.com/"
        "static/fconline/meta/spid.json"
    )


    response = httpx.get(
        url,
        timeout=20.0,
    )


    if response.status_code != 200:

        raise HTTPException(
            status_code=response.status_code,
            detail="선수 메타데이터 조회 실패",
        )


    SPID_METADATA_CACHE = {
        player["id"]: player["name"]
        for player in response.json()
    }


    return SPID_METADATA_CACHE

def get_season_metadata():

    global SEASON_METADATA_CACHE


    if SEASON_METADATA_CACHE is not None:
        return SEASON_METADATA_CACHE


    url = (
        "https://open.api.nexon.com/"
        "static/fconline/meta/seasonid.json"
    )


    response = httpx.get(
        url,
        timeout=20.0,
    )


    if response.status_code != 200:

        raise HTTPException(
            status_code=response.status_code,
            detail="시즌 메타데이터 조회 실패",
        )


    SEASON_METADATA_CACHE = {
        int(season["seasonId"]): {
            "season_id":
                int(season["seasonId"]),

            "class_name":
                season["className"],

            "season_image_url":
                season["seasonImg"],
        }

        for season
        in response.json()
    }


    return SEASON_METADATA_CACHE


def get_player_season_info(
    sp_id,
    season_metadata,
):

    try:
        season_id = (
            int(sp_id)
            // 1_000_000
        )

    except (
        TypeError,
        ValueError,
    ):
        return None


    return season_metadata.get(
        season_id
    )

def ensure_fconline_season_snapshot(
    sp_id,
):

    season_id = (
        int(sp_id)
        // 1_000_000
    )


    # =========================
    # 이미 저장됐으면 종료
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    season_id,
                    class_name
                FROM fconline_season_snapshots
                WHERE season_id = %s
                """,
                (
                    season_id,
                ),
            )


            existing_snapshot = cursor.fetchone()


    if existing_snapshot:

        return {
            "season_id":
                existing_snapshot[
                    "season_id"
                ],

            "class_name":
                existing_snapshot[
                    "class_name"
                ],

            "season_image_url":
                (
                    "/api/fconline/"
                    "metadata/seasons/"
                    f"{season_id}/image"
                ),
        }


    # =========================
    # Nexon 시즌 메타데이터
    # =========================

    season_metadata = get_season_metadata()


    season = season_metadata.get(
            season_id
        )


    if not season:

        raise HTTPException(
            status_code=404,
            detail=(
                f"시즌 메타데이터 없음: "
                f"{season_id}"
            ),
        )


    source_image_url = season[
            "season_image_url"
        ]


    # =========================
    # 시즌 이미지 실제 다운로드
    # =========================

    image_response = httpx.get(
            source_image_url,
            timeout=20.0,
            follow_redirects=True,
        )


    if image_response.status_code != 200:

        raise HTTPException(
            status_code=
                image_response.status_code,

            detail=(
                "시즌 이미지 다운로드 실패"
            ),
        )


    image_content_type = (
        image_response.headers.get(
            "content-type",
            "image/png",
        )
        .split(";")[0]
        .strip()
    )


    image_data = image_response.content


    if not image_data:

        raise HTTPException(
            status_code=502,
            detail="시즌 이미지 데이터 없음",
        )


    # =========================
    # Neon 영구 Snapshot
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO
                    fconline_season_snapshots (
                        season_id,
                        class_name,
                        source_image_url,
                        image_data,
                        image_content_type
                    )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )

                ON CONFLICT (
                    season_id
                )
                DO NOTHING
                """,
                (
                    season_id,

                    season[
                        "class_name"
                    ],

                    source_image_url,
                    image_data,
                    image_content_type,
                ),
            )


        connection.commit()


    return {
        "season_id":
            season_id,

        "class_name":
            season[
                "class_name"
            ],

        "season_image_url":
            (
                "/api/fconline/"
                "metadata/seasons/"
                f"{season_id}/image"
            ),
    }

def get_player_name(
    sp_id,
    spid_metadata,
):

    return spid_metadata.get(
        sp_id,
        f"알 수 없는 선수 ({sp_id})",
    )


def get_player_image_url(sp_id):

    return (
        "https://fco.dn.nexoncdn.co.kr/"
        "live/externalAssets/common/"
        f"playersAction/p{sp_id}.png"
    )


# =========================
# PostgreSQL
# =========================

def parse_kst_datetime(value):


    if isinstance(value, datetime):
        parsed = value

    else:
        parsed = datetime.fromisoformat(value)


    if parsed.tzinfo is None:
        parsed = parsed.replace(
            tzinfo=ZoneInfo("Asia/Seoul")
        )


    return parsed.astimezone(
        ZoneInfo("Asia/Seoul")
    )

def parse_nexon_datetime(value):

    if isinstance(value, datetime):
        parsed = value

    else:
        parsed = datetime.fromisoformat(
            value
        )


    if parsed.tzinfo is None:

        parsed = parsed.replace(
            tzinfo=ZoneInfo("UTC")
        )


    return parsed.astimezone(
        ZoneInfo("Asia/Seoul")
    )

def get_participant_ouid(
    participant_id,
    fc_nickname,
    saved_ouid,
):

    if saved_ouid:
        return saved_ouid


    ouid = get_ouid_by_nickname(
        fc_nickname
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE participants

                SET
                    ouid = %s,
                    updated_at = NOW()

                WHERE id = %s
                """,
                (
                    ouid,
                    participant_id,
                ),
            )


        connection.commit()


    return ouid


def get_db_connection():

    if not DATABASE_URL:

        raise HTTPException(
            status_code=500,
            detail="DATABASE_URL이 설정되지 않았습니다.",
        )

    return psycopg.connect(
        DATABASE_URL,
        row_factory=dict_row,
    )

# =========================
# POINTS
# =========================

def change_user_points(
    cursor,
    user_id: int,
    amount: int,
    transaction_type: str,
    reference_type: str | None = None,
    reference_id: int | None = None,
    description: str | None = None,
):

    if amount == 0:

        raise ValueError(
            "포인트 변동 금액은 "
            "0일 수 없습니다."
        )


    # =========================
    # 사용자 행 잠금
    # =========================

    cursor.execute(
        """
        SELECT
            id,
            points

        FROM users

        WHERE id = %s

        FOR UPDATE
        """,
        (
            user_id,
        ),
    )


    user = (
        cursor.fetchone()
    )


    if not user:

        raise HTTPException(
            status_code=404,
            detail=(
                "사용자를 찾을 수 없습니다."
            ),
        )


    current_balance = int(
        user["points"]
    )


    new_balance = (
        current_balance
        + amount
    )


    # =========================
    # 잔액 부족 방지
    # =========================

    if new_balance < 0:

        raise HTTPException(
            status_code=400,
            detail=(
                "보유 포인트가 부족합니다."
            ),
        )


    # =========================
    # 현재 잔액 수정
    # =========================

    cursor.execute(
        """
        UPDATE users

        SET
            points = %s,
            updated_at = NOW()

        WHERE id = %s
        """,
        (
            new_balance,
            user_id,
        ),
    )


    # =========================
    # 거래 원장 기록
    # =========================

    cursor.execute(
        """
        INSERT INTO point_transactions (
            user_id,
            amount,
            balance_after,
            transaction_type,
            reference_type,
            reference_id,
            description
        )

        VALUES (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s
        )

        RETURNING
            id,
            created_at
        """,
        (
            user_id,
            amount,
            new_balance,
            transaction_type,
            reference_type,
            reference_id,
            description,
        ),
    )


    transaction = (
        cursor.fetchone()
    )


    return {
        "transaction_id":
            transaction["id"],

        "amount":
            amount,

        "balance_before":
            current_balance,

        "balance_after":
            new_balance,

        "transaction_type":
            transaction_type,

        "reference_type":
            reference_type,

        "reference_id":
            reference_id,

        "description":
            description,

        "created_at":
            transaction["created_at"],
    }


def save_series_set_squad_players(
    series_id,
    team_a_id,
    nickname_a,
    team_b_id,
    nickname_b,
    detected_matches,
):

    spid_metadata = get_spid_metadata()

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES의 SET ID 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    set_number

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )

            series_sets = cursor.fetchall()


            set_id_map = {
                series_set["set_number"]:
                    series_set["id"]

                for series_set in series_sets
            }


            # =========================
            # 기존 Snapshot 초기화
            #
            # 재동기화 시 같은 SET의
            # 스쿼드를 중복 저장하지 않음
            # =========================

            cursor.execute(
                """
                DELETE FROM
                    series_set_squad_players

                WHERE series_set_id IN (
                    SELECT id

                    FROM series_sets

                    WHERE series_id = %s
                )
                """,
                (
                    series_id,
                ),
            )


            inserted_count = 0
            ensured_season_ids = set()


            # =========================
            # SET별 Snapshot 저장
            # =========================

            for (
                set_number,
                detected_match,
            ) in enumerate(
                detected_matches,
                start=1,
            ):

                series_set_id = (
                    set_id_map.get(
                        set_number
                    )
                )


                if series_set_id is None:

                    raise HTTPException(
                        status_code=500,
                        detail=(
                            f"{set_number}세트의 "
                            "DB 정보를 찾을 수 없습니다."
                        ),
                    )


                match_data = (
                    detected_match["data"]
                )


                participant_map = {
                    match_info["nickname"]:
                        match_info

                    for match_info
                    in match_data["matchInfo"]
                }


                squad_sides = [
                    (
                        "team_a",
                        team_a_id,
                        nickname_a,
                    ),
                    (
                        "team_b",
                        team_b_id,
                        nickname_b,
                    ),
                ]


                for (
                    side,
                    participant_id,
                    nickname,
                ) in squad_sides:

                    match_info = (
                        participant_map.get(
                            nickname
                        )
                    )


                    if match_info is None:

                        raise HTTPException(
                            status_code=500,
                            detail=(
                                f"{set_number}세트의 "
                                f"{nickname} 스쿼드를 "
                                "찾을 수 없습니다."
                            ),
                        )


                    for (
                        source_order,
                        player,
                    ) in enumerate(
                        match_info["player"]
                    ):

                        status = player["status"]

                        sp_id = player["spId"]


                        # =========================
                        # 시즌 이미지 Snapshot 보장
                        # =========================

                        season_id = (
                            int(sp_id)
                            // 1_000_000
                        )


                        if season_id not in ensured_season_ids:

                            ensure_fconline_season_snapshot(
                                sp_id
                            )

                            ensured_season_ids.add(
                                season_id
                            )


                        player_name = (
                            get_player_name(
                                sp_id,
                                spid_metadata,
                            )
                        )


                        cursor.execute(
                            """
                            INSERT INTO
                                series_set_squad_players (
                                    series_set_id,
                                    participant_id,
                                    side,
                                    source_order,

                                    sp_id,
                                    player_name,

                                    sp_position,
                                    sp_grade,

                                    rating,
                                    goals,
                                    assists,

                                    image_url
                                )

                            VALUES (
                                %s,
                                %s,
                                %s,
                                %s,

                                %s,
                                %s,

                                %s,
                                %s,

                                %s,
                                %s,
                                %s,

                                %s
                            )
                            """,
                            (
                                series_set_id,
                                participant_id,
                                side,
                                source_order,

                                sp_id,
                                player_name,

                                player["spPosition"],
                                player["spGrade"],

                                float(
                                    status["spRating"]
                                ),

                                int(
                                    status["goal"]
                                ),

                                int(
                                    status["assist"]
                                ),

                                get_player_image_url(
                                    sp_id
                                ),
                            ),
                        )


                        inserted_count += 1


        connection.commit()


    return inserted_count

def save_series_player_stats(
    series_id,
    team_a_id,
    nickname_a,
    team_b_id,
    nickname_b,
    player_stats,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            for player in player_stats:

                # =========================
                # FC 닉네임 -> 참가자 ID
                # =========================

                if (
                    player["nickname"]
                    == nickname_a
                ):

                    participant_id = (
                        team_a_id
                    )

                elif (
                    player["nickname"]
                    == nickname_b
                ):

                    participant_id = (
                        team_b_id
                    )

                else:

                    continue


                cursor.execute(
                    """
                    INSERT INTO series_player_stats (
                        series_id,
                        participant_id,

                        sp_id,
                        player_name,

                        sets_played,

                        rating_total,
                        average_rating,

                        goals,
                        assists,

                        image_url
                    )

                    VALUES (
                        %s,
                        %s,

                        %s,
                        %s,

                        %s,

                        %s,
                        %s,

                        %s,
                        %s,

                        %s
                    )

                    ON CONFLICT (
                        series_id,
                        participant_id,
                        player_name
                    )

                    DO UPDATE SET
                        sp_id =
                            EXCLUDED.sp_id,

                        sets_played =
                            EXCLUDED.sets_played,

                        rating_total =
                            EXCLUDED.rating_total,

                        average_rating =
                            EXCLUDED.average_rating,

                        goals =
                            EXCLUDED.goals,

                        assists =
                            EXCLUDED.assists,

                        image_url =
                            EXCLUDED.image_url,

                        updated_at =
                            NOW()
                    """,
                    (
                        series_id,
                        participant_id,

                        player["sp_id"],
                        player["player_name"],

                        player["sets_played"],

                        player["rating_total"],
                        player["average_rating"],

                        player["goals"],
                        player["assists"],

                        player["image_url"],
                    ),
                )


        connection.commit()

def save_series_mvp(
    series_id,
    participant_id,
    mvp,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO series_mvp (
                    series_id,
                    participant_id,

                    sp_id,
                    player_name,

                    sets_played,

                    rating_total,
                    average_rating,

                    goals,
                    assists,

                    image_url
                )

                VALUES (
                    %s,
                    %s,

                    %s,
                    %s,

                    %s,

                    %s,
                    %s,

                    %s,
                    %s,

                    %s
                )

                ON CONFLICT (
                    series_id
                )

                DO UPDATE SET

                    participant_id =
                        EXCLUDED.participant_id,

                    sp_id =
                        EXCLUDED.sp_id,

                    player_name =
                        EXCLUDED.player_name,

                    sets_played =
                        EXCLUDED.sets_played,

                    rating_total =
                        EXCLUDED.rating_total,

                    average_rating =
                        EXCLUDED.average_rating,

                    goals =
                        EXCLUDED.goals,

                    assists =
                        EXCLUDED.assists,

                    image_url =
                        EXCLUDED.image_url
                """,
                (
                    series_id,
                    participant_id,

                    mvp["sp_id"],
                    mvp["player_name"],

                    mvp["sets_played"],

                    mvp["rating_total"],
                    mvp["average_rating"],

                    mvp["goals"],
                    mvp["assists"],

                    mvp["image_url"],
                ),
            )


        connection.commit()

def initialize_database():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # USERS
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    id BIGSERIAL PRIMARY KEY,

                    email VARCHAR(255)
                        NOT NULL
                        UNIQUE,

                    password_hash TEXT
                        NOT NULL,

                    nickname VARCHAR(50)
                        NOT NULL
                        UNIQUE,

                    points INTEGER
                        NOT NULL
                        DEFAULT 0,

                    is_admin BOOLEAN
                        NOT NULL
                        DEFAULT FALSE,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    updated_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT users_points_check
                    CHECK (
                        points >= 0
                    )
                )
                """
            )

            # =========================
            # ATTENDANCE RECORDS
            # 출석체크 기록
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    attendance_records (
                        id BIGSERIAL PRIMARY KEY,

                        user_id BIGINT
                            NOT NULL
                            REFERENCES users(id)
                            ON DELETE CASCADE,

                        attendance_date DATE
                            NOT NULL,

                        streak_count INTEGER
                            NOT NULL
                            DEFAULT 1,

                        base_reward_points INTEGER
                            NOT NULL
                            DEFAULT 0,

                        streak_bonus_points INTEGER
                            NOT NULL
                            DEFAULT 0,

                        reward_points INTEGER
                            NOT NULL
                            DEFAULT 0,

                        created_at TIMESTAMPTZ
                            NOT NULL
                            DEFAULT NOW(),

                        CONSTRAINT
                            attendance_streak_count_check
                        CHECK (
                            streak_count >= 1
                        ),

                        CONSTRAINT
                            attendance_reward_points_check
                        CHECK (
                            base_reward_points >= 0
                            AND
                            streak_bonus_points >= 0
                            AND
                            reward_points >= 0
                        )
                    )
                """
            )


            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS
                    ux_attendance_user_date

                ON attendance_records (
                    user_id,
                    attendance_date
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_attendance_user_date

                ON attendance_records (
                    user_id,
                    attendance_date DESC
                )
                """
            )

            # =========================
            # POINT TRANSACTIONS
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS point_transactions (
                    id BIGSERIAL PRIMARY KEY,

                    user_id BIGINT
                        NOT NULL
                        REFERENCES users(id)
                        ON DELETE CASCADE,

                    amount INTEGER
                        NOT NULL,

                    balance_after INTEGER
                        NOT NULL,

                    transaction_type VARCHAR(50)
                        NOT NULL,

                    reference_type VARCHAR(50),

                    reference_id BIGINT,

                    description TEXT,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT
                        point_transactions_balance_check
                    CHECK (
                        balance_after >= 0
                    ),

                    CONSTRAINT
                        point_transactions_amount_check
                    CHECK (
                        amount <> 0
                    )
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_point_transactions_user_id

                ON point_transactions (
                    user_id
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_point_transactions_created_at

                ON point_transactions (
                    created_at
                )
                """
            )

            # =========================
            # POINT SHOP PRODUCTS
            # 교환소 상품
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    point_shop_products (
                        id BIGSERIAL PRIMARY KEY,

                        name VARCHAR(100)
                            NOT NULL,

                        category VARCHAR(50),

                        description TEXT,

                        price_points INTEGER
                            NOT NULL,

                        image_url TEXT,

                        is_active BOOLEAN
                            NOT NULL
                            DEFAULT TRUE,

                        sort_order INTEGER
                            NOT NULL
                            DEFAULT 0,

                        created_at TIMESTAMPTZ
                            NOT NULL
                            DEFAULT NOW(),

                        updated_at TIMESTAMPTZ
                            NOT NULL
                            DEFAULT NOW(),

                        CONSTRAINT
                            point_shop_products_price_check
                        CHECK (
                            price_points > 0
                        )
                    )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_point_shop_products_active

                ON point_shop_products (
                    is_active,
                    sort_order,
                    id
                )
                """
            )


            # =========================
            # POINT SHOP EXCHANGES
            # 상품 교환 내역
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    point_shop_exchanges (
                        id BIGSERIAL PRIMARY KEY,

                        user_id BIGINT
                            NOT NULL
                            REFERENCES users(id)
                            ON DELETE CASCADE,

                        product_id BIGINT
                            NOT NULL
                            REFERENCES point_shop_products(id),

                        product_name VARCHAR(100)
                            NOT NULL,

                        price_points INTEGER
                            NOT NULL,

                        status VARCHAR(20)
                            NOT NULL
                            DEFAULT 'requested',

                        created_at TIMESTAMPTZ
                            NOT NULL
                            DEFAULT NOW(),

                        completed_at TIMESTAMPTZ,

                        CONSTRAINT
                            point_shop_exchanges_price_check
                        CHECK (
                            price_points > 0
                        ),

                        CONSTRAINT
                            point_shop_exchanges_status_check
                        CHECK (
                            status IN (
                                'requested',
                                'completed',
                                'cancelled'
                            )
                        )
                    )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_point_shop_exchanges_user

                ON point_shop_exchanges (
                    user_id,
                    created_at DESC,
                    id DESC
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_point_shop_exchanges_status

                ON point_shop_exchanges (
                    status,
                    created_at ASC
                )
                """
            )

            # =========================
            # USER SESSIONS
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS user_sessions (
                    id BIGSERIAL PRIMARY KEY,

                    user_id BIGINT
                        NOT NULL
                        REFERENCES users(id)
                        ON DELETE CASCADE,

                    token_hash VARCHAR(64)
                        NOT NULL
                        UNIQUE,

                    expires_at TIMESTAMPTZ
                        NOT NULL,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW()
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_user_sessions_user_id

                ON user_sessions (
                    user_id
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_user_sessions_expires_at

                ON user_sessions (
                    expires_at
                )
                """
            )

            # =========================
            # 참가자
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS participants (
                    id BIGSERIAL PRIMARY KEY,

                    fcl_name VARCHAR(50)
                        NOT NULL
                        UNIQUE,

                    fc_nickname VARCHAR(100)
                        UNIQUE,

                    ouid VARCHAR(100)
                        UNIQUE,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    updated_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW()
                )
                """
            )

            # =========================
            # SERIES
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS series (
                    id BIGSERIAL PRIMARY KEY,

                    series_type VARCHAR(20)
                        NOT NULL,

                    team_a_id BIGINT
                        NOT NULL
                        REFERENCES participants(id),

                    team_b_id BIGINT
                        NOT NULL
                        REFERENCES participants(id),

                    match_type INTEGER
                        NOT NULL
                        DEFAULT 40,

                    scheduled_date DATE,

                    round_number INTEGER,

                    fixture_number INTEGER,

                    playoff_stage VARCHAR(20),

                    best_of INTEGER,

                    wins_required INTEGER,

                    started_at TIMESTAMPTZ,

                    completed_at TIMESTAMPTZ,

                    cancelled_at TIMESTAMPTZ,

                    finished_at TIMESTAMPTZ,

                    stats_sync_status VARCHAR(20)
                        NOT NULL
                        DEFAULT 'pending',

                    status VARCHAR(20)
                        NOT NULL
                        DEFAULT 'scheduled',

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT series_check
                    CHECK (
                        team_a_id <> team_b_id
                    ),

                    CONSTRAINT series_series_type_check
                    CHECK (
                        series_type IN (
                            '프리시즌',
                            '정규리그',
                            '플레이오프'
                        )
                    ),

                    CONSTRAINT series_status_check
                    CHECK (
                        status IN (
                            'scheduled',
                            'active',
                            'completed',
                            'cancelled'
                        )
                    ),

                    CONSTRAINT chk_series_stats_sync_status
                    CHECK (
                        stats_sync_status IN (
                            'pending',
                            'synced',
                            'conflict'
                        )
                    ),

                    CONSTRAINT chk_series_playoff_config
                    CHECK (
                        (
                            series_type IN (
                                '프리시즌',
                                '정규리그'
                            )

                            AND playoff_stage IS NULL
                            AND best_of IS NULL
                            AND wins_required IS NULL
                        )

                        OR

                        (
                            series_type = '플레이오프'

                            AND playoff_stage IS NOT NULL
                            AND best_of IS NOT NULL
                            AND wins_required IS NOT NULL

                            AND (
                                (
                                    playoff_stage IN (
                                        '준플레이오프',
                                        '플레이오프'
                                    )

                                    AND best_of = 5
                                    AND wins_required = 3
                                )

                                OR

                                (
                                    playoff_stage = '결승시리즈'

                                    AND best_of = 7
                                    AND wins_required = 4
                                )
                            )
                        )
                    )
                )
                """
            )



            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS
                    ux_series_regular_fixture_number

                ON series (
                    fixture_number
                )

                WHERE
                    series_type = '정규리그'
                """
            )


            # =========================
            # SERIES 세트
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS series_sets (
                    id BIGSERIAL PRIMARY KEY,

                    series_id BIGINT
                        NOT NULL
                        REFERENCES series(id)
                        ON DELETE CASCADE,

                    set_number INTEGER
                        NOT NULL,

                    nexon_match_id VARCHAR(100)
                        UNIQUE,

                    played_at TIMESTAMPTZ
                        NOT NULL,

                    team_a_score INTEGER
                        NOT NULL,

                    team_b_score INTEGER
                        NOT NULL,

                    score_source VARCHAR(20)
                        NOT NULL
                        DEFAULT 'nexon',

                    winner_side VARCHAR(10),

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    UNIQUE (
                        series_id,
                        set_number
                    ),

                    CONSTRAINT series_sets_set_number_check
                    CHECK (
                        set_number
                        BETWEEN 1 AND 7
                    ),

                    CONSTRAINT series_sets_team_a_score_check
                    CHECK (
                        team_a_score >= 0
                    ),

                    CONSTRAINT series_sets_team_b_score_check
                    CHECK (
                        team_b_score >= 0
                    ),

                    CONSTRAINT chk_series_sets_score_source
                    CHECK (
                        score_source IN (
                            'nexon',
                            'manual'
                        )
                    ),

                    CONSTRAINT chk_series_sets_winner_side
                    CHECK (
                        winner_side IS NULL

                        OR winner_side IN (
                            'team_a',
                            'team_b',
                            'draw'
                        )
                    )
                )
                """
            )

            # =========================
            # SERIES 세트별 스쿼드 Snapshot
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS series_set_squad_players (
                    id BIGSERIAL PRIMARY KEY,

                    series_set_id BIGINT
                        NOT NULL
                        REFERENCES series_sets(id)
                        ON DELETE CASCADE,

                    participant_id BIGINT
                        NOT NULL
                        REFERENCES participants(id),

                    side VARCHAR(10)
                        NOT NULL,

                    source_order INTEGER
                        NOT NULL,

                    sp_id BIGINT
                        NOT NULL,

                    player_name VARCHAR(100)
                        NOT NULL,

                    sp_position INTEGER
                        NOT NULL,

                    sp_grade INTEGER
                        NOT NULL,

                    rating NUMERIC(5, 2)
                        NOT NULL,

                    goals INTEGER
                        NOT NULL
                        DEFAULT 0,

                    assists INTEGER
                        NOT NULL
                        DEFAULT 0,

                    image_url TEXT,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT chk_series_set_squad_side
                    CHECK (
                        side IN (
                            'team_a',
                            'team_b'
                        )
                    ),

                    CONSTRAINT chk_series_set_squad_source_order
                    CHECK (
                        source_order >= 0
                    ),

                    CONSTRAINT chk_series_set_squad_rating
                    CHECK (
                        rating >= 0
                    ),

                    CONSTRAINT chk_series_set_squad_goals
                    CHECK (
                        goals >= 0
                    ),

                    CONSTRAINT chk_series_set_squad_assists
                    CHECK (
                        assists >= 0
                    ),

                    UNIQUE (
                        series_set_id,
                        side,
                        source_order
                    )
                )
                """
            )


            # =========================
            # SERIES MVP
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS series_mvp (
                    series_id BIGINT
                        PRIMARY KEY
                        REFERENCES series(id)
                        ON DELETE CASCADE,

                    participant_id BIGINT
                        NOT NULL
                        REFERENCES participants(id),

                    sp_id BIGINT
                        NOT NULL,

                    player_name VARCHAR(100)
                        NOT NULL,

                    sets_played INTEGER
                        NOT NULL,

                    rating_total NUMERIC(6, 2)
                        NOT NULL,

                    average_rating NUMERIC(5, 2)
                        NOT NULL,

                    goals INTEGER
                        NOT NULL,

                    assists INTEGER
                        NOT NULL,

                    image_url TEXT,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW()
                )
                """
            )

            # =========================
            # SERIES 선수 기록
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS series_player_stats (
                    id BIGSERIAL PRIMARY KEY,

                    series_id BIGINT
                        NOT NULL
                        REFERENCES series(id)
                        ON DELETE CASCADE,

                    participant_id BIGINT
                        NOT NULL
                        REFERENCES participants(id),

                    sp_id BIGINT
                        NOT NULL,

                    player_name VARCHAR(100)
                        NOT NULL,

                    sets_played INTEGER
                        NOT NULL,

                    rating_total NUMERIC(6, 2)
                        NOT NULL,

                    average_rating NUMERIC(5, 2)
                        NOT NULL,

                    goals INTEGER
                        NOT NULL,

                    assists INTEGER
                        NOT NULL,

                    image_url TEXT,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    updated_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    UNIQUE (
                        series_id,
                        participant_id,
                        player_name
                    )
                )
                """
            )

            # =========================
            # COMMUNITY NOTICE
            # =========================

            cursor.execute(
                """
                ALTER TABLE community_posts

                ADD COLUMN IF NOT EXISTS
                    is_notice BOOLEAN
                    NOT NULL
                    DEFAULT FALSE
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_community_posts_notice

                ON community_posts (
                    is_notice,
                    created_at DESC
                )
                """
            )

        connection.commit()


    seed_participants()

def seed_participants():

    participant_data = [
        (
            "문권기",
            "공기",
        ),
        (
            "이준석",
            "똭똭",
        ),
        (
            "주은성",
            "펜쉬차일드",
        ),
        (
            "이상",
            "지수사",
        ),
        (
            "서종원",
            "붉은심장베컴",
        ),
    ]


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            for (
                fcl_name,
                fc_nickname,
            ) in participant_data:

                cursor.execute(
                    """
                    INSERT INTO participants (
                        fcl_name,
                        fc_nickname
                    )

                    VALUES (
                        %s,
                        %s
                    )

                    ON CONFLICT (
                        fcl_name
                    )

                    DO UPDATE SET

                        fc_nickname =
                            COALESCE(
                                EXCLUDED.fc_nickname,
                                participants.fc_nickname
                            ),

                        updated_at = NOW()
                    """,
                    (
                        fcl_name,
                        fc_nickname,
                    ),
                )


        connection.commit()

def get_nexon_headers():

    if not NEXON_API_KEY:
        raise HTTPException(
            status_code=500,
            detail="NEXON_API_KEY가 설정되지 않았습니다.",
        )

    return {
        "x-nxopen-api-key": NEXON_API_KEY,
    }

def get_ouid_by_nickname(nickname):

    response = httpx.get(
        f"{NEXON_API_BASE_URL}/id",
        headers=get_nexon_headers(),
        params={
            "nickname": nickname,
        },
        timeout=10.0,
    )

    if response.status_code != 200:
        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )

    return response.json()["ouid"]


def get_user_match_ids(
    ouid,
    match_type,
    limit=50,
):

    response = httpx.get(
        f"{NEXON_API_BASE_URL}/user/match",
        headers=get_nexon_headers(),
        params={
            "ouid": ouid,
            "matchtype": match_type,
            "offset": 0,
            "limit": limit,
        },
        timeout=10.0,
    )

    if response.status_code != 200:
        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )

    return response.json()

# =========================
# FC Online 경기 상세 조회
# =========================

def get_match_detail(
    match_id,
):

    response = httpx.get(
        f"{NEXON_API_BASE_URL}/match-detail",
        headers=get_nexon_headers(),
        params={
            "matchid": match_id,
        },
        timeout=10.0,
    )


    if response.status_code != 200:

        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )


    return response.json()

def get_match_winner_side(
    match_data,
    nickname_a,
    nickname_b,
    series_type=None,
):

    participant_map = {
        match_info["nickname"]:
            match_info

        for match_info
        in match_data.get(
            "matchInfo",
            [],
        )
    }


    team_a_info = participant_map.get(
        nickname_a
    )


    team_b_info = participant_map.get(
        nickname_b
    )


    if (
        team_a_info is None
        or
        team_b_info is None
    ):

        return None


    # =========================
    # 실제 경기 점수
    #
    # FCL 경기 스코어 기준은
    # NEXON goalTotal 사용
    # =========================

    team_a_score = int(
        team_a_info
        .get(
            "shoot",
            {},
        )
        .get(
            "goalTotal",
            0,
        )
        or 0
    )


    team_b_score = int(
        team_b_info
        .get(
            "shoot",
            {},
        )
        .get(
            "goalTotal",
            0,
        )
        or 0
    )


    # =========================
    # 점수로 승자가 명확한 경우
    # =========================

    if (
        team_a_score
        >
        team_b_score
    ):

        return "team_a"


    if (
        team_b_score
        >
        team_a_score
    ):

        return "team_b"


    # =========================
    # 프리시즌 / 정규리그
    #
    # 스코어가 같으면 무승부
    # =========================

    if (
        series_type
        != "플레이오프"
    ):

        return "draw"


    # =========================
    # 플레이오프 동점
    #
    # 연장 / 승부차기 결과는
    # NEXON matchResult 사용
    # =========================

    team_a_result = (
        team_a_info
        .get(
            "matchDetail",
            {},
        )
        .get(
            "matchResult"
        )
    )


    team_b_result = (
        team_b_info
        .get(
            "matchDetail",
            {},
        )
        .get(
            "matchResult"
        )
    )


    if (
        team_a_result == "승"
        and
        team_b_result == "패"
    ):

        return "team_a"


    if (
        team_a_result == "패"
        and
        team_b_result == "승"
    ):

        return "team_b"


    # 플레이오프인데
    # 동점 스코어 + 승자 판별 실패

    return None

# =========================
# FC Online 경기 데이터 정합성 검증
# =========================

def get_match_integrity_conflict(
    match_data,
    nickname_a,
    nickname_b,
    series_type=None,
):

    participant_map = {
        match_info["nickname"]:
            match_info

        for match_info
        in match_data.get(
            "matchInfo",
            [],
        )
    }


    team_a_info = (
        participant_map.get(
            nickname_a
        )
    )


    team_b_info = (
        participant_map.get(
            nickname_b
        )
    )


    # =========================
    # 참가자 정합성
    # =========================

    if (
        team_a_info is None
        or
        team_b_info is None
    ):

        return (
            "NEXON 경기 데이터에서 "
            "SERIES 참가자를 찾을 수 없습니다."
        )


    # =========================
    # 팀 점수 / 선수 득점 정합성
    # =========================

    for (
        nickname,
        match_info,
    ) in (
        (
            nickname_a,
            team_a_info,
        ),

        (
            nickname_b,
            team_b_info,
        ),
    ):

        shoot = (
            match_info.get(
                "shoot",
                {},
            )
        )


        goal_total = int(
            shoot.get(
                "goalTotal",
                0,
            )
            or 0
        )


        own_goal = int(
            shoot.get(
                "ownGoal",
                0,
            )
            or 0
        )


        player_goal_sum = sum(
            int(
                player
                .get(
                    "status",
                    {},
                )
                .get(
                    "goal",
                    0,
                )
                or 0
            )

            for player
            in match_info.get(
                "player",
                [],
            )
        )


        # =========================
        # 자책골이 없는 경우
        #
        # 선수 득점 합계와
        # 실제 팀 득점이 같아야 함
        # =========================

        if (
            own_goal == 0
            and
            player_goal_sum
            !=
            goal_total
        ):

            return (
                "NEXON 선수 득점 합계와 "
                "팀 득점이 일치하지 않습니다. "
                f"({nickname}: "
                f"playerGoals="
                f"{player_goal_sum}, "
                f"goalTotal={goal_total})"
            )


    # =========================
    # 플레이오프 동점 경기 검증
    # =========================

    team_a_score = int(
        team_a_info
        .get(
            "shoot",
            {},
        )
        .get(
            "goalTotal",
            0,
        )
        or 0
    )


    team_b_score = int(
        team_b_info
        .get(
            "shoot",
            {},
        )
        .get(
            "goalTotal",
            0,
        )
        or 0
    )


    if (
        series_type == "플레이오프"
        and
        team_a_score == team_b_score
    ):

        winner_side = (
            get_match_winner_side(
                match_data,
                nickname_a,
                nickname_b,
                series_type,
            )
        )


        if (
            winner_side
            not in (
                "team_a",
                "team_b",
            )
        ):

            team_a_result = (
                team_a_info
                .get(
                    "matchDetail",
                    {},
                )
                .get(
                    "matchResult"
                )
            )


            team_b_result = (
                team_b_info
                .get(
                    "matchDetail",
                    {},
                )
                .get(
                    "matchResult"
                )
            )


            return (
                "플레이오프 동점 경기의 "
                "승자를 확인할 수 없습니다. "
                f"({nickname_a}: {team_a_result}, "
                f"{nickname_b}: {team_b_result})"
            )


    return None

# =========================
# FCL SERIES MVP 계산
# =========================

def calculate_series_mvp_from_matches(
    matches,
):

    spid_metadata = get_spid_metadata()

    player_totals = {}


    for match_data in matches:

        for match_info in match_data["matchInfo"]:

            nickname = match_info["nickname"]


            fcl_name = next(
                (
                    name
                    for name, fc_nickname
                    in FCONLINE_NICKNAMES.items()
                    if fc_nickname == nickname
                ),
                nickname,
            )


            for player in match_info["player"]:

                status = player["status"]

                rating = float(
                    status["spRating"]
                )


                # 출전하지 않은 선수 제외
                if rating <= 0:
                    continue


                sp_id = player["spId"]

                player_name = get_player_name(
                    sp_id,
                    spid_metadata,
                )


                # 같은 실제 선수는 시즌 카드가 달라도
                # 한 선수로 합산
                player_key = (
                    nickname,
                    player_name,
                )


                if player_key not in player_totals:

                    player_totals[player_key] = {
                        "nickname": nickname,
                        "fcl_name": fcl_name,

                        "player_name": player_name,
                        "sp_id": sp_id,

                        "sets_played": 0,

                        "rating_total": 0,
                        "ratings": [],

                        "goals": 0,
                        "assists": 0,

                        "best_single_rating":
                            rating,
                    }


                record = player_totals[
                    player_key
                ]


                # 다른 시즌 카드를 사용했다면
                # 가장 높은 평점을 받은 카드 이미지 사용
                if (
                    rating
                    >
                    record[
                        "best_single_rating"
                    ]
                ):

                    record[
                        "best_single_rating"
                    ] = rating

                    record["sp_id"] = sp_id


                record["sets_played"] += 1

                record["rating_total"] += (
                    rating
                )

                record["ratings"].append(
                    rating
                )

                record["goals"] += int(
                    status["goal"]
                )

                record["assists"] += int(
                    status["assist"]
                )


    # =========================
    # 전체 선수 기록
    # =========================

    all_player_stats = []


    for record in player_totals.values():

        record["rating_total"] = round(
            record["rating_total"],
            2,
        )


        record["average_rating"] = round(
            record["rating_total"]
            / record["sets_played"],
            2,
        )


        record["image_url"] = (
            get_player_image_url(
                record["sp_id"]
            )
        )


        record.pop(
            "best_single_rating",
            None,
        )


        all_player_stats.append(
            record
        )


    # =========================
    # MVP 후보
    # 최소 2세트 출전
    # =========================

    mvp_rankings = [
        player
        for player in all_player_stats
        if player["sets_played"] >= 2
    ]


    mvp_rankings.sort(
        key=lambda player: (
            -player["rating_total"],
            -player["goals"],
            -player["assists"],
            -player["average_rating"],
        )
    )


    if not mvp_rankings:

        return (
            None,
            [],
            all_player_stats,
        )


    return (
        mvp_rankings[0],
        mvp_rankings,
        all_player_stats,
    )

# =========================
# FCL 3세트 MVP 테스트
# =========================

@app.get("/api/fconline/series-mvp-test")
def get_series_mvp_test():

    match_ids = [
        "6a81cdffa962a502d85e1eaa",
        "6a81cb89bdd2ff3b3f6807a8",
        "6a81c8f6115d94f2dc8ca43c",
    ]


    matches = [
        get_match_detail(match_id)
        for match_id in match_ids
    ]


    matches.sort(
        key=lambda match:
            match["matchDate"]
    )


    (
        mvp,
        ranking,
        player_stats,
    ) = calculate_series_mvp_from_matches(
        matches
    )


    return {
        "mvp": mvp,
        "ranking": ranking,
        "player_stats": player_stats,
    }


# =========================
# 참가자
# =========================

PARTICIPANTS = [
    "문권기",
    "이준석",
    "주은성",
    "이상",
    "서종원",
]

FCONLINE_NICKNAMES = {
    "문권기": "공기",
    "이준석": "똭똭",
    "주은성": "펜쉬차일드",
    "이상": "지수사",
    "서종원": "붉은심장베컴",
}

# =========================
# PREDICTIONS
# 승부예측 운영 설정
# =========================

PREDICTION_FIXED_ODDS = 2.50
PREDICTION_MAX_STAKE_POINTS = 1000

# =========================
# ATTENDANCE EVENT
# 출석체크 운영 설정
# =========================

ATTENDANCE_DAILY_REWARD = 50
ATTENDANCE_STREAK_DAYS = 7
ATTENDANCE_STREAK_BONUS = 150

# =========================
# PREDICTIONS
# 세트별 승부예측 자동 정산
# =========================

def settle_predictions_for_series(
    cursor,
    series_id: int,
):

    # =========================
    # SERIES 확인
    # =========================

    cursor.execute(
        """
        SELECT
            id,
            series_type,
            team_a_id,
            team_b_id,
            status

        FROM series

        WHERE id = %s

        FOR UPDATE
        """,
        (
            series_id,
        ),
    )


    series = cursor.fetchone()


    if not series:

        raise RuntimeError(
            "승부예측 정산 대상 SERIES를 "
            "찾을 수 없습니다."
        )


    # 정규리그만 승부예측 정산
    if (
        series["series_type"]
        != "정규리그"
    ):

        return {
            "settled": 0,
            "wins": 0,
            "losses": 0,
            "payout_points": 0,
        }


    # 완료된 경기만 정산
    if (
        series["status"]
        != "completed"
    ):

        return {
            "settled": 0,
            "wins": 0,
            "losses": 0,
            "payout_points": 0,
        }


    # =========================
    # 실제 1 / 2 / 3세트 결과
    # =========================

    cursor.execute(
        """
        SELECT
            set_number,
            team_a_score,
            team_b_score

        FROM series_sets

        WHERE
            series_id = %s

            AND set_number
                BETWEEN 1 AND 3

        ORDER BY
            set_number

        FOR UPDATE
        """,
        (
            series_id,
        ),
    )


    set_rows = cursor.fetchall()


    set_map = {
        int(
            row["set_number"]
        ):
            row

        for row
        in set_rows
    }


    if (
        set(
            set_map.keys()
        )
        !=
        {
            1,
            2,
            3,
        }
    ):

        raise RuntimeError(
            "승부예측 정산에 필요한 "
            "1~3세트 결과가 모두 존재하지 않습니다."
        )


    # =========================
    # 아직 정산되지 않은 예측
    #
    # pending만 가져오므로
    # 같은 경기 sync를 다시 눌러도
    # 중복 지급되지 않음
    # =========================

    cursor.execute(
        """
        SELECT
            id,
            user_id,
            set_number,
            prediction_type,
            predicted_participant_id,
            stake_points,
            odds

        FROM predictions

        WHERE
            series_id = %s

            AND status =
                'pending'

        ORDER BY
            id

        FOR UPDATE
        """,
        (
            series_id,
        ),
    )


    predictions = (
        cursor.fetchall()
    )


    settled_count = 0
    win_count = 0
    loss_count = 0
    total_payout_points = 0


    for prediction in predictions:

        set_number = int(
            prediction[
                "set_number"
            ]
        )


        set_result = (
            set_map[
                set_number
            ]
        )


        team_a_score = int(
            set_result[
                "team_a_score"
            ]
        )

        team_b_score = int(
            set_result[
                "team_b_score"
            ]
        )


        # =========================
        # 실제 결과 판단
        # =========================

        if (
            team_a_score
            ==
            team_b_score
        ):

            winning_type = (
                "draw"
            )

            winning_participant_id = (
                None
            )


        elif (
            team_a_score
            >
            team_b_score
        ):

            winning_type = (
                "participant"
            )

            winning_participant_id = int(
                series[
                    "team_a_id"
                ]
            )


        else:

            winning_type = (
                "participant"
            )

            winning_participant_id = int(
                series[
                    "team_b_id"
                ]
            )


        # =========================
        # 예측 적중 여부
        # =========================

        is_win = False


        if (
            prediction[
                "prediction_type"
            ]
            ==
            "draw"

            and

            winning_type
            ==
            "draw"
        ):

            is_win = True


        elif (
            prediction[
                "prediction_type"
            ]
            ==
            "participant"

            and

            winning_type
            ==
            "participant"

            and

            int(
                prediction[
                    "predicted_participant_id"
                ]
            )
            ==
            winning_participant_id
        ):

            is_win = True


        # =========================
        # 적중
        # =========================

        if is_win:

            payout_points = int(
                prediction[
                    "stake_points"
                ]
                *
                prediction[
                    "odds"
                ]
            )


            change_user_points(
                cursor,

                prediction[
                    "user_id"
                ],

                payout_points,

                "prediction_win",

                reference_type=
                    "prediction",

                reference_id=
                    prediction[
                        "id"
                    ],

                description=(
                    "FCL 승부예측 "
                    f"{set_number}세트 적중"
                ),
            )


            prediction_status = (
                "win"
            )

            win_count += 1

            total_payout_points += (
                payout_points
            )


        # =========================
        # 실패
        # =========================

        else:

            payout_points = 0

            prediction_status = (
                "loss"
            )

            loss_count += 1


        # =========================
        # 예측 정산 완료
        # =========================

        cursor.execute(
            """
            UPDATE predictions

            SET
                status = %s,
                payout_points = %s,
                settled_at = NOW(),
                updated_at = NOW()

            WHERE
                id = %s
                AND status = 'pending'
            """,
            (
                prediction_status,
                payout_points,
                prediction[
                    "id"
                ],
            ),
        )


        settled_count += 1


    return {
        "settled":
            settled_count,

        "wins":
            win_count,

        "losses":
            loss_count,

        "payout_points":
            total_payout_points,
    }

class PointShopExchangeRequest(
    BaseModel
):
    product_id: int

class PredictionCreateRequest(
    BaseModel
): 
    series_id: int
    set_number: int
    prediction_type: str
    participant_id: int | None = None
    stake_points: int

class PredictionSettlementTestRequest(
    BaseModel
):
    series_id: int

    set1_team_a: int
    set1_team_b: int

    set2_team_a: int
    set2_team_b: int

    set3_team_a: int
    set3_team_b: int

class UserSignupRequest(
    BaseModel
):
    email: str
    password: str
    nickname: str

class UserLoginRequest(
    BaseModel
):
    email: str
    password: str

class AdminUserPointRequest(
    BaseModel
):
    amount: int
    description: str | None = None


class AdminUserRoleRequest(
    BaseModel
):
    is_admin: bool

class AdminPointShopProductCreateRequest(
    BaseModel
):
    name: str
    category: str | None = None
    description: str | None = None
    price_points: int
    image_url: str | None = None
    is_active: bool = True
    sort_order: int = 0


class AdminPointShopProductUpdateRequest(
    BaseModel
):
    name: str
    category: str | None = None
    description: str | None = None
    price_points: int
    image_url: str | None = None
    is_active: bool
    sort_order: int

class AdminCommunityPostUpdateRequest(
    BaseModel
):
    title: str
    content: str

class AdminCommunityNoticeRequest(
    BaseModel
):
    title: str
    content: str

class CommunityCommentCreateRequest(
    BaseModel
):
    content: str

class CommunityCommentUpdateRequest(
    BaseModel
):
    content: str

class AdminLoginRequest(BaseModel):
    password: str

class AdminParticipantTeamUpdateRequest(
    BaseModel
):
    current_team_name: str
    current_team_logo_path: str

class SeriesStartRequest(BaseModel):
    team_a: str
    team_b: str

    series_type: str = "프리시즌"

    scheduled_date: str | None = None

class ManualSeriesCompleteRequest(BaseModel):

    set1_team_a: int
    set1_team_b: int
    set1_winner_side: str | None = None

    set2_team_a: int
    set2_team_b: int
    set2_winner_side: str | None = None

    set3_team_a: int
    set3_team_b: int
    set3_winner_side: str | None = None

    set4_team_a: int | None = None
    set4_team_b: int | None = None
    set4_winner_side: str | None = None

    set5_team_a: int | None = None
    set5_team_b: int | None = None
    set5_winner_side: str | None = None

    set6_team_a: int | None = None
    set6_team_b: int | None = None
    set6_winner_side: str | None = None

    set7_team_a: int | None = None
    set7_team_b: int | None = None
    set7_winner_side: str | None = None

class AdminSeriesResultUpdateRequest(
    BaseModel
):
    set1_team_a: int
    set1_team_b: int
    set1_winner_side: str | None = None

    set2_team_a: int
    set2_team_b: int
    set2_winner_side: str | None = None

    set3_team_a: int
    set3_team_b: int
    set3_winner_side: str | None = None

    set4_team_a: int | None = None
    set4_team_b: int | None = None
    set4_winner_side: str | None = None

    set5_team_a: int | None = None
    set5_team_b: int | None = None
    set5_winner_side: str | None = None

    set6_team_a: int | None = None
    set6_team_b: int | None = None
    set6_winner_side: str | None = None

    set7_team_a: int | None = None
    set7_team_b: int | None = None
    set7_winner_side: str | None = None

class HistorySeriesImportRequest(BaseModel):
    team_a: str
    team_b: str
    match_date: str

class PlayoffInitializeRequest(
    BaseModel
):
    scheduled_date: str

class PlayoffAdvanceRequest(
    BaseModel
):
    scheduled_date: str

class AdminRegularScheduleUpdateRequest(
    BaseModel
):
    scheduled_date: str

# =========================
# USER AUTH
# =========================

PASSWORD_HASH_ITERATIONS = (
    600_000
)

USER_SESSION_SECONDS = (
    30 * 24 * 60 * 60
)

SIGNUP_BONUS_POINTS = 1000

user_bearer_scheme = HTTPBearer(
    auto_error=False
)


def hash_user_password(
    password: str,
):

    salt = secrets.token_bytes(
        16
    )


    password_hash = (
        hashlib.pbkdf2_hmac(
            "sha256",
            password.encode(
                "utf-8"
            ),
            salt,
            PASSWORD_HASH_ITERATIONS,
        )
    )


    return (
        "pbkdf2_sha256"
        f"${PASSWORD_HASH_ITERATIONS}"
        f"${salt.hex()}"
        f"${password_hash.hex()}"
    )


def verify_user_password(
    password: str,
    stored_password_hash: str,
):

    try:

        (
            algorithm,
            iterations_text,
            salt_hex,
            password_hash_hex,
        ) = stored_password_hash.split(
            "$",
            3,
        )


        if (
            algorithm
            != "pbkdf2_sha256"
        ):

            return False


        iterations = int(
            iterations_text
        )


        salt = bytes.fromhex(
            salt_hex
        )


        expected_hash = (
            bytes.fromhex(
                password_hash_hex
            )
        )


        calculated_hash = (
            hashlib.pbkdf2_hmac(
                "sha256",
                password.encode(
                    "utf-8"
                ),
                salt,
                iterations,
            )
        )


        return hmac.compare_digest(
            calculated_hash,
            expected_hash,
        )


    except (
        ValueError,
        TypeError,
    ):

        return False

def hash_user_session_token(
    token: str,
):

    return hashlib.sha256(
        token.encode(
            "utf-8"
        )
    ).hexdigest()


def create_user_session(
    user_id: int,
):

    token = secrets.token_urlsafe(
        48
    )


    token_hash = (
        hash_user_session_token(
            token
        )
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # 만료된 세션 정리
            cursor.execute(
                """
                DELETE FROM user_sessions

                WHERE expires_at <= NOW()
                """
            )


            cursor.execute(
                """
                INSERT INTO user_sessions (
                    user_id,
                    token_hash,
                    expires_at
                )

                VALUES (
                    %s,
                    %s,
                    NOW()
                    + (
                        %s
                        * INTERVAL '1 second'
                    )
                )

                RETURNING
                    expires_at
                """,
                (
                    user_id,
                    token_hash,
                    USER_SESSION_SECONDS,
                ),
            )


            session = (
                cursor.fetchone()
            )


        connection.commit()


    return (
        token,
        session["expires_at"],
    )


def get_user_bearer_token(
    authorization: str | None,
):

    if not authorization:

        raise HTTPException(
            status_code=401,
            detail=(
                "로그인이 필요합니다."
            ),
        )


    parts = authorization.split(
        " ",
        1,
    )


    if (
        len(parts) != 2
        or
        parts[0].lower()
        != "bearer"
        or
        not parts[1].strip()
    ):

        raise HTTPException(
            status_code=401,
            detail=(
                "로그인 정보가 "
                "올바르지 않습니다."
            ),
        )


    return parts[1].strip()

def require_user(
    credentials:
        HTTPAuthorizationCredentials
        | None
        = Depends(
            user_bearer_scheme
        )
):

    if not credentials:

        raise HTTPException(
            status_code=401,
            detail="로그인이 필요합니다.",
        )


    token = (
        credentials.credentials
    )


    token_hash = (
        hash_user_session_token(
            token
        )
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    users.id,
                    users.email,
                    users.nickname,
                    users.points,
                    users.is_admin,
                    users.created_at

                FROM user_sessions

                JOIN users
                    ON users.id
                    = user_sessions.user_id

                WHERE
                    user_sessions.token_hash
                    = %s

                    AND
                    user_sessions.expires_at
                    > NOW()
                """,
                (
                    token_hash,
                ),
            )


            user = (
                cursor.fetchone()
            )


    if not user:

        raise HTTPException(
            status_code=401,
            detail=(
                "로그인이 만료되었거나 "
                "유효하지 않습니다."
            ),
        )


    return user

def require_user_admin(
    user = Depends(
        require_user
    )
):

    if not user["is_admin"]:

        raise HTTPException(
            status_code=403,
            detail=(
                "관리자 권한이 필요합니다."
            ),
        )

    return user


@app.post(
    "/api/auth/signup"
)
def signup_user(
    request: UserSignupRequest
):

    email = (
        request.email
        .strip()
        .lower()
    )


    nickname = (
        request.nickname
        .strip()
    )


    password = (
        request.password
    )


    # =========================
    # 이메일 검증
    # =========================

    if (
        not email
        or
        "@" not in email
        or
        len(email) > 255
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "올바른 이메일을 "
                "입력해주세요."
            ),
        )


    # =========================
    # 닉네임 검증
    # =========================

    if (
        len(nickname) < 2
        or
        len(nickname) > 20
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "닉네임은 2자 이상 "
                "20자 이하로 입력해주세요."
            ),
        )


    # =========================
    # 비밀번호 검증
    # =========================

    if (
        len(password) < 8
        or
        len(password) > 128
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "비밀번호는 8자 이상 "
                "128자 이하로 입력해주세요."
            ),
        )


    # =========================
    # 중복 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT id

                FROM users

                WHERE email = %s
                """,
                (
                    email,
                ),
            )


            if cursor.fetchone():

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "이미 사용 중인 "
                        "이메일입니다."
                    ),
                )


            cursor.execute(
                """
                SELECT id

                FROM users

                WHERE nickname = %s
                """,
                (
                    nickname,
                ),
            )


            if cursor.fetchone():

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "이미 사용 중인 "
                        "닉네임입니다."
                    ),
                )


            # =========================
            # 비밀번호 해시
            # =========================

            password_hash = (
                hash_user_password(
                    password
                )
            )


            # =========================
            # 회원 생성
            # =========================

            cursor.execute(
                """
                INSERT INTO users (
                    email,
                    password_hash,
                    nickname
                )

                VALUES (
                    %s,
                    %s,
                    %s
                )

                RETURNING
                    id,
                    email,
                    nickname,
                    points,
                    is_admin,
                    created_at
                """,
                (
                    email,
                    password_hash,
                    nickname,
                ),
            )


            user = (
                cursor.fetchone()
            )

            point_transaction = (
                change_user_points(
                    cursor,
                    user["id"],
                    SIGNUP_BONUS_POINTS,
                    "signup_bonus",
                    description=(
                        "FCL 신규 회원 가입 보너스"
                    ),
                )
            )


            user["points"] = (
                point_transaction[
                    "balance_after"
                ]
            )


        connection.commit()


    return {
        "message":
            "회원가입이 완료되었습니다.",

        "user":
            user,
    }

@app.post(
    "/api/auth/login"
)
def login_user(
    request: UserLoginRequest
):

    email = (
        request.email
        .strip()
        .lower()
    )


    password = (
        request.password
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    email,
                    password_hash,
                    nickname,
                    points,
                    is_admin,
                    created_at

                FROM users

                WHERE email = %s
                """,
                (
                    email,
                ),
            )


            user = (
                cursor.fetchone()
            )


    if (
        not user
        or
        not verify_user_password(
            password,
            user["password_hash"],
        )
    ):

        raise HTTPException(
            status_code=401,
            detail=(
                "이메일 또는 비밀번호가 "
                "올바르지 않습니다."
            ),
        )


    (
        token,
        expires_at,
    ) = create_user_session(
        user["id"]
    )


    return {
        "message":
            "로그인되었습니다.",

        "token_type":
            "Bearer",

        "token":
            token,

        "expires_in":
            USER_SESSION_SECONDS,

        "expires_at":
            expires_at,

        "user": {
            "id":
                user["id"],

            "email":
                user["email"],

            "nickname":
                user["nickname"],

            "points":
                user["points"],

            "is_admin":
                user["is_admin"],

            "created_at":
                user["created_at"],
        },
    }

@app.get(
    "/api/auth/me"
)
def get_current_user(
    user = Depends(
        require_user
    )
):

    return {
        "user":
            user,
    }


@app.post(
    "/api/auth/logout"
)
def logout_user(
    credentials:
        HTTPAuthorizationCredentials
        = Depends(
            user_bearer_scheme
        ),

    user = Depends(
        require_user
    ),
):

    if not credentials:

        raise HTTPException(
            status_code=401,
            detail="로그인이 필요합니다.",
        )


    token = (
        credentials.credentials
    )


    token_hash = (
        hash_user_session_token(
            token
        )
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                DELETE FROM user_sessions

                WHERE token_hash = %s
                """,
                (
                    token_hash,
                ),
            )


        connection.commit()


    return {
        "message":
            "로그아웃되었습니다."
    }

# =========================
# MYPAGE
# =========================

@app.get(
    "/api/mypage"
)
def get_mypage(
    user = Depends(
        require_user
    )
):

    return {
        "user": {
            "id":
                user["id"],

            "email":
                user["email"],

            "nickname":
                user["nickname"],

            "points":
                user["points"],

            "is_admin":
                user["is_admin"],

            "created_at":
                user["created_at"],
        }
    }


@app.get(
    "/api/mypage/point-transactions"
)
def get_mypage_point_transactions(
    limit: int = 50,

    user = Depends(
        require_user
    ),
):

    # =========================
    # 조회 개수 제한
    # =========================

    limit = max(
        1,
        min(
            limit,
            100,
        ),
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    amount,
                    balance_after,
                    transaction_type,
                    reference_type,
                    reference_id,
                    description,
                    created_at

                FROM point_transactions

                WHERE user_id = %s

                ORDER BY
                    created_at DESC,
                    id DESC

                LIMIT %s
                """,
                (
                    user["id"],
                    limit,
                ),
            )


            transactions = (
                cursor.fetchall()
            )


    return {
        "current_points":
            user["points"],

        "transactions":
            transactions,
    }


# =========================
# ADMIN AUTH
# =========================

def create_admin_session():

    token = secrets.token_urlsafe(
        32
    )


    ADMIN_SESSIONS[token] = (
        time.time()
        + ADMIN_SESSION_SECONDS
    )


    return token


def require_admin(
    x_admin_token: str | None =
        Header(default=None)
):

    if not x_admin_token:

        raise HTTPException(
            status_code=401,
            detail="관리자 로그인이 필요합니다.",
        )


    expires_at = (
        ADMIN_SESSIONS.get(
            x_admin_token
        )
    )


    if (
        expires_at is None
        or
        expires_at < time.time()
    ):

        ADMIN_SESSIONS.pop(
            x_admin_token,
            None,
        )


        raise HTTPException(
            status_code=401,
            detail="관리자 로그인이 만료되었습니다.",
        )


    return x_admin_token


@app.post(
    "/api/admin/login"
)
def admin_login(
    request: AdminLoginRequest
):

    if not ADMIN_PASSWORD:

        raise HTTPException(
            status_code=500,
            detail=(
                "ADMIN_PASSWORD가 "
                "설정되지 않았습니다."
            ),
        )


    if not secrets.compare_digest(
        request.password,
        ADMIN_PASSWORD,
    ):

        raise HTTPException(
            status_code=401,
            detail="비밀번호가 올바르지 않습니다.",
        )


    token = create_admin_session()


    return {
        "admin": True,
        "token": token,
        "expires_in":
            ADMIN_SESSION_SECONDS,
    }

# =========================
# ADMIN USERS
# =========================

@app.get(
    "/api/admin/users"
)
def get_admin_users(
    admin_token = Depends(
        require_admin
    )
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    email,
                    nickname,
                    points,
                    is_admin,
                    created_at,
                    updated_at

                FROM users

                ORDER BY
                    created_at DESC,
                    id DESC
                """
            )


            users = (
                cursor.fetchall()
            )


    return {
        "users":
            users
    }

@app.post(
    "/api/admin/users/{user_id}/points"
)
def change_admin_user_points(
    user_id: int,

    request:
        AdminUserPointRequest,

    admin_token = Depends(
        require_admin
    ),
):

    amount = (
        request.amount
    )


    if amount == 0:

        raise HTTPException(
            status_code=400,
            detail=(
                "포인트 변동 금액은 "
                "0일 수 없습니다."
            ),
        )


    description = (
        request.description.strip()
        if request.description
        else None
    )


    if (
        description
        and
        len(description) > 200
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "포인트 사유는 "
                "200자 이하로 입력해주세요."
            ),
        )


    transaction_type = (
        "admin_grant"
        if amount > 0
        else "admin_deduct"
    )


    if not description:

        description = (
            "관리자 포인트 지급"
            if amount > 0
            else "관리자 포인트 차감"
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            transaction = (
                change_user_points(
                    cursor,
                    user_id,
                    amount,
                    transaction_type,
                    description=
                        description,
                )
            )


        connection.commit()


    return {
        "message":
            "포인트가 변경되었습니다.",

        "transaction":
            transaction,
    }

@app.patch(
    "/api/admin/users/{user_id}/role"
)
def change_admin_user_role(
    user_id: int,

    request:
        AdminUserRoleRequest,

    admin_token = Depends(
        require_admin
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE users

                SET
                    is_admin = %s,
                    updated_at = NOW()

                WHERE id = %s

                RETURNING
                    id,
                    email,
                    nickname,
                    points,
                    is_admin,
                    created_at
                """,
                (
                    request.is_admin,
                    user_id,
                ),
            )


            user = (
                cursor.fetchone()
            )


            if not user:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "사용자를 찾을 수 없습니다."
                    ),
                )


        connection.commit()


    return {
        "message":
            (
                "관리자 권한이 부여되었습니다."
                if request.is_admin
                else
                "관리자 권한이 해제되었습니다."
            ),

        "user":
            user,
    }


# =========================
# ADMIN PARTICIPANTS
# 참가자 / 현재 팀 조회
# =========================

@app.get(
    "/api/admin/participants"
)
def admin_get_participants(
    admin_token: str =
        Depends(
            require_admin
        )
):
    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname,

                    current_team_name,
                    current_team_logo_path

                FROM participants

                ORDER BY id
                """
            )

            participants = (
                cursor.fetchall()
            )

    return participants

# =========================
# ADMIN PARTICIPANT TEAM UPDATE
# 현재 팀 변경
# =========================

@app.put(
    "/api/admin/participants/"
    "{participant_id}/team"
)
def admin_update_participant_team(
    participant_id: int,
    request:
        AdminParticipantTeamUpdateRequest,
    admin_token: str =
        Depends(
            require_admin
        ),
):
    current_team_name = (
        request.current_team_name.strip()
    )

    current_team_logo_path = (
        request
            .current_team_logo_path
            .strip()
    )


    # =========================
    # 기본 검증
    # =========================

    if not current_team_name:

        raise HTTPException(
            status_code=400,
            detail=(
                "현재 팀 이름을 "
                "입력해주세요."
            ),
        )


    if not current_team_logo_path:

        raise HTTPException(
            status_code=400,
            detail=(
                "현재 팀 로고 경로를 "
                "입력해주세요."
            ),
        )


    # =========================
    # 역사 보존용 로고만 허용
    # =========================

    history_directory = (
        FRONTEND_DIR
        / "assets"
        / "images"
        / "teams"
        / "history"
    ).resolve()


    logo_relative_path = Path(
        current_team_logo_path.removeprefix(
            "./"
        )
    )


    logo_file_path = (
        FRONTEND_DIR
        / logo_relative_path
    ).resolve()


    if (
        history_directory
        not in logo_file_path.parents
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "현재 팀 로고는 "
                "history 폴더의 파일만 "
                "사용할 수 있습니다."
            ),
        )


    if not logo_file_path.is_file():

        raise HTTPException(
            status_code=400,
            detail=(
                "지정한 팀 로고 파일을 "
                "찾을 수 없습니다."
            ),
        )


    # =========================
    # 참가자 현재 팀 변경
    #
    # series Snapshot은
    # 절대 수정하지 않음
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 참가자 존재 확인 + 잠금
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name

                FROM participants

                WHERE id = %s

                FOR UPDATE
                """,
                (
                    participant_id,
                ),
            )


            existing_participant = (
                cursor.fetchone()
            )


            if not existing_participant:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "참가자를 "
                        "찾을 수 없습니다."
                    ),
                )


            # =========================
            # 진행 중 SERIES 보호
            #
            # 경기 진행 중에는
            # 현재 팀 변경 금지
            #
            # 시작 순간 저장된 Snapshot과
            # participants 현재 팀 정보가
            # 서로 달라지는 상황 방지
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    playoff_stage

                FROM series

                WHERE
                    status = 'active'

                    AND (
                        team_a_id = %s
                        OR
                        team_b_id = %s
                    )

                LIMIT 1

                FOR UPDATE
                """,
                (
                    participant_id,
                    participant_id,
                ),
            )


            active_series = (
                cursor.fetchone()
            )


            if active_series:

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "진행 중인 경기가 있는 "
                        "참가자의 팀은 "
                        "변경할 수 없습니다. "
                        "경기 종료 후 다시 시도해주세요."
                    ),
                )


            # =========================
            # 참가자 현재 팀 변경
            #
            # 과거 SERIES Snapshot은
            # 절대 수정하지 않음
            # =========================

            cursor.execute(
                """
                UPDATE participants

                SET
                    current_team_name = %s,
                    current_team_logo_path = %s,
                    updated_at = NOW()

                WHERE id = %s

                RETURNING
                    id,
                    fcl_name,
                    fc_nickname,
                    current_team_name,
                    current_team_logo_path
                """,
                (
                    current_team_name,
                    current_team_logo_path,
                    participant_id,
                ),
            )


            participant = (
                cursor.fetchone()
            )


        connection.commit()


    return participant

# =========================
# ADMIN TEAM LOGO INTEGRITY
# 현재 / 과거 팀 로고 파일 무결성 검사
# =========================

@app.get(
    "/api/admin/team-logo-integrity"
)
def admin_check_team_logo_integrity(
    admin_token: str =
        Depends(
            require_admin
        )
):
    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 현재 참가자 팀 로고
            # =========================

            cursor.execute(
                """
                SELECT DISTINCT
                    current_team_logo_path
                        AS logo_path

                FROM participants

                WHERE
                    current_team_logo_path
                    IS NOT NULL
                """
            )

            current_logo_rows = (
                cursor.fetchall()
            )


            # =========================
            # 과거 SERIES Snapshot 로고
            # =========================

            cursor.execute(
                """
                SELECT DISTINCT
                    logo_path

                FROM (
                    SELECT
                        team_a_snapshot_logo_path
                            AS logo_path

                    FROM series

                    UNION

                    SELECT
                        team_b_snapshot_logo_path
                            AS logo_path

                    FROM series
                ) AS snapshot_logos

                WHERE
                    logo_path IS NOT NULL
                """
            )

            snapshot_logo_rows = (
                cursor.fetchall()
            )


    referenced_logo_paths = set()


    for row in current_logo_rows:

        referenced_logo_paths.add(
            row["logo_path"]
        )


    for row in snapshot_logo_rows:

        referenced_logo_paths.add(
            row["logo_path"]
        )


    history_directory = (
        FRONTEND_DIR
        / "assets"
        / "images"
        / "teams"
        / "history"
    ).resolve()


    missing_logo_paths = []
    invalid_logo_paths = []


    for logo_path in sorted(
        referenced_logo_paths
    ):

        logo_relative_path = Path(
            logo_path.removeprefix("./")
        )


        logo_file_path = (
            FRONTEND_DIR
            / logo_relative_path
        ).resolve()


        # =========================
        # history 밖의 잘못된 경로
        # =========================

        if (
            history_directory
            not in logo_file_path.parents
        ):

            invalid_logo_paths.append(
                logo_path
            )

            continue


        # =========================
        # 실제 파일 존재 여부
        # =========================

        if not logo_file_path.is_file():

            missing_logo_paths.append(
                logo_path
            )


    return {
        "ok":
            (
                len(missing_logo_paths) == 0
                and
                len(invalid_logo_paths) == 0
            ),

        "referenced_count":
            len(
                referenced_logo_paths
            ),

        "missing_count":
            len(
                missing_logo_paths
            ),

        "invalid_count":
            len(
                invalid_logo_paths
            ),

        "missing_logo_paths":
            missing_logo_paths,

        "invalid_logo_paths":
            invalid_logo_paths,
    }


@app.get(
    "/api/admin/check"
)
def admin_check(
    admin_token: str =
        Depends(
            require_admin
        )
):

    return {
        "admin": True
    }

def get_round_number(match_index):
    return (match_index // 5) + 1


# =========================
# ADMIN SERIES SQUAD BACKFILL
# 기존 완료 경기 세트별 스쿼드 Snapshot 생성
# =========================

@app.post(
    "/api/admin/series/{series_id}/squad-backfill"
)
def admin_backfill_series_squad(
    series_id: int,
    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # SERIES + 참가자 조회
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.status,

                    s.team_a_id,
                    team_a.fc_nickname
                        AS nickname_a,

                    s.team_b_id,
                    team_b.fc_nickname
                        AS nickname_b

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s
                """,
                (
                    series_id,
                ),
            )

            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            if series["status"] != "completed":

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "완료된 경기만 "
                        "스쿼드 Snapshot을 "
                        "생성할 수 있습니다."
                    ),
                )


            cursor.execute(
                """
                SELECT
                    id,
                    set_number,
                    nexon_match_id,
                    played_at

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )

            saved_sets = cursor.fetchall()


    # =========================
    # 세트 확인
    # =========================

    if not saved_sets:

        raise HTTPException(
            status_code=400,
            detail=(
                "저장된 세트가 없습니다."
            ),
        )


    nickname_a = series[
        "nickname_a"
    ]

    nickname_b = series[
        "nickname_b"
    ]


    if not nickname_a or not nickname_b:

        raise HTTPException(
            status_code=400,
            detail=(
                "FC Online 닉네임 정보가 "
                "없습니다."
            ),
        )


    # =========================
    # 저장된 matchId 기준
    # Nexon 원본 다시 조회
    # =========================

    detected_matches = []


    for saved_set in saved_sets:

        match_id = saved_set[
            "nexon_match_id"
        ]


        if not match_id:

            raise HTTPException(
                status_code=400,
                detail=(
                    f"{saved_set['set_number']}세트에 "
                    "Nexon matchId가 없습니다."
                ),
            )


        match_data = get_match_detail(
            match_id
        )


        match_nicknames = {
            match_info["nickname"]

            for match_info
            in match_data["matchInfo"]
        }


        if match_nicknames != {
            nickname_a,
            nickname_b,
        }:

            raise HTTPException(
                status_code=400,
                detail=(
                    f"{saved_set['set_number']}세트의 "
                    "Nexon 참가자 정보가 "
                    "SERIES와 일치하지 않습니다."
                ),
            )


        detected_matches.append(
            {
                "data":
                    match_data,

                "played_at":
                    parse_nexon_datetime(
                        match_data[
                            "matchDate"
                        ]
                    ),
            }
        )


    # =========================
    # Snapshot 저장
    # =========================

    inserted_count = (
        save_series_set_squad_players(
            series_id,

            series[
                "team_a_id"
            ],
            nickname_a,

            series[
                "team_b_id"
            ],
            nickname_b,

            detected_matches,
        )
    )


    return {
        "series_id":
            series_id,

        "set_count":
            len(saved_sets),

        "player_snapshot_count":
            inserted_count,

        "message":
            (
                "세트별 스쿼드 Snapshot "
                "생성이 완료되었습니다."
            ),
    }

# =========================
# ADMIN REGULAR SCHEDULE
# 정규리그 일정 조회
# =========================

@app.get(
    "/api/admin/regular-schedule"
)
def admin_get_regular_schedule(
    admin_token: str =
        Depends(
            require_admin
        )
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id AS series_id,
                    s.fixture_number,
                    s.round_number,
                    s.scheduled_date,
                    s.status,

                    team_a.fcl_name
                        AS team_a,

                    team_b.fcl_name
                        AS team_b

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE
                    s.series_type =
                        '정규리그'

                ORDER BY
                    s.fixture_number,
                    s.id
                """
            )


            schedule_rows = cursor.fetchall()


    schedules = []


    for schedule_row in schedule_rows:

        scheduled_date = schedule_row [
                "scheduled_date"
            ]


        schedules.append(
            {
                "series_id":
                    schedule_row[
                        "series_id"
                    ],

                "fixture_number":
                    schedule_row[
                        "fixture_number"
                    ],

                "round":
                    schedule_row[
                        "round_number"
                    ],

                "date":
                    (
                        scheduled_date.isoformat()
                        if scheduled_date
                        else None
                    ),

                "team_a":
                    schedule_row[
                        "team_a"
                    ],

                "team_b":
                    schedule_row[
                        "team_b"
                    ],

                "status":
                    schedule_row[
                        "status"
                    ],
            }
        )


    return schedules

# =========================
# ADMIN REGULAR SCHEDULE UPDATE
# 정규리그 일정 변경
# =========================

@app.put(
    "/api/admin/series/{series_id}/schedule"
)
def admin_update_regular_schedule(
    series_id: int,
    request:
        AdminRegularScheduleUpdateRequest,
    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # 날짜 형식 확인
    # =========================

    try:

        new_scheduled_date = (
            datetime.strptime(
                request.scheduled_date,
                "%Y-%m-%d",
            ).date()
        )

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail=(
                "경기 날짜 형식이 "
                "올바르지 않습니다."
            ),
        )


    # =========================
    # 과거 날짜 금지
    # =========================

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    if new_scheduled_date < today:

        raise HTTPException(
            status_code=400,
            detail=(
                "지난 날짜로는 "
                "일정을 변경할 수 없습니다."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    scheduled_date,
                    status

                FROM series

                WHERE id = %s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "경기를 찾을 수 없습니다."
                    ),
                )


            # =========================
            # 정규리그만 허용
            # =========================

            if (
                series["series_type"]
                != "정규리그"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "정규리그 경기만 "
                        "일정을 변경할 수 있습니다."
                    ),
                )


            # =========================
            # 예정 경기만 허용
            # =========================

            if (
                series["status"]
                != "scheduled"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "예정 상태의 경기만 "
                        "일정을 변경할 수 있습니다."
                    ),
                )


            previous_date = (
                series["scheduled_date"]
            )


            # =========================
            # 날짜 변경
            # =========================

            cursor.execute(
                """
                UPDATE series

                SET
                    scheduled_date = %s

                WHERE id = %s

                RETURNING
                    id,
                    fixture_number,
                    round_number,
                    scheduled_date,
                    status
                """,
                (
                    new_scheduled_date,
                    series_id,
                ),
            )


            updated_series = cursor.fetchone()


        connection.commit()


    return {
        "series_id":
            updated_series["id"],

        "fixture_number":
            updated_series[
                "fixture_number"
            ],

        "round":
            updated_series[
                "round_number"
            ],

        "previous_date":
            (
                previous_date.isoformat()
                if previous_date
                else None
            ),

        "scheduled_date":
            updated_series[
                "scheduled_date"
            ].isoformat(),

        "status":
            updated_series[
                "status"
            ],

        "message":
            "정규리그 일정이 변경되었습니다.",
    }



# =========================
# ADMIN SERIES DELETE
# =========================

@app.delete(
    "/api/admin/series/{series_id}"
)
def admin_delete_series(
    series_id: int,
    admin_token: str =
        Depends(
            require_admin
        )
):

    removed_downstream_series = []

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    status,
                    scheduled_date
                FROM series
                WHERE id = %s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            if (
                series["status"]
                != "completed"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "완료된 경기만 "
                        "삭제할 수 있습니다."
                    ),
                )


            # =========================
            # 프리시즌
            #
            # SERIES 자체 삭제
            # 자식 데이터는 CASCADE 삭제
            # =========================

            if (
                series["series_type"]
                == "프리시즌"
            ):

                cursor.execute(
                    """
                    DELETE FROM series
                    WHERE id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                result_action = (
                    "deleted"
                )


            # =========================
            # 정규리그
            #
            # 일정은 유지하고
            # 경기 결과만 초기화
            # =========================

            elif (
                series["series_type"]
                == "정규리그"
            ):

                cursor.execute(
                    """
                    DELETE FROM series_sets
                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    DELETE FROM series_mvp
                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    DELETE FROM series_player_stats
                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    UPDATE series

                    SET
                        status = 'scheduled',
                        started_at = NULL,
                        completed_at = NULL,
                        finished_at = NULL,

                        stats_sync_status = 'pending',

                        team_a_snapshot_name = NULL,
                        team_a_snapshot_logo_path = NULL,

                        team_b_snapshot_name = NULL,
                        team_b_snapshot_logo_path = NULL

                    WHERE id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                result_action = (
                    "reset"
                )

            # =========================
            # 플레이오프
            #
            # 현재 결과는 scheduled로 복구
            # 이후 자동 생성된 단계는 제거
            #
            # 단, 이후 단계가 이미
            # active/completed면 삭제 금지
            # =========================

            elif (
                series["series_type"]
                == "플레이오프"
            ):

                # =========================
                # 현재 단계 확인
                # =========================

                cursor.execute(
                    """
                    SELECT
                        playoff_stage

                    FROM series

                    WHERE id = %s

                    FOR UPDATE
                    """,
                    (
                        series_id,
                    ),
                )


                playoff_series = (
                    cursor.fetchone()
                )


                playoff_stage = (
                    playoff_series[
                        "playoff_stage"
                    ]
                )


                if playoff_stage not in (
                    "준플레이오프",
                    "플레이오프",
                    "결승시리즈",
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "플레이오프 단계를 "
                            "확인할 수 없습니다."
                        ),
                    )


                # =========================
                # 이후 단계 목록
                # =========================

                if (
                    playoff_stage
                    == "준플레이오프"
                ):

                    downstream_stages = [
                        "플레이오프",
                        "결승시리즈",
                    ]

                elif (
                    playoff_stage
                    == "플레이오프"
                ):

                    downstream_stages = [
                        "결승시리즈",
                    ]

                else:

                    downstream_stages = []


                downstream_series = []


                # =========================
                # 이후 단계 상태 확인
                # =========================

                if downstream_stages:

                    cursor.execute(
                        """
                        SELECT
                            id,
                            playoff_stage,
                            status

                        FROM series

                        WHERE
                            series_type =
                                '플레이오프'

                            AND
                            playoff_stage =
                                ANY(%s)

                            AND
                            status <>
                                'cancelled'

                        ORDER BY id

                        FOR UPDATE
                        """,
                        (
                            downstream_stages,
                        ),
                    )


                    downstream_series = (
                        cursor.fetchall()
                    )


                    for downstream in (
                        downstream_series
                    ):

                        if (
                            downstream["status"]
                            != "scheduled"
                        ):

                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    f"{downstream['playoff_stage']}가 "
                                    "이미 시작되었거나 완료되어 "
                                    "이전 단계 결과를 "
                                    "삭제할 수 없습니다."
                                ),
                            )


                # =========================
                # 이후 자동 생성 SERIES 제거
                #
                # scheduled 상태만 여기까지
                # 통과했으므로 제거 가능
                # =========================

                removed_downstream_series = []


                for downstream in reversed(
                    downstream_series
                ):

                    cursor.execute(
                        """
                        DELETE FROM series

                        WHERE id = %s
                        """,
                        (
                            downstream["id"],
                        ),
                    )


                    removed_downstream_series.append(
                        {
                            "series_id":
                                downstream["id"],

                            "playoff_stage":
                                downstream[
                                    "playoff_stage"
                                ],
                        }
                    )


                # =========================
                # 현재 플레이오프 결과 제거
                # =========================

                cursor.execute(
                    """
                    DELETE FROM series_sets

                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    DELETE FROM series_mvp

                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    DELETE FROM series_player_stats

                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                # =========================
                # 현재 SERIES는 일정 유지
                # scheduled 상태로 복구
                # =========================

                cursor.execute(
                    """
                    UPDATE series

                    SET
                        status = 'scheduled',

                        started_at = NULL,
                        completed_at = NULL,
                        finished_at = NULL,

                        stats_sync_status =
                            'pending',

                        team_a_snapshot_name =
                            NULL,

                        team_a_snapshot_logo_path =
                            NULL,

                        team_b_snapshot_name =
                            NULL,

                        team_b_snapshot_logo_path =
                            NULL

                    WHERE id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                result_action = (
                    "playoff_reset"
                )


            else:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "삭제할 수 없는 "
                        "경기 종류입니다."
                    ),
                )


        connection.commit()


    return {
        "series_id":
            series_id,

        "series_type":
            series["series_type"],

        "action":
            result_action,

        "removed_downstream_series":
            removed_downstream_series,

        "message":
            (
                "경기 결과가 "
                "삭제되었습니다."
            ),
    }

# =========================
# ADMIN SERIES RESULT UPDATE
# =========================

@app.put(
    "/api/admin/series/{series_id}/result"
)
def admin_update_series_result(
    series_id: int,
    request:
        AdminSeriesResultUpdateRequest,
    admin_token: str =
        Depends(
            require_admin
        ),
):

    edited_at = datetime.now(
        ZoneInfo("Asia/Seoul")
    )


    requested_sets = [
        (
            1,
            request.set1_team_a,
            request.set1_team_b,
            request.set1_winner_side,
        ),
        (
            2,
            request.set2_team_a,
            request.set2_team_b,
            request.set2_winner_side,
        ),
        (
            3,
            request.set3_team_a,
            request.set3_team_b,
            request.set3_winner_side,
        ),
        (
            4,
            request.set4_team_a,
            request.set4_team_b,
            request.set4_winner_side,
        ),
        (
            5,
            request.set5_team_a,
            request.set5_team_b,
            request.set5_winner_side,
        ),
        (
            6,
            request.set6_team_a,
            request.set6_team_b,
            request.set6_winner_side,
        ),
        (
            7,
            request.set7_team_a,
            request.set7_team_b,
            request.set7_winner_side,
        ),
    ]


    bracket_updated = False
    next_stage = None
    next_series_missing = False


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    playoff_stage,
                    best_of,
                    wins_required,
                    status,

                    team_a_id,
                    team_b_id,

                    completed_at,
                    finished_at

                FROM series

                WHERE id = %s

                FOR UPDATE
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            if (
                series["status"]
                != "completed"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "완료된 경기 결과만 "
                        "수정할 수 있습니다."
                    ),
                )


            # =========================
            # 경기 종류별 세트 수
            # =========================

            is_playoff = (
                series["series_type"]
                == "플레이오프"
            )


            if is_playoff:

                if (
                    series["best_of"] is None
                    or
                    series["wins_required"] is None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "플레이오프 진행 정보가 "
                            "올바르지 않습니다."
                        ),
                    )


                max_sets = int(
                    series["best_of"]
                )

                wins_required = int(
                    series["wins_required"]
                )

            else:

                max_sets = 3
                wins_required = None


            # =========================
            # 허용 세트 초과 방어
            # =========================

            for (
                set_number,
                team_a_score,
                team_b_score,
                explicit_winner_side,
            ) in requested_sets[max_sets:]:

                if (
                    team_a_score is not None
                    or team_b_score is not None
                    or explicit_winner_side is not None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{max_sets}세트를 초과하여 "
                            "결과를 입력할 수 없습니다."
                        ),
                    )


            # =========================
            # 기존 세트
            #
            # 기존 played_at 보존
            # 기존 플레이오프 승자 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    set_number,
                    played_at,
                    winner_side

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )


            saved_sets = cursor.fetchall()


            if not saved_sets:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "저장된 경기 결과가 없습니다."
                    ),
                )


            saved_played_at = {
                saved_set["set_number"]:
                    saved_set["played_at"]

                for saved_set in saved_sets
            }


            current_winner_side = None


            if is_playoff:

                current_team_a_wins = 0
                current_team_b_wins = 0


                for saved_set in saved_sets:

                    if current_winner_side is not None:

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "기존 플레이오프 결과에 "
                                "선승 도달 이후 세트가 "
                                "존재합니다."
                            ),
                        )


                    winner_side = (
                        saved_set[
                            "winner_side"
                        ]
                    )


                    if winner_side == "team_a":

                        current_team_a_wins += 1

                    elif winner_side == "team_b":

                        current_team_b_wins += 1

                    else:

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "기존 플레이오프의 "
                                "세트 승패 정보가 "
                                "올바르지 않습니다."
                            ),
                        )


                    if (
                        current_team_a_wins
                        >= wins_required
                    ):

                        current_winner_side = (
                            "team_a"
                        )

                    elif (
                        current_team_b_wins
                        >= wins_required
                    ):

                        current_winner_side = (
                            "team_b"
                        )


                if current_winner_side is None:

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "기존 플레이오프의 "
                            "SERIES 승자를 "
                            "확인할 수 없습니다."
                        ),
                    )


            # =========================
            # 수정 결과 검증
            # =========================

            updated_sets = []

            gap_found = False

            team_a_wins = 0
            team_b_wins = 0

            series_winner_side = None
            winning_set_number = None


            for (
                set_number,
                team_a_score,
                team_b_score,
                explicit_winner_side,
            ) in requested_sets[:max_sets]:

                # =========================
                # 둘 다 비어 있음
                # =========================

                if (
                    team_a_score is None
                    and team_b_score is None
                ):

                    if (
                        explicit_winner_side
                        is not None
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 점수 없이 "
                                "승자만 지정할 수 없습니다."
                            ),
                        )


                    gap_found = True
                    continue


                # =========================
                # 한쪽 점수만 입력
                # =========================

                if (
                    team_a_score is None
                    or team_b_score is None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{set_number}세트의 "
                            "양쪽 점수를 모두 입력해주세요."
                        ),
                    )


                # =========================
                # 중간 세트 공백
                # =========================

                if gap_found:

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "중간 세트를 비워둔 채 "
                            "다음 세트를 입력할 수 없습니다."
                        ),
                    )


                # =========================
                # 음수 방어
                # =========================

                if (
                    team_a_score < 0
                    or team_b_score < 0
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "점수는 0 이상의 "
                            "정수여야 합니다."
                        ),
                    )


                # =========================
                # 세트 승자
                # =========================

                if team_a_score > team_b_score:

                    winner_side = "team_a"


                    if (
                        explicit_winner_side
                        is not None
                        and
                        explicit_winner_side
                        != winner_side
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 승자와 "
                                "입력 점수가 일치하지 않습니다."
                            ),
                        )


                elif team_b_score > team_a_score:

                    winner_side = "team_b"


                    if (
                        explicit_winner_side
                        is not None
                        and
                        explicit_winner_side
                        != winner_side
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 승자와 "
                                "입력 점수가 일치하지 않습니다."
                            ),
                        )


                else:

                    if not is_playoff:

                        winner_side = "draw"

                    else:

                        if (
                            explicit_winner_side
                            not in (
                                "team_a",
                                "team_b",
                            )
                        ):

                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    f"{set_number}세트가 동점입니다. "
                                    "플레이오프에서는 실제 승자를 "
                                    "지정해야 합니다."
                                ),
                            )


                        winner_side = (
                            explicit_winner_side
                        )


                # =========================
                # 선승 이후 추가 세트 방어
                # =========================

                if (
                    is_playoff
                    and
                    series_winner_side
                    is not None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "선승 도달 이후의 "
                            "추가 세트가 입력되었습니다."
                        ),
                    )


                updated_sets.append(
                    (
                        set_number,
                        team_a_score,
                        team_b_score,
                        winner_side,
                    )
                )


                # =========================
                # 플레이오프 승수
                # =========================

                if is_playoff:

                    if winner_side == "team_a":

                        team_a_wins += 1

                    elif winner_side == "team_b":

                        team_b_wins += 1


                    if (
                        team_a_wins
                        >= wins_required
                    ):

                        series_winner_side = (
                            "team_a"
                        )

                        winning_set_number = (
                            set_number
                        )

                    elif (
                        team_b_wins
                        >= wins_required
                    ):

                        series_winner_side = (
                            "team_b"
                        )

                        winning_set_number = (
                            set_number
                        )


            # =========================
            # 일반 경기
            # 정확히 3세트
            # =========================

            if not is_playoff:

                if len(updated_sets) != 3:

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "프리시즌과 정규리그는 "
                            "3세트를 모두 입력해야 합니다."
                        ),
                    )


            # =========================
            # 플레이오프
            # 선승 도달 필수
            # =========================

            else:

                if series_winner_side is None:

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{wins_required}승에 도달한 "
                            "참가자가 없습니다."
                        ),
                    )


            # =========================
            # 플레이오프 승자 변경 시
            # 다음 단계 대진 보호
            # =========================

            if (
                is_playoff
                and
                series[
                    "playoff_stage"
                ]
                in (
                    "준플레이오프",
                    "플레이오프",
                )
            ):

                if (
                    series["playoff_stage"]
                    == "준플레이오프"
                ):

                    next_stage = "플레이오프"

                else:

                    next_stage = "결승시리즈"


                cursor.execute(
                    """
                    SELECT
                        id,
                        status,
                        team_a_id,
                        team_b_id

                    FROM series

                    WHERE
                        series_type =
                            '플레이오프'

                        AND
                        playoff_stage = %s

                        AND
                        status <>
                            'cancelled'

                    ORDER BY id DESC

                    LIMIT 1

                    FOR UPDATE
                    """,
                    (
                        next_stage,
                    ),
                )


                next_series = (
                    cursor.fetchone()
                )


                if next_series:

                    current_winner_id = (
                        series["team_a_id"]
                        if
                        current_winner_side
                        == "team_a"
                        else
                        series["team_b_id"]
                    )


                    new_winner_id = (
                        series["team_a_id"]
                        if
                        series_winner_side
                        == "team_a"
                        else
                        series["team_b_id"]
                    )


                    # =========================
                    # 승자가 바뀌는 경우만
                    # 다음 단계 영향
                    # =========================

                    if (
                        new_winner_id
                        != current_winner_id
                    ):

                        if (
                            next_series["status"]
                            != "scheduled"
                        ):

                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    f"{next_stage}가 이미 "
                                    "시작되었거나 완료되어 "
                                    "이전 단계 승자를 "
                                    "변경할 수 없습니다."
                                ),
                            )


                        if (
                            next_series["team_a_id"]
                            == new_winner_id
                        ):

                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    "수정된 승자와 "
                                    "다음 단계 시드 참가자가 "
                                    "같습니다."
                                ),
                            )


                        cursor.execute(
                            """
                            SELECT COUNT(*) AS count

                            FROM series_sets

                            WHERE series_id = %s
                            """,
                            (
                                next_series["id"],
                            ),
                        )


                        next_set_count = int(
                            cursor.fetchone()[
                                "count"
                            ]
                        )


                        if next_set_count > 0:

                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    "다음 단계에 이미 "
                                    "세트 결과가 존재하여 "
                                    "승자를 변경할 수 없습니다."
                                ),
                            )


                        # 다음 단계의 team_b는
                        # 이전 단계 승자
                        cursor.execute(
                            """
                            UPDATE series

                            SET
                                team_b_id = %s,

                                team_b_snapshot_name =
                                    NULL,

                                team_b_snapshot_logo_path =
                                    NULL

                            WHERE id = %s
                            """,
                            (
                                new_winner_id,
                                next_series["id"],
                            ),
                        )


                        bracket_updated = True

                else:

                    next_series_missing = True


            # =========================
            # 기존 결과 제거
            # =========================

            cursor.execute(
                """
                DELETE FROM series_sets
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            cursor.execute(
                """
                DELETE FROM series_mvp
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            cursor.execute(
                """
                DELETE FROM series_player_stats
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            # =========================
            # 수정 세트 저장
            # =========================

            for (
                set_number,
                team_a_score,
                team_b_score,
                winner_side,
            ) in updated_sets:

                played_at = (
                    saved_played_at.get(
                        set_number
                    )
                    or
                    series["finished_at"]
                    or
                    series["completed_at"]
                    or
                    edited_at
                )


                cursor.execute(
                    """
                    INSERT INTO series_sets (
                        series_id,
                        set_number,
                        nexon_match_id,
                        played_at,
                        team_a_score,
                        team_b_score,
                        score_source,
                        winner_side
                    )

                    VALUES (
                        %s,
                        %s,
                        NULL,
                        %s,
                        %s,
                        %s,
                        'manual',
                        %s
                    )
                    """,
                    (
                        series_id,
                        set_number,
                        played_at,
                        team_a_score,
                        team_b_score,
                        winner_side,
                    ),
                )


            # =========================
            # 관리자 수정 후
            # NEXON 통계 재동기화 대기
            # =========================

            cursor.execute(
                """
                UPDATE series

                SET
                    stats_sync_status =
                        'pending'

                WHERE id = %s
                """,
                (
                    series_id,
                ),
            )


        connection.commit()


    # =========================
    # 다음 단계가 없는 경우
    # 자동 진행 복구 시도
    # =========================

    progression_warning = None


    if (
        is_playoff
        and
        next_series_missing
    ):

        try:

            create_next_playoff_if_ready(
                series_id
            )

        except Exception as error:

            print(
                "[PLAYOFF ADMIN EDIT "
                "ADVANCE ERROR]",
                repr(error),
            )

            progression_warning = (
                "결과 수정은 완료됐지만 "
                "다음 플레이오프 생성에 "
                "실패했습니다."
            )


    return {
        "series_id":
            series_id,

        "status":
            "completed",

        "stats_sync_status":
            "pending",

        "series_type":
            series["series_type"],

        "playoff_stage":
            series["playoff_stage"],

        "best_of":
            series["best_of"],

        "wins_required":
            series["wins_required"],

        "team_a_wins":
            (
                team_a_wins
                if is_playoff
                else None
            ),

        "team_b_wins":
            (
                team_b_wins
                if is_playoff
                else None
            ),

        "winner_side":
            (
                series_winner_side
                if is_playoff
                else None
            ),

        "winning_set":
            (
                winning_set_number
                if is_playoff
                else None
            ),

        "bracket_updated":
            bracket_updated,

        "progression_warning":
            progression_warning,

        "sets": [
            {
                "set":
                    set_number,

                "team_a_score":
                    team_a_score,

                "team_b_score":
                    team_b_score,

                "winner_side":
                    winner_side,
            }

            for (
                set_number,
                team_a_score,
                team_b_score,
                winner_side,
            ) in updated_sets
        ],

        "message":
            "경기 결과가 수정되었습니다.",
    }

# =========================
# PREDICTIONS
# 승부예측 가능 경기 조회
# =========================

# =========================
# PREDICTIONS
# 승부예측 대상 경기 조회
# =========================

@app.get(
    "/api/predictions/matches"
)
def get_prediction_matches():

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 표시할 경기 날짜 결정
            #
            # 1. 오늘 경기 있으면 오늘
            # 2. 없으면 가장 가까운 다음 경기일
            # =========================

            cursor.execute(
                """
                SELECT
                    MIN(scheduled_date)
                        AS target_date

                FROM series

                WHERE
                    series_type = '정규리그'

                    AND scheduled_date
                        IS NOT NULL

                    AND scheduled_date >= %s

                    AND status <> 'cancelled'
                """,
                (
                    today,
                ),
            )


            target_row = (
                cursor.fetchone()
            )


            target_date = (
                target_row["target_date"]
                if target_row
                else None
            )


            # 앞으로 남은 정규리그가 없음
            if target_date is None:

                return []


            # =========================
            # 해당 날짜의 경기 전부 조회
            #
            # 같은 날 2경기면
            # 둘 다 반환
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id AS series_id,
                    s.fixture_number,
                    s.round_number,
                    s.scheduled_date,
                    s.status,
                    s.started_at,

                    team_a.id
                        AS team_a_id,

                    team_a.fcl_name
                        AS team_a_fcl_name,

                    team_a.current_team_name
                        AS team_a_team_name,

                    team_a.current_team_logo_path
                        AS team_a_logo_path,

                    team_b.id
                        AS team_b_id,

                    team_b.fcl_name
                        AS team_b_fcl_name,

                    team_b.current_team_name
                        AS team_b_team_name,

                    team_b.current_team_logo_path
                        AS team_b_logo_path

                FROM series AS s

                INNER JOIN participants
                    AS team_a

                    ON team_a.id =
                        s.team_a_id

                INNER JOIN participants
                    AS team_b

                    ON team_b.id =
                        s.team_b_id

                WHERE
                    s.series_type =
                        '정규리그'

                    AND s.scheduled_date =
                        %s

                    AND s.status <>
                        'cancelled'

                ORDER BY
                    s.fixture_number ASC,
                    s.id ASC
                """,
                (
                    target_date,
                ),
            )


            rows = cursor.fetchall()


    matches = []


    for row in rows:

        # =========================
        # 예측 가능 여부
        #
        # 경기 당일부터는 마감
        # =========================

        is_open = (
            row["scheduled_date"] > today
            and row["status"] == "scheduled"
        )


        matches.append(
            {
                "series_id":
                    row["series_id"],

                "fixture_number":
                    row[
                        "fixture_number"
                    ],

                "round":
                    row[
                        "round_number"
                    ],

                "date":
                    (
                        row[
                            "scheduled_date"
                        ].isoformat()

                        if row[
                            "scheduled_date"
                        ]

                        else None
                    ),

                "status":
                    row["status"],

                "is_open":
                    is_open,

                "odds": {
                    "team_a":
                        PREDICTION_FIXED_ODDS,

                    "draw":
                        PREDICTION_FIXED_ODDS,

                    "team_b":
                        PREDICTION_FIXED_ODDS,
                },

                "max_stake_points":
                    PREDICTION_MAX_STAKE_POINTS,

                "team_a": {
                    "participant_id":
                        row[
                            "team_a_id"
                        ],

                    "fcl_name":
                        row[
                            "team_a_fcl_name"
                        ],

                    "team_name":
                        row[
                            "team_a_team_name"
                        ],

                    "logo_path":
                        row[
                            "team_a_logo_path"
                        ],
                },

                "team_b": {
                    "participant_id":
                        row[
                            "team_b_id"
                        ],

                    "fcl_name":
                        row[
                            "team_b_fcl_name"
                        ],

                    "team_name":
                        row[
                            "team_b_team_name"
                        ],

                    "logo_path":
                        row[
                            "team_b_logo_path"
                        ],
                },
            }
        )


    return matches

# =========================
# ATTENDANCE EVENT
# 출석체크 상태 조회
# =========================

@app.get(
    "/api/events/attendance"
)
def get_attendance_status(
    user = Depends(require_user)
):

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 전체 출석 횟수
            # =========================

            cursor.execute(
                """
                SELECT
                    COUNT(*) AS total_count

                FROM attendance_records

                WHERE user_id = %s
                """,
                (
                    user["id"],
                ),
            )


            total_row = cursor.fetchone()


            total_count = int(
                total_row[
                    "total_count"
                ]
                or 0
            )


            # =========================
            # 오늘 출석 여부
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    attendance_date,
                    streak_count,
                    base_reward_points,
                    streak_bonus_points,
                    reward_points

                FROM attendance_records

                WHERE
                    user_id = %s
                    AND
                    attendance_date = %s

                LIMIT 1
                """,
                (
                    user["id"],
                    today,
                ),
            )


            today_record = cursor.fetchone()


            # =========================
            # 가장 최근 출석
            # =========================

            cursor.execute(
                """
                SELECT
                    attendance_date,
                    streak_count,

                    (
                        attendance_date
                        =
                        (%s::date - 1)
                    ) AS is_yesterday

                FROM attendance_records

                WHERE user_id = %s

                ORDER BY
                    attendance_date DESC

                LIMIT 1
                """,
                (
                    today,
                    user["id"],
                ),
            )


            latest_record = cursor.fetchone()

            # =========================
            # 이번 달 출석 기록
            # =========================

            month_start = (
                today.replace(
                    day=1
                )
            )


            if today.month == 12:

                next_month_start = (
                    today.replace(
                        year=today.year + 1,
                        month=1,
                        day=1,
                    )
                )

            else:

                next_month_start = (
                    today.replace(
                        month=today.month + 1,
                        day=1,
                    )
                )


            cursor.execute(
                """
                SELECT
                    attendance_date

                FROM attendance_records

                WHERE
                    user_id = %s
                    AND
                    attendance_date >= %s
                    AND
                    attendance_date < %s

                ORDER BY
                    attendance_date ASC
                """,
                (
                    user["id"],
                    month_start,
                    next_month_start,
                ),
            )


            month_records = (
                cursor.fetchall()
            )

    month_attendance_dates = [
        record[
            "attendance_date"
        ].isoformat()

        for record
        in month_records
    ]


    current_month = (
        today.strftime(
            "%Y-%m"
        )
    )


    # =========================
    # 이미 오늘 출석함
    # =========================

    if today_record:

        current_streak = int(
            today_record[
                "streak_count"
            ]
        )


        today_reward_points = int(
            today_record[
                "reward_points"
            ]
        )


        streak_bonus_points = int(
            today_record[
                "streak_bonus_points"
            ]
        )

        month_attendance_dates = [
            record[
                "attendance_date"
            ].isoformat()

            for record
            in month_records
        ]


        current_month = (
            today.strftime(
                "%Y-%m"
            )
        )


        return {
            "today_attended":
                True,

            "attendance_date":
                today.isoformat(),

            "current_month":
                current_month,

            "month_attendance_dates":
                month_attendance_dates,

            "total_count":
                total_count,

            "current_streak":
                current_streak,

            "daily_reward_points":
                ATTENDANCE_DAILY_REWARD,

            "streak_days":
                ATTENDANCE_STREAK_DAYS,

            "streak_bonus_points":
                ATTENDANCE_STREAK_BONUS,

            "today_streak_bonus":
                streak_bonus_points,

            "today_reward_points":
                today_reward_points,
        }


    # =========================
    # 아직 오늘 출석하지 않음
    # 오늘 출석 시 연속 일수 계산
    # =========================

    if (
        latest_record
        and
        latest_record[
            "is_yesterday"
        ]
    ):

        current_streak = int(
            latest_record[
                "streak_count"
            ]
        )


        next_streak = (
            current_streak
            + 1
        )

    else:

        current_streak = 0

        next_streak = 1


    will_get_streak_bonus = (
        next_streak
        % ATTENDANCE_STREAK_DAYS
        == 0
    )


    today_streak_bonus = (
        ATTENDANCE_STREAK_BONUS
        if will_get_streak_bonus
        else 0
    )


    today_reward_points = (
        ATTENDANCE_DAILY_REWARD
        +
        today_streak_bonus
    )

    month_attendance_dates = [
        record[
            "attendance_date"
        ].isoformat()

        for record
        in month_records
    ]


    current_month = (
        today.strftime(
            "%Y-%m"
        )
    )


    return {
        "today_attended":
            False,

        "attendance_date":
            today.isoformat(),

        "current_month":
            current_month,

        "month_attendance_dates":
            month_attendance_dates,

        "total_count":
            total_count,

        "current_streak":
            current_streak,

        "next_streak":
            next_streak,

        "daily_reward_points":
            ATTENDANCE_DAILY_REWARD,

        "streak_days":
            ATTENDANCE_STREAK_DAYS,

        "streak_bonus_points":
            ATTENDANCE_STREAK_BONUS,

        "today_streak_bonus":
            today_streak_bonus,

        "today_reward_points":
            today_reward_points,
    }


# =========================
# ATTENDANCE EVENT
# 출석체크 실행
# =========================

@app.post(
    "/api/events/attendance"
)
def check_attendance(
    user = Depends(require_user)
):

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 회원 잠금
            # 동시에 여러 번 요청되어도
            # 중복 지급 방지
            # =========================

            cursor.execute(
                """
                SELECT
                    id

                FROM users

                WHERE id = %s

                FOR UPDATE
                """,
                (
                    user["id"],
                ),
            )


            locked_user = cursor.fetchone()


            if not locked_user:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "회원 정보를 "
                        "찾을 수 없습니다."
                    ),
                )


            # =========================
            # 오늘 이미 출석했는지 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id

                FROM attendance_records

                WHERE
                    user_id = %s
                    AND
                    attendance_date = %s

                LIMIT 1
                """,
                (
                    user["id"],
                    today,
                ),
            )


            existing_record = cursor.fetchone()


            if existing_record:

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "오늘은 이미 "
                        "출석체크를 완료했습니다."
                    ),
                )


            # =========================
            # 가장 최근 출석 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    attendance_date,
                    streak_count,

                    (
                        attendance_date
                        =
                        (%s::date - 1)
                    ) AS is_yesterday

                FROM attendance_records

                WHERE user_id = %s

                ORDER BY
                    attendance_date DESC

                LIMIT 1
                """,
                (
                    today,
                    user["id"],
                ),
            )


            latest_record = cursor.fetchone()


            # =========================
            # 연속 출석 계산
            # =========================

            if (
                latest_record
                and
                latest_record[
                    "is_yesterday"
                ]
            ):

                streak_count = (
                    int(
                        latest_record[
                            "streak_count"
                        ]
                    )
                    + 1
                )

            else:

                streak_count = 1


            # =========================
            # 7일 단위 보너스
            # 7 / 14 / 21 / 28 ...
            # =========================

            if (
                streak_count
                % ATTENDANCE_STREAK_DAYS
                == 0
            ):

                streak_bonus_points = (
                    ATTENDANCE_STREAK_BONUS
                )

            else:

                streak_bonus_points = 0


            reward_points = (
                ATTENDANCE_DAILY_REWARD
                +
                streak_bonus_points
            )


            # =========================
            # 출석 기록 저장
            # =========================

            cursor.execute(
                """
                INSERT INTO attendance_records (
                    user_id,
                    attendance_date,
                    streak_count,
                    base_reward_points,
                    streak_bonus_points,
                    reward_points
                )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )

                RETURNING id
                """,
                (
                    user["id"],
                    today,
                    streak_count,
                    ATTENDANCE_DAILY_REWARD,
                    streak_bonus_points,
                    reward_points,
                ),
            )


            attendance_row = cursor.fetchone()


            attendance_id = (
                attendance_row[
                    "id"
                ]
            )


            # =========================
            # 포인트 지급
            # =========================

            description = (
                f"FCL 출석체크 "
                f"{streak_count}일차"
            )


            if streak_bonus_points > 0:

                description += (
                    " + 연속 출석 보너스"
                )


            change_user_points(
                cursor,
                user["id"],
                reward_points,
                "attendance_reward",
                reference_type=
                    "attendance",
                reference_id=
                    attendance_id,
                description=
                    description,
            )


            # =========================
            # 지급 후 포인트 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    points

                FROM users

                WHERE id = %s
                """,
                (
                    user["id"],
                ),
            )


            user_row = cursor.fetchone()


            current_points = int(
                user_row[
                    "points"
                ]
            )


        connection.commit()


    return {
        "status":
            "success",

        "attendance_id":
            attendance_id,

        "attendance_date":
            today.isoformat(),

        "streak_count":
            streak_count,

        "base_reward_points":
            ATTENDANCE_DAILY_REWARD,

        "streak_bonus_points":
            streak_bonus_points,

        "reward_points":
            reward_points,

        "current_points":
            current_points,

        "message":
            (
                f"출석체크 완료! "
                f"{reward_points}P가 지급되었습니다."
            ),
    }

# =========================
# ADMIN POINT SHOP
# 교환 신청 목록 조회
# =========================

@app.get(
    "/api/admin/point-shop/exchanges"
)
def get_admin_point_shop_exchanges(
    admin_token = Depends(
        require_admin
    )
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    e.id,
                    e.user_id,
                    u.nickname,
                    u.email,

                    e.product_id,
                    e.product_name,
                    e.price_points,

                    e.status,
                    e.created_at,
                    e.completed_at

                FROM point_shop_exchanges e

                JOIN users u
                    ON u.id = e.user_id

                ORDER BY
                    CASE
                        WHEN e.status = 'requested'
                            THEN 0
                        WHEN e.status = 'completed'
                            THEN 1
                        ELSE 2
                    END,

                    e.created_at DESC,
                    e.id DESC
                """
            )


            exchanges = cursor.fetchall()


    return [
        {
            "id":
                exchange["id"],

            "user_id":
                exchange["user_id"],

            "nickname":
                exchange["nickname"],

            "email":
                exchange["email"],

            "product_id":
                exchange["product_id"],

            "product_name":
                exchange["product_name"],

            "price_points":
                int(
                    exchange[
                        "price_points"
                    ]
                ),

            "status":
                exchange["status"],

            "created_at":
                exchange[
                    "created_at"
                ].isoformat(),

            "completed_at":
                (
                    exchange[
                        "completed_at"
                    ].isoformat()

                    if exchange[
                        "completed_at"
                    ]

                    else None
                ),
        }

        for exchange
        in exchanges
    ]

# =========================
# ADMIN POINT SHOP
# 교환 처리 완료
# =========================

@app.patch(
    "/api/admin/point-shop/exchanges/{exchange_id}/complete"
)
def complete_admin_point_shop_exchange(
    exchange_id: int,

    admin_token = Depends(
        require_admin
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 교환 신청 잠금
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    user_id,
                    product_name,
                    price_points,
                    status,
                    created_at,
                    completed_at

                FROM point_shop_exchanges

                WHERE
                    id = %s

                FOR UPDATE
                """,
                (
                    exchange_id,
                ),
            )


            exchange = (
                cursor.fetchone()
            )


            if not exchange:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "교환 신청을 찾을 수 없습니다."
                    ),
                )


            if (
                exchange["status"]
                == "completed"
            ):

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "이미 처리 완료된 교환 신청입니다."
                    ),
                )


            if (
                exchange["status"]
                != "requested"
            ):

                raise HTTPException(
                    status_code=409,
                    detail=(
                        "처리할 수 없는 교환 상태입니다."
                    ),
                )


            # =========================
            # 완료 처리
            # =========================

            cursor.execute(
                """
                UPDATE point_shop_exchanges

                SET
                    status = 'completed',
                    completed_at = NOW()

                WHERE
                    id = %s

                RETURNING
                    id,
                    user_id,
                    product_name,
                    price_points,
                    status,
                    created_at,
                    completed_at
                """,
                (
                    exchange_id,
                ),
            )


            completed_exchange = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "status":
            "success",

        "message":
            "교환 신청이 처리 완료되었습니다.",

        "exchange": {
            "id":
                completed_exchange["id"],

            "user_id":
                completed_exchange["user_id"],

            "product_name":
                completed_exchange["product_name"],

            "price_points":
                int(
                    completed_exchange[
                        "price_points"
                    ]
                ),

            "status":
                completed_exchange["status"],

            "created_at":
                completed_exchange[
                    "created_at"
                ].isoformat(),

            "completed_at":
                completed_exchange[
                    "completed_at"
                ].isoformat(),
        },
    }

# =========================
# ADMIN POINT SHOP
# 상품 전체 조회
# =========================

@app.get(
    "/api/admin/point-shop/products"
)
def get_admin_point_shop_products(
    admin_token = Depends(
        require_admin
    )
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    name,
                    category,
                    description,
                    price_points,
                    image_url,
                    is_active,
                    sort_order,
                    created_at,
                    updated_at

                FROM point_shop_products

                ORDER BY
                    sort_order ASC,
                    id ASC
                """
            )


            products = (
                cursor.fetchall()
            )


    return [
        {
            "id":
                product["id"],

            "name":
                product["name"],

            "category":
                product["category"],

            "description":
                product["description"],

            "price_points":
                int(
                    product[
                        "price_points"
                    ]
                ),

            "image_url":
                product["image_url"],

            "is_active":
                product["is_active"],

            "sort_order":
                product["sort_order"],

            "created_at":
                product[
                    "created_at"
                ].isoformat(),

            "updated_at":
                product[
                    "updated_at"
                ].isoformat(),
        }

        for product
        in products
    ]


# =========================
# ADMIN POINT SHOP
# 상품 등록
# =========================

@app.post(
    "/api/admin/point-shop/products"
)
def create_admin_point_shop_product(
    request:
        AdminPointShopProductCreateRequest,

    admin_token = Depends(
        require_admin
    ),
):

    name = (
        request.name.strip()
    )


    if not name:

        raise HTTPException(
            status_code=400,
            detail=(
                "상품명을 입력해주세요."
            ),
        )


    if (
        request.price_points
        <= 0
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "상품 가격은 1P 이상이어야 합니다."
            ),
        )


    if (
        request.sort_order
        < 0
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "정렬 순서는 0 이상이어야 합니다."
            ),
        )


    category = (
        request.category.strip()
        if request.category
        else None
    )


    description = (
        request.description.strip()
        if request.description
        else None
    )


    image_url = (
        request.image_url.strip()
        if request.image_url
        else None
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO
                    point_shop_products (
                        name,
                        category,
                        description,
                        price_points,
                        image_url,
                        is_active,
                        sort_order
                    )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )

                RETURNING
                    id,
                    name,
                    category,
                    description,
                    price_points,
                    image_url,
                    is_active,
                    sort_order,
                    created_at,
                    updated_at
                """,
                (
                    name,
                    category,
                    description,
                    request.price_points,
                    image_url,
                    request.is_active,
                    request.sort_order,
                ),
            )


            product = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "status":
            "success",

        "message":
            "교환 상품이 등록되었습니다.",

        "product": {
            "id":
                product["id"],

            "name":
                product["name"],

            "category":
                product["category"],

            "description":
                product["description"],

            "price_points":
                int(
                    product[
                        "price_points"
                    ]
                ),

            "image_url":
                product["image_url"],

            "is_active":
                product["is_active"],

            "sort_order":
                product["sort_order"],
        },
    }


# =========================
# ADMIN POINT SHOP
# 상품 수정 / 판매 상태 변경
# =========================

@app.patch(
    "/api/admin/point-shop/products/{product_id}"
)
def update_admin_point_shop_product(
    product_id: int,

    request:
        AdminPointShopProductUpdateRequest,

    admin_token = Depends(
        require_admin
    ),
):

    name = (
        request.name.strip()
    )


    if not name:

        raise HTTPException(
            status_code=400,
            detail=(
                "상품명을 입력해주세요."
            ),
        )


    if (
        request.price_points
        <= 0
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "상품 가격은 1P 이상이어야 합니다."
            ),
        )


    if (
        request.sort_order
        < 0
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "정렬 순서는 0 이상이어야 합니다."
            ),
        )


    category = (
        request.category.strip()
        if request.category
        else None
    )


    description = (
        request.description.strip()
        if request.description
        else None
    )


    image_url = (
        request.image_url.strip()
        if request.image_url
        else None
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE point_shop_products

                SET
                    name = %s,
                    category = %s,
                    description = %s,
                    price_points = %s,
                    image_url = %s,
                    is_active = %s,
                    sort_order = %s,
                    updated_at = NOW()

                WHERE
                    id = %s

                RETURNING
                    id,
                    name,
                    category,
                    description,
                    price_points,
                    image_url,
                    is_active,
                    sort_order,
                    created_at,
                    updated_at
                """,
                (
                    name,
                    category,
                    description,
                    request.price_points,
                    image_url,
                    request.is_active,
                    request.sort_order,
                    product_id,
                ),
            )


            product = (
                cursor.fetchone()
            )


            if not product:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "교환 상품을 "
                        "찾을 수 없습니다."
                    ),
                )


        connection.commit()


    return {
        "status":
            "success",

        "message":
            "교환 상품이 수정되었습니다.",

        "product": {
            "id":
                product["id"],

            "name":
                product["name"],

            "category":
                product["category"],

            "description":
                product["description"],

            "price_points":
                int(
                    product[
                        "price_points"
                    ]
                ),

            "image_url":
                product["image_url"],

            "is_active":
                product["is_active"],

            "sort_order":
                product["sort_order"],
        },
    }

# =========================
# POINT SHOP
# 상품 목록
# =========================

@app.get(
    "/api/point-shop/products"
)
def get_point_shop_products():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    name,
                    category,
                    description,
                    price_points,
                    image_url,
                    sort_order

                FROM point_shop_products

                WHERE
                    is_active = TRUE

                ORDER BY
                    sort_order ASC,
                    id ASC
                """
            )


            products = (
                cursor.fetchall()
            )


    return [
        {
            "id":
                product["id"],

            "name":
                product["name"],

            "category":
                product["category"],

            "description":
                product["description"],

            "price_points":
                int(
                    product[
                        "price_points"
                    ]
                ),

            "image_url":
                product["image_url"],
        }

        for product
        in products
    ]


# =========================
# POINT SHOP
# 상품 교환
# =========================

@app.post(
    "/api/point-shop/exchanges"
)
def create_point_shop_exchange(
    request:
        PointShopExchangeRequest,

    user = Depends(
        require_user
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 상품 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    name,
                    price_points,
                    is_active

                FROM point_shop_products

                WHERE
                    id = %s

                FOR SHARE
                """,
                (
                    request.product_id,
                ),
            )


            product = (
                cursor.fetchone()
            )


            if (
                not product
                or
                not product[
                    "is_active"
                ]
            ):

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "교환 가능한 상품을 "
                        "찾을 수 없습니다."
                    ),
                )


            price_points = int(
                product[
                    "price_points"
                ]
            )


            # =========================
            # 교환 내역 먼저 생성
            #
            # 뒤 포인트 차감 실패 시
            # 같은 트랜잭션이므로
            # 이 INSERT도 자동 롤백
            # =========================

            cursor.execute(
                """
                INSERT INTO
                    point_shop_exchanges (
                        user_id,
                        product_id,
                        product_name,
                        price_points,
                        status
                    )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    'requested'
                )

                RETURNING
                    id,
                    created_at
                """,
                (
                    user["id"],
                    product["id"],
                    product["name"],
                    price_points,
                ),
            )


            exchange = (
                cursor.fetchone()
            )


            # =========================
            # 포인트 차감
            # =========================

            point_transaction = (
                change_user_points(
                    cursor,

                    user["id"],

                    -price_points,

                    "point_shop_exchange",

                    reference_type=
                        "point_shop_exchange",

                    reference_id=
                        exchange["id"],

                    description=(
                        "포인트 교환소 - "
                        f"{product['name']}"
                    ),
                )
            )


        connection.commit()


    return {
        "status":
            "success",

        "message":
            (
                f"{product['name']} "
                "교환 신청이 완료되었습니다."
            ),

        "exchange": {
            "id":
                exchange["id"],

            "product_id":
                product["id"],

            "product_name":
                product["name"],

            "price_points":
                price_points,

            "status":
                "requested",

            "created_at":
                exchange[
                    "created_at"
                ].isoformat(),
        },

        "points": {
            "spent":
                price_points,

            "balance":
                point_transaction[
                    "balance_after"
                ],
        },
    }


# =========================
# POINT SHOP
# 내 교환 내역
# =========================

@app.get(
    "/api/point-shop/exchanges/me"
)
def get_my_point_shop_exchanges(
    user = Depends(
        require_user
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    product_id,
                    product_name,
                    price_points,
                    status,
                    created_at,
                    completed_at

                FROM point_shop_exchanges

                WHERE
                    user_id = %s

                ORDER BY
                    created_at DESC,
                    id DESC
                """,
                (
                    user["id"],
                ),
            )


            exchanges = (
                cursor.fetchall()
            )


    return [
        {
            "id":
                exchange["id"],

            "product_id":
                exchange["product_id"],

            "product_name":
                exchange["product_name"],

            "price_points":
                int(
                    exchange[
                        "price_points"
                    ]
                ),

            "status":
                exchange["status"],

            "created_at":
                exchange[
                    "created_at"
                ].isoformat(),

            "completed_at":
                (
                    exchange[
                        "completed_at"
                    ].isoformat()

                    if exchange[
                        "completed_at"
                    ]

                    else None
                ),
        }

        for exchange
        in exchanges
    ]

# =========================
# PREDICTIONS
# 세트별 승부예측 참여
# =========================

@app.post(
    "/api/predictions"
)
def create_prediction(
    request:
        PredictionCreateRequest,

    user = Depends(
        require_user
    ),
):

    # =========================
    # 기본 검증
    # =========================

    if request.set_number not in (
        1,
        2,
        3,
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "세트 번호는 "
                "1, 2, 3 중 하나여야 합니다."
            ),
        )


    if request.stake_points <= 0:

        raise HTTPException(
            status_code=400,
            detail=(
                "예측 포인트는 "
                "1P 이상이어야 합니다."
            ),
        )

    if (
        request.stake_points
        >
        PREDICTION_MAX_STAKE_POINTS
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "한 세트에는 최대 "
                f"{PREDICTION_MAX_STAKE_POINTS:,}P까지 "
                "예측할 수 있습니다."
            ),
        )


    prediction_type = (
        request.prediction_type
        .strip()
        .lower()
    )


    if prediction_type not in (
        "participant",
        "draw",
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "올바르지 않은 "
                "승부예측 유형입니다."
            ),
        )


    # 현재 배당은 기존과 동일
    prediction_odds = (
        PREDICTION_FIXED_ODDS
    )


    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    with get_db_connection() as connection:

        try:

            with connection.cursor() as cursor:

                # =========================
                # SERIES 확인
                # =========================

                cursor.execute(
                    """
                    SELECT
                        id,
                        series_type,
                        team_a_id,
                        team_b_id,
                        status,
                        scheduled_date,
                        started_at

                    FROM series

                    WHERE id = %s

                    FOR UPDATE
                    """,
                    (
                        request.series_id,
                    ),
                )


                series = cursor.fetchone()


                if not series:

                    raise HTTPException(
                        status_code=404,
                        detail=(
                            "경기를 찾을 수 없습니다."
                        ),
                    )


                # =========================
                # 정규리그만 가능
                # =========================

                if (
                    series[
                        "series_type"
                    ]
                    != "정규리그"
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "정규리그 경기만 "
                            "승부예측할 수 있습니다."
                        ),
                    )


                # =========================
                # 경기 날짜 확인
                # =========================

                if (
                    series[
                        "scheduled_date"
                    ]
                    is None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "경기 날짜가 "
                            "등록되지 않았습니다."
                        ),
                    )


                # =========================
                # 예측 마감
                #
                # 경기 전날까지만 가능
                # 경기 당일 00:00부터 마감
                # =========================

                if (
                    series[
                        "scheduled_date"
                    ]
                    <= today
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "승부예측은 "
                            "경기 전날까지만 "
                            "참여할 수 있습니다."
                        ),
                    )


                # =========================
                # SERIES 상태 확인
                # =========================

                if (
                    series["status"]
                    != "scheduled"
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "이미 시작되었거나 "
                            "마감된 경기입니다."
                        ),
                    )


                # =========================
                # 예측 대상
                # =========================

                predicted_participant_id = (
                    None
                )


                if (
                    prediction_type
                    == "participant"
                ):

                    if (
                        request.participant_id
                        is None
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "예측할 참가자를 "
                                "선택해주세요."
                            ),
                        )


                    valid_participant_ids = {
                        int(
                            series[
                                "team_a_id"
                            ]
                        ),

                        int(
                            series[
                                "team_b_id"
                            ]
                        ),
                    }


                    if (
                        request.participant_id
                        not in
                        valid_participant_ids
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "해당 경기에 "
                                "참가하는 선수만 "
                                "선택할 수 있습니다."
                            ),
                        )


                    predicted_participant_id = (
                        request.participant_id
                    )


                elif (
                    prediction_type
                    == "draw"
                ):

                    if (
                        request.participant_id
                        is not None
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "무승부 예측에는 "
                                "참가자 ID가 "
                                "필요하지 않습니다."
                            ),
                        )


                # =========================
                # 같은 세트 중복 확인
                # =========================

                cursor.execute(
                    """
                    SELECT
                        id

                    FROM predictions

                    WHERE
                        user_id = %s

                        AND series_id = %s

                        AND set_number = %s
                    """,
                    (
                        user["id"],

                        request.series_id,

                        request.set_number,
                    ),
                )


                existing_prediction = (
                    cursor.fetchone()
                )


                if existing_prediction:

                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"이미 {request.set_number}세트 "
                            "승부예측에 참여했습니다."
                        ),
                    )


                # =========================
                # 예측 저장
                # =========================

                cursor.execute(
                    """
                    INSERT INTO predictions (
                        user_id,
                        series_id,
                        set_number,
                        predicted_participant_id,
                        prediction_type,
                        stake_points,
                        odds,
                        status
                    )

                    VALUES (
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        'pending'
                    )

                    RETURNING
                        id,
                        user_id,
                        series_id,
                        set_number,
                        predicted_participant_id,
                        prediction_type,
                        stake_points,
                        odds,
                        status,
                        created_at
                    """,
                    (
                        user["id"],

                        request.series_id,

                        request.set_number,

                        predicted_participant_id,

                        prediction_type,

                        request.stake_points,

                        prediction_odds,
                    ),
                )


                prediction = (
                    cursor.fetchone()
                )


                # =========================
                # 포인트 차감
                # =========================

                point_transaction = (
                    change_user_points(
                        cursor,

                        user["id"],

                        -request.stake_points,

                        "prediction_bet",

                        reference_type=
                            "prediction",

                        reference_id=
                            prediction["id"],

                        description=(
                            "FCL 승부예측 "
                            f"{request.set_number}세트 참여"
                        ),
                    )
                )


            connection.commit()


        except psycopg.errors.UniqueViolation:

            connection.rollback()

            raise HTTPException(
                status_code=409,
                detail=(
                    f"이미 {request.set_number}세트 "
                    "승부예측에 참여했습니다."
                ),
            )


    return {
        "status":
            "success",

        "message":
            (
                f"{request.set_number}세트 "
                "승부예측 참여가 완료되었습니다."
            ),

        "prediction": {

            "id":
                prediction["id"],

            "series_id":
                prediction[
                    "series_id"
                ],

            "set_number":
                prediction[
                    "set_number"
                ],

            "prediction_type":
                prediction[
                    "prediction_type"
                ],

            "participant_id":
                prediction[
                    "predicted_participant_id"
                ],

            "stake_points":
                prediction[
                    "stake_points"
                ],

            "odds":
                float(
                    prediction["odds"]
                ),

            "status":
                prediction[
                    "status"
                ],

            "created_at":
                prediction[
                    "created_at"
                ].isoformat(),
        },

        "points": {

            "spent":
                request.stake_points,

            "balance":
                point_transaction[
                    "balance_after"
                ],
        },
    }


# =========================
# PREDICTIONS
# 내 세트별 승부예측 조회
# =========================

@app.get(
    "/api/predictions/me"
)
def get_my_predictions(
    user = Depends(
        require_user
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    p.id,
                    p.series_id,
                    p.set_number,
                    p.prediction_type,
                    p.predicted_participant_id,
                    p.stake_points,
                    p.odds,
                    p.status,
                    p.payout_points,
                    p.settled_at,
                    p.created_at,

                    s.fixture_number,
                    s.round_number,
                    s.scheduled_date,
                    s.status
                        AS series_status,

                    s.team_a_id,
                    s.team_b_id,

                    team_a.fcl_name
                        AS team_a_fcl_name,

                    team_a.current_team_name
                        AS team_a_team_name,

                    team_a.current_team_logo_path
                        AS team_a_logo_path,

                    team_b.fcl_name
                        AS team_b_fcl_name,

                    team_b.current_team_name
                        AS team_b_team_name,

                    team_b.current_team_logo_path
                        AS team_b_logo_path

                FROM predictions AS p

                INNER JOIN series AS s
                    ON s.id =
                        p.series_id

                INNER JOIN participants
                    AS team_a

                    ON team_a.id =
                        s.team_a_id

                INNER JOIN participants
                    AS team_b

                    ON team_b.id =
                        s.team_b_id

                WHERE
                    p.user_id = %s

                ORDER BY
                    s.scheduled_date DESC,
                    s.fixture_number DESC,
                    p.set_number ASC,
                    p.id ASC
                """,
                (
                    user["id"],
                ),
            )


            rows = cursor.fetchall()


    predictions = []


    for row in rows:

        if (
            row["prediction_type"]
            == "draw"
        ):

            selection_side = "draw"
            selection_name = "무승부"


        elif (
            row[
                "predicted_participant_id"
            ]
            ==
            row["team_a_id"]
        ):

            selection_side = "team_a"

            selection_name = (
                row[
                    "team_a_fcl_name"
                ]
            )


        elif (
            row[
                "predicted_participant_id"
            ]
            ==
            row["team_b_id"]
        ):

            selection_side = "team_b"

            selection_name = (
                row[
                    "team_b_fcl_name"
                ]
            )


        else:

            selection_side = "unknown"
            selection_name = "알 수 없음"


        predictions.append(
            {
                "id":
                    row["id"],

                "series_id":
                    row[
                        "series_id"
                    ],

                "set_number":
                    row[
                        "set_number"
                    ],

                "fixture_number":
                    row[
                        "fixture_number"
                    ],

                "round":
                    row[
                        "round_number"
                    ],

                "date":
                    (
                        row[
                            "scheduled_date"
                        ].isoformat()

                        if row[
                            "scheduled_date"
                        ]

                        else None
                    ),

                "series_status":
                    row[
                        "series_status"
                    ],

                "prediction_type":
                    row[
                        "prediction_type"
                    ],

                "selection_side":
                    selection_side,

                "selection_name":
                    selection_name,

                "stake_points":
                    row[
                        "stake_points"
                    ],

                "odds":
                    float(
                        row["odds"]
                    ),

                "expected_payout":
                    int(
                        row[
                            "stake_points"
                        ]
                        *
                        float(
                            row["odds"]
                        )
                    ),

                "status":
                    row["status"],

                "payout_points":
                    row[
                        "payout_points"
                    ],

                "team_a": {

                    "participant_id":
                        row[
                            "team_a_id"
                        ],

                    "fcl_name":
                        row[
                            "team_a_fcl_name"
                        ],

                    "team_name":
                        row[
                            "team_a_team_name"
                        ],

                    "logo_path":
                        row[
                            "team_a_logo_path"
                        ],
                },

                "team_b": {

                    "participant_id":
                        row[
                            "team_b_id"
                        ],

                    "fcl_name":
                        row[
                            "team_b_fcl_name"
                        ],

                    "team_name":
                        row[
                            "team_b_team_name"
                        ],

                    "logo_path":
                        row[
                            "team_b_logo_path"
                        ],
                },

                "created_at":
                    row[
                        "created_at"
                    ].isoformat(),

                "settled_at":
                    (
                        row[
                            "settled_at"
                        ].isoformat()

                        if row[
                            "settled_at"
                        ]

                        else None
                    ),
            }
        )


    return predictions

# =========================
# PREDICTIONS
# 자동 정산 테스트
#
# 실제 DB 변경 없음
# 마지막에 무조건 ROLLBACK
# =========================

@app.post(
    "/api/admin/predictions/settlement-test"
)
def test_prediction_settlement(
    request:
        PredictionSettlementTestRequest,

    admin_user = Depends(
        require_user_admin
    ),
):

    with get_db_connection() as connection:

        try:

            with connection.cursor() as cursor:

                # =========================
                # SERIES 확인
                # =========================

                cursor.execute(
                    """
                    SELECT
                        id,
                        series_type,
                        status,
                        team_a_id,
                        team_b_id

                    FROM series

                    WHERE id = %s

                    FOR UPDATE
                    """,
                    (
                        request.series_id,
                    ),
                )


                series = cursor.fetchone()


                if not series:

                    raise HTTPException(
                        status_code=404,
                        detail=(
                            "경기를 찾을 수 없습니다."
                        ),
                    )


                if (
                    series["series_type"]
                    != "정규리그"
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "정규리그 경기만 "
                            "정산 테스트할 수 있습니다."
                        ),
                    )


                # =========================
                # 현재 pending 예측 조회
                # =========================

                cursor.execute(
                    """
                    SELECT
                        p.id,
                        p.user_id,
                        p.set_number,
                        p.prediction_type,
                        p.predicted_participant_id,
                        p.stake_points,
                        p.odds,
                        p.status,

                        u.nickname,
                        u.points

                    FROM predictions AS p

                    INNER JOIN users AS u
                        ON u.id =
                            p.user_id

                    WHERE
                        p.series_id = %s

                        AND p.status =
                            'pending'

                    ORDER BY
                        p.set_number,
                        p.id
                    """,
                    (
                        request.series_id,
                    ),
                )


                before_predictions = (
                    cursor.fetchall()
                )


                if (
                    len(
                        before_predictions
                    )
                    == 0
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "이 경기에 정산 테스트할 "
                            "pending 예측이 없습니다."
                        ),
                    )


                before_points = {
                    int(
                        row["user_id"]
                    ):
                        int(
                            row["points"]
                        )

                    for row
                    in before_predictions
                }


                # =========================
                # 테스트용 경기 완료 처리
                #
                # ROLLBACK되므로
                # 실제 상태는 바뀌지 않음
                # =========================

                cursor.execute(
                    """
                    UPDATE series

                    SET
                        status =
                            'completed'

                    WHERE id = %s
                    """,
                    (
                        request.series_id,
                    ),
                )


                # =========================
                # 테스트 세트 결과
                # =========================

                test_sets = [
                    (
                        1,
                        request.set1_team_a,
                        request.set1_team_b,
                    ),
                    (
                        2,
                        request.set2_team_a,
                        request.set2_team_b,
                    ),
                    (
                        3,
                        request.set3_team_a,
                        request.set3_team_b,
                    ),
                ]


                for (
                    set_number,
                    team_a_score,
                    team_b_score,
                ) in test_sets:

                    if (
                        team_a_score < 0
                        or
                        team_b_score < 0
                    ):

                        raise HTTPException(
                            status_code=400,
                            detail=(
                                "세트 점수는 "
                                "0 이상이어야 합니다."
                            ),
                        )


                    if (
                        team_a_score
                        >
                        team_b_score
                    ):

                        winner_side = (
                            "team_a"
                        )


                    elif (
                        team_a_score
                        <
                        team_b_score
                    ):

                        winner_side = (
                            "team_b"
                        )


                    else:

                        winner_side = (
                            "draw"
                        )


                    cursor.execute(
                        """
                        INSERT INTO series_sets (
                            series_id,
                            set_number,
                            nexon_match_id,
                            played_at,
                            team_a_score,
                            team_b_score,
                            score_source,
                            winner_side
                        )

                        VALUES (
                            %s,
                            %s,
                            NULL,
                            NOW(),
                            %s,
                            %s,
                            'manual',
                            %s
                        )

                        ON CONFLICT (
                            series_id,
                            set_number
                        )

                        DO UPDATE SET
                            team_a_score =
                                EXCLUDED.team_a_score,

                            team_b_score =
                                EXCLUDED.team_b_score,

                            score_source =
                                'manual',

                            winner_side =
                                EXCLUDED.winner_side
                        """,
                        (
                            request.series_id,
                            set_number,
                            team_a_score,
                            team_b_score,
                            winner_side,
                        ),
                    )


                # =========================
                # 실제 정산 함수 실행
                # =========================

                settlement_result = (
                    settle_predictions_for_series(
                        cursor,
                        request.series_id,
                    )
                )


                # =========================
                # 정산 후 예측 상태
                # =========================

                cursor.execute(
                    """
                    SELECT
                        p.id,
                        p.user_id,
                        p.set_number,
                        p.prediction_type,
                        p.predicted_participant_id,
                        p.stake_points,
                        p.odds,
                        p.status,
                        p.payout_points,

                        u.nickname,
                        u.points

                    FROM predictions AS p

                    INNER JOIN users AS u
                        ON u.id =
                            p.user_id

                    WHERE
                        p.series_id = %s

                    ORDER BY
                        p.set_number,
                        p.id
                    """,
                    (
                        request.series_id,
                    ),
                )


                after_predictions = (
                    cursor.fetchall()
                )


                prediction_results = []


                for row in after_predictions:

                    user_id = int(
                        row["user_id"]
                    )


                    points_before = (
                        before_points.get(
                            user_id,
                            int(
                                row["points"]
                            ),
                        )
                    )


                    points_after = int(
                        row["points"]
                    )


                    prediction_results.append(
                        {
                            "prediction_id":
                                row["id"],

                            "nickname":
                                row[
                                    "nickname"
                                ],

                            "set_number":
                                row[
                                    "set_number"
                                ],

                            "prediction_type":
                                row[
                                    "prediction_type"
                                ],

                            "participant_id":
                                row[
                                    "predicted_participant_id"
                                ],

                            "stake_points":
                                row[
                                    "stake_points"
                                ],

                            "odds":
                                float(
                                    row[
                                        "odds"
                                    ]
                                ),

                            "result_status":
                                row[
                                    "status"
                                ],

                            "payout_points":
                                row[
                                    "payout_points"
                                ],

                            "points_before":
                                points_before,

                            "points_after":
                                points_after,

                            "point_change":
                                int(
                                    row[
                                        "payout_points"
                                    ]
                                    or 0
                                ),
                            "user_total_point_change":
                                (
                                    points_after
                                    -
                                    points_before
                                ),
                        }
                    )


            # =========================
            # 핵심
            #
            # 모든 테스트 변경 취소
            # =========================

            connection.rollback()


        except Exception:

            connection.rollback()

            raise


    return {
        "status":
            "success",

        "database_change":
            "rolled_back",

        "message":
            (
                "정산 테스트가 완료되었습니다. "
                "실제 DB 변경은 모두 취소되었습니다."
            ),

        "settlement":
            settlement_result,

        "predictions":
            prediction_results,
    }


# =========================
# 전체 경기 일정
# =========================

@app.get("/api/matches")
def get_matches():

    matches = []

    database_match_keys = set()


    # =========================================
    # 1. Neon DB 일정
    #
    # 정규리그 + 프리시즌
    # =========================================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id AS series_id,
                    s.fixture_number,
                    s.series_type,
                    s.scheduled_date,
                    s.round_number,
                    s.status,

                    team_a.fcl_name
                        AS team_a,

                    team_b.fcl_name
                        AS team_b

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE
                    s.scheduled_date
                        IS NOT NULL

                    AND
                    s.status <>
                        'cancelled'

                ORDER BY
                    s.scheduled_date,
                    s.fixture_number,
                    s.id
                """
            )


            series_rows = cursor.fetchall()


    for series_row in series_rows:

        match_date = (
            series_row[
                "scheduled_date"
            ].isoformat()
        )


        team_a = (
            series_row[
                "team_a"
            ]
        )

        team_b = (
            series_row[
                "team_b"
            ]
        )


        series_type = (
            series_row[
                "series_type"
            ]
        )


        teams = sorted(
            [
                team_a,
                team_b,
            ]
        )


        database_match_keys.add(
            (
                match_date,
                series_type,
                teams[0],
                teams[1],
            )
        )


        matches.append(
            {
                "series_id":
                    series_row[
                        "series_id"
                    ],

                "fixture_number":
                    series_row[
                        "fixture_number"
                    ],

                "source":
                    "database",

                "date":
                    match_date,

                "round":
                    series_row[
                        "round_number"
                    ],

                "match_type":
                    series_type,

                "team_a":
                    team_a,

                "team_b":
                    team_b,

                "status":
                    series_row[
                        "status"
                    ],
            }
        )


    # =========================================
    # 2. 기존 Excel
    #
    # 프리시즌만 임시 유지
    # 정규리그는 이제 DB 사용
    # =========================================

    workbook = load_workbook(
        EXCEL_PATH
    )


    worksheet = workbook[
        "경기일정"
    ]


    for row in worksheet.iter_rows(
        min_row=2,
        max_col=4,
        values_only=True,
    ):

        (
            match_date,
            team_a,
            team_b,
            match_type,
        ) = row


        if match_date is None:
            continue


        # 경기구분이 비어 있으면
        # 기존 정규리그 데이터이므로 제외
        if match_type is None:
            continue


        # 정규리그는 DB에서만 조회
        if (
            match_type
            != "프리시즌"
        ):
            continue


        if hasattr(
            match_date,
            "strftime"
        ):

            match_date = (
                match_date.strftime(
                    "%Y-%m-%d"
                )
            )


        match_date = str(
            match_date
        )


        teams = sorted(
            [
                team_a,
                team_b,
            ]
        )


        match_key = (
            match_date,
            "프리시즌",
            teams[0],
            teams[1],
        )


        # 같은 프리시즌이 DB에 있으면
        # DB 데이터 우선
        if (
            match_key
            in database_match_keys
        ):
            continue


        matches.append(
            {
                "series_id":
                    None,

                "fixture_number":
                    None,

                "source":
                    "excel",

                "date":
                    match_date,

                "round":
                    None,

                "match_type":
                    "프리시즌",

                "team_a":
                    team_a,

                "team_b":
                    team_b,

                "status":
                    None,
            }
        )


    workbook.close()

    # =========================================
    # 참가자 현재 팀 정보 연결
    #
    # 일정 화면은 항상 participants의
    # 현재 팀 정보를 사용
    #
    # 과거 경기 결과 Snapshot과는 별개
    # =========================================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    fcl_name,
                    current_team_name,
                    current_team_logo_path

                FROM participants
                """
            )

            participant_team_rows = (
                cursor.fetchall()
            )


    participant_team_map = {
        row["fcl_name"]: row
        for row in participant_team_rows
    }


    for match in matches:

        team_a_current = (
            participant_team_map.get(
                match["team_a"]
            )
        )

        team_b_current = (
            participant_team_map.get(
                match["team_b"]
            )
        )


        if team_a_current:

            match[
                "team_a_current_team_name"
            ] = (
                team_a_current[
                    "current_team_name"
                ]
            )

            match[
                "team_a_current_team_logo_path"
            ] = (
                team_a_current[
                    "current_team_logo_path"
                ]
            )

        else:

            match[
                "team_a_current_team_name"
            ] = None

            match[
                "team_a_current_team_logo_path"
            ] = None


        if team_b_current:

            match[
                "team_b_current_team_name"
            ] = (
                team_b_current[
                    "current_team_name"
                ]
            )

            match[
                "team_b_current_team_logo_path"
            ] = (
                team_b_current[
                    "current_team_logo_path"
                ]
            )

        else:

            match[
                "team_b_current_team_name"
            ] = None

            match[
                "team_b_current_team_logo_path"
            ] = None


    # =========================================
    # 날짜순
    # =========================================

    matches.sort(
        key=lambda match: (
            match["date"],
            (
                match[
                    "fixture_number"
                ]
                or 9999
            ),
            match["team_a"],
            match["team_b"],
        )
    )


    return matches

# =========================
# 오늘 경기
# =========================

@app.get("/api/matches/today")
def get_today_matches():

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    today_string = (
        today.isoformat()
    )


    # 기존 Excel 일정 사용
    all_matches = get_matches()


    matches = [
        match
        for match in all_matches
        if match["date"] == today_string
    ]


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            for match in matches:

                cursor.execute(
                    """
                    SELECT
                        s.id AS series_id,
                        s.status AS series_status

                    FROM series AS s

                    JOIN participants AS team_a
                        ON team_a.id =
                            s.team_a_id

                    JOIN participants AS team_b
                        ON team_b.id =
                            s.team_b_id

                    WHERE
                        s.scheduled_date = %s

                        AND
                        s.series_type = %s

                        AND
                        s.status <> 'cancelled'

                        AND (
                            (
                                team_a.fcl_name = %s
                                AND
                                team_b.fcl_name = %s
                            )

                            OR

                            (
                                team_a.fcl_name = %s
                                AND
                                team_b.fcl_name = %s
                            )
                        )

                    ORDER BY
                        CASE s.status

                            WHEN 'active'
                                THEN 1

                            WHEN 'scheduled'
                                THEN 2

                            WHEN 'completed'
                                THEN 3

                            ELSE 4

                        END,

                        s.id DESC

                    LIMIT 1
                    """,
                    (
                        today,

                        match["match_type"],

                        match["team_a"],
                        match["team_b"],

                        match["team_b"],
                        match["team_a"],
                    ),
                )


                series = cursor.fetchone()


                if series:

                    match["series_id"] = (
                        series["series_id"]
                    )

                    match["series_status"] = (
                        series["series_status"]
                    )

                else:

                    match["series_id"] = None

                    match["series_status"] = (
                        "not_started"
                    )


    return matches

# =========================
# 경기 결과
# =========================

@app.get("/api/results")
def get_results():
    workbook = load_workbook(RESULTS_PATH)
    worksheet = workbook["경기결과"]

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()

    results = []

    regular_match_index = 0

    for row in worksheet.iter_rows(
        min_row=2,
        max_col=10,
        values_only=True,
    ):
        (
            match_date,
            team_a,
            set1_team_a,
            set1_team_b,
            set2_team_a,
            set2_team_b,
            set3_team_a,
            set3_team_b,
            team_b,
            match_type,
        ) = row


        if match_date is None:
            continue


        if match_type is None:
            match_type = "정규리그"


        # ----------------------------------
        # 라운드 계산
        # 프리시즌은 라운드 계산에서 제외
        # ----------------------------------

        if match_type == "프리시즌":
            round_number = None

        else:
            round_number = get_round_number(
                regular_match_index
            )

            regular_match_index += 1


        if hasattr(match_date, "date"):
            match_date = match_date.date()


        # 오늘 경기와 미래 경기는 결과에서 제외
        if match_date >= today:
            continue


        scores = [
            set1_team_a,
            set1_team_b,
            set2_team_a,
            set2_team_b,
            set3_team_a,
            set3_team_b,
        ]


        # 세트 점수가 전부 입력된 경기만 출력
        if any(score is None for score in scores):
            continue


        team_a_total_score = (
            set1_team_a
            + set2_team_a
            + set3_team_a
        )

        team_b_total_score = (
            set1_team_b
            + set2_team_b
            + set3_team_b
        )


        results.append(
            {
                "date": match_date.strftime(
                    "%Y-%m-%d"
                ),

                "round": round_number,

                "match_type": match_type,

                "team_a": team_a,

                "team_b": team_b,

                "team_a_score": team_a_total_score,

                "team_b_score": team_b_total_score,

                "sets": [
                    {
                        "set": 1,
                        "team_a_score": set1_team_a,
                        "team_b_score": set1_team_b,
                    },
                    {
                        "set": 2,
                        "team_a_score": set2_team_a,
                        "team_b_score": set2_team_b,
                    },
                    {
                        "set": 3,
                        "team_a_score": set3_team_a,
                        "team_b_score": set3_team_b,
                    },
                ],
            }
        )


    workbook.close()

    return results

# =========================
# 팀 순위
# Excel + Neon PostgreSQL
# =========================

@app.get("/api/standings")
def get_standings():

    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    # =========================
    # 기본 순위 데이터
    # =========================

    standings = {}


    for participant in PARTICIPANTS:

        standings[participant] = {
            "name": participant,

            # SERIES 수
            "played": 0,

            # 세트 기준
            "wins": 0,
            "draws": 0,
            "losses": 0,

            "goals_for": 0,
            "goals_against": 0,
            "goal_difference": 0,

            "points": 0,
        }


    # =========================
    # 경기 중복 판별 키
    # =========================

    def make_match_key(
        match_date,
        team_a,
        team_b,
    ):

        teams = sorted(
            [
                team_a,
                team_b,
            ]
        )


        return (
            match_date.isoformat(),
            teams[0],
            teams[1],
        )


    # =========================
    # 실제 순위 반영 함수
    # =========================

    def apply_match_result(
        team_a,
        team_b,
        sets,
    ):

        if (
            team_a not in standings
            or
            team_b not in standings
        ):
            return


        # FCL 정규리그는 3세트
        if len(sets) != 3:
            return


        team_a_record = (
            standings[team_a]
        )

        team_b_record = (
            standings[team_b]
        )


        # SERIES 경기 수
        team_a_record["played"] += 1
        team_b_record["played"] += 1


        # =========================
        # 세트별 계산
        # =========================

        for (
            team_a_score,
            team_b_score,
        ) in sets:

            team_a_score = int(
                team_a_score
            )

            team_b_score = int(
                team_b_score
            )


            # 득점 / 실점
            team_a_record[
                "goals_for"
            ] += team_a_score

            team_a_record[
                "goals_against"
            ] += team_b_score


            team_b_record[
                "goals_for"
            ] += team_b_score

            team_b_record[
                "goals_against"
            ] += team_a_score


            # =========================
            # 팀A 승
            # =========================

            if (
                team_a_score
                >
                team_b_score
            ):

                team_a_record[
                    "wins"
                ] += 1

                team_a_record[
                    "points"
                ] += 3

                team_b_record[
                    "losses"
                ] += 1


            # =========================
            # 팀B 승
            # =========================

            elif (
                team_a_score
                <
                team_b_score
            ):

                team_b_record[
                    "wins"
                ] += 1

                team_b_record[
                    "points"
                ] += 3

                team_a_record[
                    "losses"
                ] += 1


            # =========================
            # 무승부
            # =========================

            else:

                team_a_record[
                    "draws"
                ] += 1

                team_b_record[
                    "draws"
                ] += 1


                team_a_record[
                    "points"
                ] += 1

                team_b_record[
                    "points"
                ] += 1


    # =========================================
    # 1. Neon 정규리그 결과
    #
    # DB를 우선으로 계산
    # =========================================

    database_series = {}


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id
                        AS series_id,

                    s.scheduled_date,
                    s.started_at,

                    team_a.fcl_name
                        AS team_a,

                    team_b.fcl_name
                        AS team_b,

                    ss.set_number,
                    ss.team_a_score,
                    ss.team_b_score

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                JOIN series_sets AS ss
                    ON ss.series_id =
                        s.id

                WHERE
                    s.status = 'completed'

                    AND
                    s.series_type = '정규리그'

                ORDER BY
                    s.id,
                    ss.set_number
                """
            )


            database_rows = (
                cursor.fetchall()
            )


    # =========================
    # SERIES별 묶기
    # =========================

    for row in database_rows:

        series_id = row[
            "series_id"
        ]


        if (
            series_id
            not in database_series
        ):

            database_series[
                series_id
            ] = {
                "scheduled_date":
                    row[
                        "scheduled_date"
                    ],

                "started_at":
                    row[
                        "started_at"
                    ],

                "team_a":
                    row[
                        "team_a"
                    ],

                "team_b":
                    row[
                        "team_b"
                    ],

                "sets": [],
            }


        database_series[
            series_id
        ][
            "sets"
        ].append(
            (
                row[
                    "team_a_score"
                ],

                row[
                    "team_b_score"
                ],
            )
        )


    # DB에 존재하는 경기 키
    database_match_keys = set()


    # =========================
    # Neon 결과 순위 반영
    # =========================

    for series_data in (
        database_series.values()
    ):

        match_date = (
            series_data[
                "scheduled_date"
            ]
        )


        # 예전 테스트 데이터 보호
        if match_date is None:

            match_date = (
                parse_kst_datetime(
                    series_data[
                        "started_at"
                    ]
                ).date()
            )


        # 미래 경기 보호
        if match_date > today:
            continue


        sets = series_data[
            "sets"
        ]


        if len(sets) != 3:
            continue


        match_key = (
            make_match_key(
                match_date,
                series_data[
                    "team_a"
                ],
                series_data[
                    "team_b"
                ],
            )
        )


        database_match_keys.add(
            match_key
        )


        apply_match_result(
            series_data[
                "team_a"
            ],

            series_data[
                "team_b"
            ],

            sets,
        )


    # =========================================
    # 2. 기존 Excel 결과
    # =========================================

    workbook = load_workbook(
        RESULTS_PATH
    )

    worksheet = workbook[
        "경기결과"
    ]


    for row in worksheet.iter_rows(
        min_row=2,
        max_col=10,
        values_only=True,
    ):

        (
            match_date,

            team_a,

            set1_team_a,
            set1_team_b,

            set2_team_a,
            set2_team_b,

            set3_team_a,
            set3_team_b,

            team_b,

            match_type,
        ) = row


        # 빈 행
        if match_date is None:
            continue


        # 경기구분 비어 있으면 정규리그
        if match_type is None:

            match_type = (
                "정규리그"
            )


        # =========================
        # 정규리그만 순위 반영
        # =========================

        if (
            match_type
            != "정규리그"
        ):
            continue


        # =========================
        # 날짜 변환
        # =========================

        if hasattr(
            match_date,
            "date"
        ):

            match_date = (
                match_date.date()
            )


        elif isinstance(
            match_date,
            str
        ):

            try:

                match_date = (
                    datetime.strptime(
                        match_date,
                        "%Y-%m-%d",
                    ).date()
                )

            except ValueError:

                continue


        # 미래 경기 제외
        if match_date > today:
            continue


        # =========================
        # 3세트 확인
        # =========================

        scores = [
            set1_team_a,
            set1_team_b,

            set2_team_a,
            set2_team_b,

            set3_team_a,
            set3_team_b,
        ]


        if any(
            score is None
            for score in scores
        ):
            continue


        # =========================
        # Neon과 중복이면 Excel 제외
        #
        # DB 데이터 우선
        # =========================

        match_key = (
            make_match_key(
                match_date,
                team_a,
                team_b,
            )
        )


        if (
            match_key
            in database_match_keys
        ):
            continue


        sets = [
            (
                set1_team_a,
                set1_team_b,
            ),
            (
                set2_team_a,
                set2_team_b,
            ),
            (
                set3_team_a,
                set3_team_b,
            ),
        ]


        apply_match_result(
            team_a,
            team_b,
            sets,
        )


    workbook.close()


    # =========================
    # 득실차
    # =========================

    for record in standings.values():

        record[
            "goal_difference"
        ] = (
            record[
                "goals_for"
            ]
            -
            record[
                "goals_against"
            ]
        )


    # =========================
    # 동률 시 참가자 기본 순서
    # =========================

    participant_order = {
        participant: index

        for index, participant
        in enumerate(
            PARTICIPANTS
        )
    }


    # =========================
    # 순위
    #
    # 1. 승점
    # 2. 득실차
    # 3. 득점
    # =========================

    sorted_standings = sorted(
        standings.values(),

        key=lambda record: (
            -record["points"],

            -record[
                "goal_difference"
            ],

            -record[
                "goals_for"
            ],

            participant_order[
                record["name"]
            ],
        ),
    )


    # =========================
    # 순위 번호
    # =========================

    for index, record in enumerate(
        sorted_standings,
        start=1,
    ):

        record["rank"] = index


    # =========================
    # 참가자 현재 팀 정보
    #
    # 현재 화면에서 사용하는 팀 정보는
    # participants 테이블을 기준으로 함
    #
    # 과거 SERIES Snapshot은
    # 절대 수정하지 않음
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    fcl_name,
                    current_team_name,
                    current_team_logo_path

                FROM participants
                """
            )

            participant_team_rows = (
                cursor.fetchall()
            )


    participant_team_map = {
        row["fcl_name"]: row
        for row in participant_team_rows
    }


    for record in sorted_standings:

        participant_team = (
            participant_team_map.get(
                record["name"]
            )
        )

        if participant_team:

            record["current_team_name"] = (
                participant_team[
                    "current_team_name"
                ]
            )

            record[
                "current_team_logo_path"
            ] = (
                participant_team[
                    "current_team_logo_path"
                ]
            )

        else:

            record["current_team_name"] = None

            record[
                "current_team_logo_path"
            ] = None


    return sorted_standings

# =========================
# PLAYOFF PREVIEW
# 현재 정규리그 순위 기준
# =========================

@app.get(
    "/api/admin/playoffs/preview"
)
def preview_playoffs(
    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # 정규리그 진행 상태
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    COUNT(*) AS total_count,

                    COUNT(*) FILTER (
                        WHERE status = 'completed'
                    ) AS completed_count

                FROM series

                WHERE
                    series_type = '정규리그'
                    AND status <> 'cancelled'
                """
            )

            regular_status = (
                cursor.fetchone()
            )


    total_count = int(
        regular_status["total_count"]
    )

    completed_count = int(
        regular_status["completed_count"]
    )


    # =========================
    # 현재 정규리그 순위
    # =========================

    standings = get_standings()


    if len(standings) < 5:

        raise HTTPException(
            status_code=400,
            detail=(
                "정규리그 순위를 "
                "확인할 수 없습니다."
            ),
        )


    first_place = standings[0]
    second_place = standings[1]
    third_place = standings[2]
    fourth_place = standings[3]
    fifth_place = standings[4]


    # =========================
    # 미리보기 반환
    # =========================

    return {
        "regular_league": {
            "total": total_count,
            "completed": completed_count,
            "is_completed": (
                total_count == 20
                and
                completed_count == 20
            ),
        },

        "standings": [
            {
                "rank": participant["rank"],
                "name": participant["name"],
                "points": participant["points"],
                "goal_difference":
                    participant[
                        "goal_difference"
                    ],
                "goals_for":
                    participant["goals_for"],
            }

            for participant
            in standings[:5]
        ],

        "playoff_bracket": {
            "준플레이오프": {
                "best_of": 5,
                "wins_required": 3,

                "team_a": {
                    "rank": 3,
                    "name":
                        third_place["name"],
                },

                "team_b": {
                    "rank": 4,
                    "name":
                        fourth_place["name"],
                },
            },

            "플레이오프": {
                "best_of": 5,
                "wins_required": 3,

                "team_a": {
                    "rank": 2,
                    "name":
                        second_place["name"],
                },

                "team_b": {
                    "source":
                        "준플레이오프 승자",
                },
            },

            "결승시리즈": {
                "best_of": 7,
                "wins_required": 4,

                "team_a": {
                    "rank": 1,
                    "name":
                        first_place["name"],
                },

                "team_b": {
                    "source":
                        "플레이오프 승자",
                },
            },
        },

        "eliminated": {
            "rank": 5,
            "name":
                fifth_place["name"],
        },
    }


# =========================
# PLAYOFF INITIALIZE
# 정규리그 3위 VS 4위
# =========================

@app.post(
    "/api/admin/playoffs/initialize"
)
def initialize_playoffs(
    request: PlayoffInitializeRequest,

    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # 경기 날짜 검증
    # =========================

    try:

        scheduled_date = (
            datetime.strptime(
                request.scheduled_date,
                "%Y-%m-%d",
            ).date()
        )

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail=(
                "경기 날짜 형식은 "
                "YYYY-MM-DD여야 합니다."
            ),
        )


    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    if scheduled_date < today:

        raise HTTPException(
            status_code=400,
            detail=(
                "지난 날짜로 플레이오프를 "
                "생성할 수 없습니다."
            ),
        )


    # =========================
    # 정규리그 완료 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    COUNT(*) AS total_count,

                    COUNT(*) FILTER (
                        WHERE status = 'completed'
                    ) AS completed_count

                FROM series

                WHERE
                    series_type = '정규리그'
                    AND status <> 'cancelled'
                """
            )


            regular_status = (
                cursor.fetchone()
            )


    total_count = int(
        regular_status[
            "total_count"
        ]
    )

    completed_count = int(
        regular_status[
            "completed_count"
        ]
    )


    if (
        total_count != 20
        or
        completed_count != 20
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "정규리그 20경기가 "
                "모두 완료된 후 "
                "플레이오프를 생성할 수 있습니다."
            ),
        )


    # =========================
    # 기존 플레이오프 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    playoff_stage,
                    status

                FROM series

                WHERE
                    series_type = '플레이오프'

                    AND
                    playoff_stage = '준플레이오프'

                    AND
                    status <> 'cancelled'

                LIMIT 1
                """
            )


            existing_series = (
                cursor.fetchone()
            )


    if existing_series:

        raise HTTPException(
            status_code=400,
            detail=(
                "이미 준플레이오프 "
                "SERIES가 생성되어 있습니다."
            ),
        )


    # =========================
    # 정규리그 최종 순위
    # =========================

    standings = get_standings()


    if len(standings) < 4:

        raise HTTPException(
            status_code=400,
            detail=(
                "플레이오프 진출자를 "
                "확정할 수 없습니다."
            ),
        )


    third_place = standings[2]

    fourth_place = standings[3]


    team_a_name = (
        third_place["name"]
    )

    team_b_name = (
        fourth_place["name"]
    )


    # =========================
    # 참가자 ID 조회
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname

                FROM participants

                WHERE fcl_name IN (
                    %s,
                    %s
                )
                """,
                (
                    team_a_name,
                    team_b_name,
                ),
            )


            participant_rows = (
                cursor.fetchall()
            )


            participant_map = {
                participant[
                    "fcl_name"
                ]:
                    participant

                for participant
                in participant_rows
            }


            if (
                team_a_name
                not in participant_map

                or

                team_b_name
                not in participant_map
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "플레이오프 참가자 정보를 "
                        "찾을 수 없습니다."
                    ),
                )


            team_a = participant_map[
                team_a_name
            ]

            team_b = participant_map[
                team_b_name
            ]


            # =========================
            # 준플레이오프 SERIES 생성
            # =========================

            cursor.execute(
                """
                INSERT INTO series (
                    series_type,

                    team_a_id,
                    team_b_id,

                    match_type,

                    scheduled_date,

                    playoff_stage,
                    best_of,
                    wins_required,

                    stats_sync_status,
                    status
                )

                VALUES (
                    '플레이오프',

                    %s,
                    %s,

                    40,

                    %s,

                    '준플레이오프',
                    5,
                    3,

                    'pending',
                    'scheduled'
                )

                RETURNING
                    id,
                    scheduled_date,
                    playoff_stage,
                    best_of,
                    wins_required,
                    status
                """,
                (
                    team_a["id"],
                    team_b["id"],
                    scheduled_date,
                ),
            )


            series_row = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "series_id":
            series_row["id"],

        "series_type":
            "플레이오프",

        "playoff_stage":
            series_row[
                "playoff_stage"
            ],

        "scheduled_date":
            series_row[
                "scheduled_date"
            ].isoformat(),

        "team_a": {
            "rank": 3,
            "fcl_name":
                team_a["fcl_name"],
            "nickname":
                team_a[
                    "fc_nickname"
                ],
        },

        "team_b": {
            "rank": 4,
            "fcl_name":
                team_b["fcl_name"],
            "nickname":
                team_b[
                    "fc_nickname"
                ],
        },

        "best_of":
            series_row[
                "best_of"
            ],

        "wins_required":
            series_row[
                "wins_required"
            ],

        "status":
            series_row["status"],
    }


# =========================
# PLAYOFF ADVANCE
# 다음 플레이오프 단계 생성
# =========================

@app.post(
    "/api/admin/playoffs/{series_id}/advance"
)
def advance_playoff_series(
    series_id: int,
    request: PlayoffAdvanceRequest,

    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # 다음 경기 날짜 검증
    # =========================

    try:

        scheduled_date = (
            datetime.strptime(
                request.scheduled_date,
                "%Y-%m-%d",
            ).date()
        )

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail=(
                "경기 날짜 형식은 "
                "YYYY-MM-DD여야 합니다."
            ),
        )


    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    if scheduled_date < today:

        raise HTTPException(
            status_code=400,
            detail=(
                "지난 날짜로 다음 "
                "플레이오프 경기를 "
                "생성할 수 없습니다."
            ),
        )


    # =========================
    # 정규리그 최종 순위
    #
    # 다음 단계의
    # 1위 / 2위 참가자 확인
    # =========================

    standings = get_standings()


    if len(standings) < 4:

        raise HTTPException(
            status_code=400,
            detail=(
                "정규리그 최종 순위를 "
                "확인할 수 없습니다."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 현재 플레이오프 SERIES
            #
            # 동시에 advance 되는 것을
            # 방지하기 위해 행 잠금
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,
                    s.status,
                    s.completed_at,

                    s.team_a_id,
                    s.team_b_id,

                    team_a.fcl_name
                        AS team_a_name,

                    team_a.fc_nickname
                        AS team_a_nickname,

                    team_b.fcl_name
                        AS team_b_name,

                    team_b.fc_nickname
                        AS team_b_nickname

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s

                FOR UPDATE OF s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            # =========================
            # 플레이오프 확인
            # =========================

            if (
                series["series_type"]
                != "플레이오프"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "플레이오프 SERIES만 "
                        "다음 단계로 "
                        "진출시킬 수 있습니다."
                    ),
                )


            # =========================
            # 완료 상태 확인
            # =========================

            if (
                series["status"]
                != "completed"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "완료된 플레이오프만 "
                        "다음 단계로 "
                        "진출시킬 수 있습니다."
                    ),
                )


            # =========================
            # 결승은 다음 단계 없음
            # =========================

            if (
                series["playoff_stage"]
                == "결승시리즈"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "결승시리즈는 "
                        "다음 단계가 없습니다."
                    ),
                )


            # =========================
            # 현재 SERIES 세트 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    set_number,
                    winner_side

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )


            saved_sets = cursor.fetchall()


            if not saved_sets:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "플레이오프 세트 결과가 "
                        "없습니다."
                    ),
                )


            # =========================
            # SERIES 승자 계산
            #
            # 선승 도달 이후 세트가
            # 존재하는지도 검증
            # =========================

            team_a_wins = 0
            team_b_wins = 0

            series_winner_side = None
            winning_set_number = None


            for saved_set in saved_sets:

                winner_side = (
                    saved_set[
                        "winner_side"
                    ]
                )


                if winner_side not in (
                    "team_a",
                    "team_b",
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "플레이오프 세트의 "
                            "승패 정보가 "
                            "올바르지 않습니다."
                        ),
                    )


                # 이미 선승 도달 후인데
                # 추가 세트가 존재하면 비정상
                if (
                    series_winner_side
                    is not None
                ):

                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "선승 도달 이후의 "
                            "추가 세트가 존재합니다."
                        ),
                    )


                if (
                    winner_side
                    == "team_a"
                ):

                    team_a_wins += 1

                else:

                    team_b_wins += 1


                if (
                    team_a_wins
                    >=
                    int(
                        series[
                            "wins_required"
                        ]
                    )
                ):

                    series_winner_side = (
                        "team_a"
                    )

                    winning_set_number = (
                        saved_set[
                            "set_number"
                        ]
                    )


                elif (
                    team_b_wins
                    >=
                    int(
                        series[
                            "wins_required"
                        ]
                    )
                ):

                    series_winner_side = (
                        "team_b"
                    )

                    winning_set_number = (
                        saved_set[
                            "set_number"
                        ]
                    )


            # =========================
            # 선승 미도달 방어
            # =========================

            if series_winner_side is None:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "SERIES 승자가 "
                        "확정되지 않았습니다."
                    ),
                )


            # =========================
            # 최대 세트 수 방어
            # =========================

            if (
                len(saved_sets)
                >
                int(
                    series["best_of"]
                )
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "허용된 경기 수보다 "
                        "많은 세트가 "
                        "저장되어 있습니다."
                    ),
                )


            # =========================
            # 현재 SERIES 승자
            # =========================

            if (
                series_winner_side
                == "team_a"
            ):

                winner_id = (
                    series[
                        "team_a_id"
                    ]
                )

                winner_name = (
                    series[
                        "team_a_name"
                    ]
                )

                winner_nickname = (
                    series[
                        "team_a_nickname"
                    ]
                )

            else:

                winner_id = (
                    series[
                        "team_b_id"
                    ]
                )

                winner_name = (
                    series[
                        "team_b_name"
                    ]
                )

                winner_nickname = (
                    series[
                        "team_b_nickname"
                    ]
                )


            # =========================
            # 다음 단계 설정
            # =========================

            if (
                series["playoff_stage"]
                == "준플레이오프"
            ):

                next_stage = (
                    "플레이오프"
                )

                seeded_rank = 2

                next_best_of = 5
                next_wins_required = 3


            elif (
                series["playoff_stage"]
                == "플레이오프"
            ):

                next_stage = (
                    "결승시리즈"
                )

                seeded_rank = 1

                next_best_of = 7
                next_wins_required = 4


            else:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "알 수 없는 "
                        "플레이오프 단계입니다."
                    ),
                )


            # =========================
            # 이미 다음 단계가
            # 생성되어 있는지 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    status

                FROM series

                WHERE
                    series_type =
                        '플레이오프'

                    AND
                    playoff_stage = %s

                    AND
                    status <> 'cancelled'

                LIMIT 1
                """,
                (
                    next_stage,
                ),
            )


            existing_next_series = (
                cursor.fetchone()
            )


            if existing_next_series:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"이미 {next_stage} "
                        "SERIES가 "
                        "생성되어 있습니다."
                    ),
                )


            # =========================
            # 1위 또는 2위 확인
            # =========================

            seeded_name = (
                standings[
                    seeded_rank - 1
                ][
                    "name"
                ]
            )


            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname

                FROM participants

                WHERE fcl_name = %s
                """,
                (
                    seeded_name,
                ),
            )


            seeded_participant = (
                cursor.fetchone()
            )


            if not seeded_participant:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "시드 참가자 정보를 "
                        "찾을 수 없습니다."
                    ),
                )


            if (
                seeded_participant["id"]
                == winner_id
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "다음 플레이오프의 "
                        "두 참가자가 같습니다."
                    ),
                )


            # =========================
            # 다음 SERIES 생성
            #
            # team_a = 상위 시드
            # team_b = 이전 단계 승자
            # =========================

            cursor.execute(
                """
                INSERT INTO series (
                    series_type,

                    team_a_id,
                    team_b_id,

                    match_type,
                    scheduled_date,

                    playoff_stage,
                    best_of,
                    wins_required,

                    stats_sync_status,
                    status
                )

                VALUES (
                    '플레이오프',

                    %s,
                    %s,

                    40,
                    %s,

                    %s,
                    %s,
                    %s,

                    'pending',
                    'scheduled'
                )

                RETURNING
                    id,
                    scheduled_date,
                    playoff_stage,
                    best_of,
                    wins_required,
                    status
                """,
                (
                    seeded_participant[
                        "id"
                    ],

                    winner_id,

                    scheduled_date,

                    next_stage,
                    next_best_of,
                    next_wins_required,
                ),
            )


            next_series = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "source_series": {
            "series_id":
                series_id,

            "playoff_stage":
                series[
                    "playoff_stage"
                ],

            "winner": {
                "side":
                    series_winner_side,

                "fcl_name":
                    winner_name,

                "nickname":
                    winner_nickname,

                "wins":
                    (
                        team_a_wins
                        if
                        series_winner_side
                        == "team_a"
                        else
                        team_b_wins
                    ),

                "winning_set":
                    winning_set_number,
            },
        },

        "next_series": {
            "series_id":
                next_series[
                    "id"
                ],

            "playoff_stage":
                next_series[
                    "playoff_stage"
                ],

            "scheduled_date":
                next_series[
                    "scheduled_date"
                ].isoformat(),

            "team_a": {
                "rank":
                    seeded_rank,

                "fcl_name":
                    seeded_participant[
                        "fcl_name"
                    ],

                "nickname":
                    seeded_participant[
                        "fc_nickname"
                    ],
            },

            "team_b": {
                "source":
                    (
                        series[
                            "playoff_stage"
                        ]
                        + " 승자"
                    ),

                "fcl_name":
                    winner_name,

                "nickname":
                    winner_nickname,
            },

            "best_of":
                next_series[
                    "best_of"
                ],

            "wins_required":
                next_series[
                    "wins_required"
                ],

            "status":
                next_series[
                    "status"
                ],
        },

        "message":
            (
                f"{next_stage} SERIES가 "
                "생성되었습니다."
            ),
    }


# =========================
# 선수 기록 동기화 상태
# 정규리그만 확인
# =========================

@app.get("/api/player-rankings/sync-status")
def get_player_rankings_sync_status():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    COUNT(*) FILTER (
                        WHERE
                            stats_sync_status =
                                'pending'
                    )
                        AS pending_count,

                    COUNT(*) FILTER (
                        WHERE
                            stats_sync_status =
                                'conflict'
                    )
                        AS conflict_count

                FROM series

                WHERE
                    status = 'completed'
                    AND
                    series_type = '정규리그'
                """
            )

            row = cursor.fetchone()


    pending_count = int(
        row["pending_count"]
        or 0
    )

    conflict_count = int(
        row["conflict_count"]
        or 0
    )


    return {
        "is_syncing":
            pending_count > 0,

        "pending_count":
            pending_count,

        "conflict_count":
            conflict_count,
    }

# =========================
# 선수 득점 순위
# Neon PostgreSQL
# 정규리그만 집계
# =========================

@app.get("/api/player-rankings")
def get_player_rankings():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                WITH base_player_rows AS (

                    SELECT
                        sssp.participant_id,

                        TRIM(
                            sssp.player_name
                        )
                            AS player_name,

                        sssp.sp_id,

                        sssp.image_url,

                        ROW_NUMBER() OVER (
                            PARTITION BY
                                sssp.participant_id,
                                TRIM(
                                    sssp.player_name
                                )

                            ORDER BY
                                sssp.created_at DESC,
                                sssp.id DESC
                        )
                            AS row_number

                    FROM series_set_squad_players
                        AS sssp

                    JOIN series_sets AS ss
                        ON ss.id =
                            sssp.series_set_id

                    JOIN series AS s
                        ON s.id =
                            ss.series_id

                    WHERE
                        s.series_type =
                            '프리시즌'

                        AND
                        s.status =
                            'completed'

                        AND
                        sssp.sp_position
                            BETWEEN 1 AND 27
                ),

                base_players AS (

                    SELECT
                        participant_id,
                        player_name,
                        sp_id,
                        image_url

                    FROM base_player_rows

                    WHERE
                        row_number = 1
                ),

                regular_stats AS (

                    SELECT
                        sps.participant_id,

                        TRIM(
                            sps.player_name
                        )
                            AS player_name,

                        (
                            ARRAY_AGG(
                                sps.sp_id

                                ORDER BY
                                    s.completed_at DESC
                                        NULLS LAST,
                                    s.id DESC
                            )
                        )[1]
                            AS sp_id,

                        (
                            ARRAY_AGG(
                                sps.image_url

                                ORDER BY
                                    s.completed_at DESC
                                        NULLS LAST,
                                    s.id DESC
                            )
                        )[1]
                            AS image_url,

                        SUM(
                            sps.sets_played
                        )
                            AS sets_played,

                        SUM(
                            sps.rating_total
                        )
                            AS rating_total,

                        SUM(
                            sps.goals
                        )
                            AS goals,

                        SUM(
                            sps.assists
                        )
                            AS assists

                    FROM series_player_stats
                        AS sps

                    JOIN series AS s
                        ON s.id =
                            sps.series_id

                    WHERE
                        s.status =
                            'completed'

                        AND
                        s.series_type =
                            '정규리그'

                    GROUP BY
                        sps.participant_id,
                        TRIM(
                            sps.player_name
                        )
                ),

                regular_mvp AS (

                    SELECT
                        sm.participant_id,

                        TRIM(
                            sm.player_name
                        )
                            AS player_name,

                        COUNT(*)
                            AS mvp_count

                    FROM series_mvp AS sm

                    JOIN series AS s
                        ON s.id =
                            sm.series_id

                    WHERE
                        s.status =
                            'completed'

                        AND
                        s.series_type =
                            '정규리그'

                    GROUP BY
                        sm.participant_id,
                        TRIM(
                            sm.player_name
                        )
                ),


                all_players AS (

                    SELECT
                        participant_id,
                        player_name

                    FROM base_players

                    UNION

                    SELECT
                        participant_id,
                        player_name

                    FROM regular_stats
                )

                SELECT
                    p.id
                        AS participant_id,

                    p.fcl_name,

                    p.fc_nickname
                        AS nickname,

                    ap.player_name,

                    COALESCE(
                        rs.sp_id,
                        bp.sp_id
                    )
                        AS sp_id,

                    COALESCE(
                        rs.image_url,
                        bp.image_url
                    )
                        AS image_url,

                    COALESCE(
                        rs.sets_played,
                        0
                    )
                        AS sets_played,

                    COALESCE(
                        rs.rating_total,
                        0
                    )
                        AS rating_total,

                    COALESCE(
                        rs.goals,
                        0
                    )
                        AS goals,

                    COALESCE(
                        rs.assists,
                        0
                    )
                        AS assists,

                    COALESCE(
                        rm.mvp_count,
                        0
                    )
                        AS mvp_count

                FROM all_players AS ap

                JOIN participants AS p
                    ON p.id =
                        ap.participant_id

                LEFT JOIN base_players AS bp
                    ON bp.participant_id =
                        ap.participant_id

                    AND
                    bp.player_name =
                        ap.player_name

                LEFT JOIN regular_stats AS rs
                    ON rs.participant_id =
                        ap.participant_id

                    AND
                    rs.player_name =
                        ap.player_name

                LEFT JOIN regular_mvp AS rm
                    ON rm.participant_id =
                        ap.participant_id

                    AND
                    rm.player_name =
                        ap.player_name

                ORDER BY
                    COALESCE(
                        rs.goals,
                        0
                    ) DESC,

                    COALESCE(
                        rs.assists,
                        0
                    ) DESC,

                    COALESCE(
                        rs.rating_total,
                        0
                    ) DESC,

                    ap.player_name ASC,

                    p.id ASC
                """
            )


            rows = cursor.fetchall()


    players = []


    for row in rows:

        sets_played = int(
            row["sets_played"]
        )


        rating_total = float(
            row["rating_total"]
        )


        goals = int(
            row["goals"]
        )


        assists = int(
            row["assists"]
        )

        mvp_count = int(
            row["mvp_count"]
        )


        if sets_played > 0:

            average_rating = round(
                rating_total
                / sets_played,
                2,
            )

        else:

            average_rating = 0


        players.append(
            {
                "fcl_name":
                    row["fcl_name"],

                "nickname":
                    row["nickname"],

                "player_name":
                    row["player_name"],

                "sp_id":
                    row["sp_id"],

                "image_url":
                    row["image_url"],

                "sets_played":
                    sets_played,

                "rating_total":
                    round(
                        rating_total,
                        2,
                    ),

                "average_rating":
                    average_rating,

                "goals":
                    goals,

                "assists":
                    assists,

                "mvp_count":
                    mvp_count,
            }
        )


    # =========================
    # 득점 공동 순위
    # =========================

    previous_goals = None
    previous_rank = 0


    for index, player in enumerate(
        players,
        start=1,
    ):

        if (
            previous_goals is None
            or
            player["goals"]
            != previous_goals
        ):

            previous_rank = index


        player["rank"] = (
            previous_rank
        )


        previous_goals = (
            player["goals"]
        )


    return players


def get_playoff_schedule_date(
    playoff_stage: str,
):

    workbook = load_workbook(
        PLAYOFFS_PATH,
        data_only=True,
    )

    worksheet = workbook[
        "플레이오프"
    ]


    try:

        for row in worksheet.iter_rows(
            min_row=2,
            max_col=4,
            values_only=True,
        ):

            (
                match_date,
                stage,
                _,
                _,
            ) = row


            if (
                match_date is None
                or
                stage is None
            ):
                continue


            stage_text = (
                str(stage).strip()
            )


            if (
                stage_text
                == "결승 시리즈"
            ):
                stage_text = (
                    "결승시리즈"
                )


            if (
                stage_text
                != playoff_stage
            ):
                continue


            if isinstance(
                match_date,
                datetime,
            ):
                return (
                    match_date.date()
                )


            if hasattr(
                match_date,
                "year",
            ):
                return match_date


            return datetime.strptime(
                str(match_date).strip(),
                "%Y-%m-%d",
            ).date()


    finally:

        workbook.close()


    return None

def create_initial_playoff_if_ready():

    # =========================
    # 정규리그 완료 여부
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    COUNT(*) AS total_count,

                    COUNT(*) FILTER (
                        WHERE status = 'completed'
                    ) AS completed_count

                FROM series

                WHERE
                    series_type =
                        '정규리그'

                    AND
                    status <>
                        'cancelled'
                """
            )


            regular_status = (
                cursor.fetchone()
            )


    total_count = int(
        regular_status[
            "total_count"
        ]
    )

    completed_count = int(
        regular_status[
            "completed_count"
        ]
    )


    # 아직 정규리그가 안 끝남
    if (
        total_count != 20
        or
        completed_count != 20
    ):

        return None


    # =========================
    # 이미 준PO가 있는지 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id

                FROM series

                WHERE
                    series_type =
                        '플레이오프'

                    AND
                    playoff_stage =
                        '준플레이오프'

                    AND
                    status <>
                        'cancelled'

                LIMIT 1
                """
            )


            existing_series = (
                cursor.fetchone()
            )


    if existing_series:

        return {
            "created": False,
            "series_id":
                existing_series["id"],
        }


    # =========================
    # 최종 순위
    # =========================

    standings = get_standings()


    if len(standings) < 4:

        return None


    third_place = standings[2]
    fourth_place = standings[3]


    # =========================
    # 기존 플레이오프 일정에서
    # 준PO 날짜 가져오기
    # =========================

    scheduled_date = (
        get_playoff_schedule_date(
            "준플레이오프"
        )
    )


    if scheduled_date is None:

        raise RuntimeError(
            "준플레이오프 일정 날짜를 "
            "찾을 수 없습니다."
        )


    # =========================
    # 참가자 조회 + SERIES 생성
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # 중복 생성 방지
            cursor.execute(
                """
                SELECT
                    pg_advisory_xact_lock(
                        20261010
                    )
                """
            )


            # lock을 기다리는 동안
            # 다른 요청이 먼저 만들었을 수 있으므로
            # 다시 확인
            cursor.execute(
                """
                SELECT
                    id

                FROM series

                WHERE
                    series_type =
                        '플레이오프'

                    AND
                    playoff_stage =
                        '준플레이오프'

                    AND
                    status <>
                        'cancelled'

                LIMIT 1
                """
            )


            existing_series = (
                cursor.fetchone()
            )


            if existing_series:

                return {
                    "created": False,
                    "series_id":
                        existing_series[
                            "id"
                        ],
                }


            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname

                FROM participants

                WHERE fcl_name IN (
                    %s,
                    %s
                )
                """,
                (
                    third_place["name"],
                    fourth_place["name"],
                ),
            )


            participant_rows = (
                cursor.fetchall()
            )


            participant_map = {
                participant[
                    "fcl_name"
                ]:
                    participant

                for participant
                in participant_rows
            }


            if (
                third_place["name"]
                not in participant_map

                or

                fourth_place["name"]
                not in participant_map
            ):

                raise RuntimeError(
                    "플레이오프 참가자 정보를 "
                    "찾을 수 없습니다."
                )


            team_a = participant_map[
                third_place["name"]
            ]

            team_b = participant_map[
                fourth_place["name"]
            ]


            cursor.execute(
                """
                INSERT INTO series (
                    series_type,
                    team_a_id,
                    team_b_id,
                    match_type,
                    scheduled_date,
                    playoff_stage,
                    best_of,
                    wins_required,
                    stats_sync_status,
                    status
                )

                VALUES (
                    '플레이오프',
                    %s,
                    %s,
                    40,
                    %s,
                    '준플레이오프',
                    5,
                    3,
                    'pending',
                    'scheduled'
                )

                RETURNING
                    id,
                    scheduled_date
                """,
                (
                    team_a["id"],
                    team_b["id"],
                    scheduled_date,
                ),
            )


            playoff_series = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "created": True,

        "series_id":
            playoff_series["id"],

        "scheduled_date":
            playoff_series[
                "scheduled_date"
            ].isoformat(),

        "team_a":
            team_a["fcl_name"],

        "team_b":
            team_b["fcl_name"],
    }

def create_next_playoff_if_ready(
    series_id: int,
):

    # =========================
    # 현재 플레이오프 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    playoff_stage,
                    status

                FROM series

                WHERE id = %s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


    if not series:
        return None


    if (
        series["series_type"]
        != "플레이오프"
    ):
        return None


    if (
        series["status"]
        != "completed"
    ):
        return None


    # =========================
    # 다음 단계
    # =========================

    if (
        series["playoff_stage"]
        == "준플레이오프"
    ):

        next_stage = "플레이오프"

    elif (
        series["playoff_stage"]
        == "플레이오프"
    ):

        next_stage = "결승시리즈"

    elif (
        series["playoff_stage"]
        == "결승시리즈"
    ):

        return {
            "created": False,
            "reason": "final_completed",
        }

    else:

        return None


    # =========================
    # 이미 다음 SERIES가 있는지
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id

                FROM series

                WHERE
                    series_type =
                        '플레이오프'

                    AND
                    playoff_stage = %s

                    AND
                    status <>
                        'cancelled'

                LIMIT 1
                """,
                (
                    next_stage,
                ),
            )


            existing_series = (
                cursor.fetchone()
            )


    if existing_series:

        return {
            "created": False,
            "series_id":
                existing_series["id"],
        }


    # =========================
    # 기존 플레이오프 일정 날짜
    # =========================

    scheduled_date = (
        get_playoff_schedule_date(
            next_stage
        )
    )


    if scheduled_date is None:

        raise RuntimeError(
            f"{next_stage} 일정 날짜를 "
            "찾을 수 없습니다."
        )


    # =========================
    # 기존 검증된 ADVANCE 로직 재사용
    # =========================

    request = PlayoffAdvanceRequest(
        scheduled_date=
            scheduled_date.isoformat()
    )


    try:

        result = advance_playoff_series(
            series_id,
            request,
            admin_token="internal",
        )

    except HTTPException:

        # 동시에 두 요청이 들어온 경우
        # 먼저 생성된 다음 SERIES 확인
        with get_db_connection() as connection:

            with connection.cursor() as cursor:

                cursor.execute(
                    """
                    SELECT
                        id

                    FROM series

                    WHERE
                        series_type =
                            '플레이오프'

                        AND
                        playoff_stage = %s

                        AND
                        status <>
                            'cancelled'

                    LIMIT 1
                    """,
                    (
                        next_stage,
                    ),
                )


                existing_series = (
                    cursor.fetchone()
                )


        if existing_series:

            return {
                "created": False,
                "series_id":
                    existing_series["id"],
            }


        raise


    return {
        "created": True,

        "series_id":
            result[
                "next_series"
            ][
                "series_id"
            ],

        "playoff_stage":
            next_stage,

        "scheduled_date":
            scheduled_date.isoformat(),
    }

# =========================
# 플레이오프 일정
#
# 날짜/단계:
# playoffs.xlsx
#
# 실제 대진/진행 상태:
# Neon SERIES
# =========================

@app.get("/api/playoffs")
def get_playoffs():

    # =========================
    # 1. 기존 플레이오프 일정
    # =========================

    workbook = load_workbook(
        PLAYOFFS_PATH,
        data_only=True,
    )

    worksheet = workbook[
        "플레이오프"
    ]

    schedule_rows = []


    for row in worksheet.iter_rows(
        min_row=2,
        max_col=4,
        values_only=True,
    ):

        (
            match_date,
            stage,
            team_a,
            team_b,
        ) = row


        if match_date is None:
            continue


        if hasattr(
            match_date,
            "strftime",
        ):

            match_date = (
                match_date.strftime(
                    "%Y-%m-%d"
                )
            )


        stage_text = (
            str(stage).strip()
            if stage is not None
            else ""
        )


        # DB에서는
        # "결승시리즈"로 저장
        if stage_text == "결승 시리즈":

            database_stage = (
                "결승시리즈"
            )

        else:

            database_stage = (
                stage_text
            )


        schedule_rows.append(
            {
                "date":
                    str(match_date),

                "stage":
                    stage_text,

                "database_stage":
                    database_stage,

                "default_team_a":
                    team_a,

                "default_team_b":
                    team_b,
            }
        )


    workbook.close()


    # =========================
    # 2. 실제 플레이오프 SERIES
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,
                    s.status,
                    s.scheduled_date,

                    team_a.fcl_name
                        AS team_a_name,

                    team_b.fcl_name
                        AS team_b_name,

                    CASE
                        WHEN
                            s.status = 'completed'
                        THEN
                            COALESCE(
                                s.team_a_snapshot_logo_path,
                                team_a.current_team_logo_path
                            )
                        ELSE
                            team_a.current_team_logo_path
                    END
                        AS team_a_logo_path,

                    CASE
                        WHEN
                            s.status = 'completed'
                        THEN
                            COALESCE(
                                s.team_b_snapshot_logo_path,
                                team_b.current_team_logo_path
                            )
                        ELSE
                            team_b.current_team_logo_path
                    END
                        AS team_b_logo_path,

                    (
                        SELECT COUNT(*)
                        FROM series_sets AS ss
                        WHERE
                            ss.series_id = s.id
                    )
                        AS set_count,

                    (
                        SELECT COUNT(*)
                        FROM series_sets AS ss
                        WHERE
                            ss.series_id = s.id
                            AND
                            ss.winner_side = 'team_a'
                    )
                        AS team_a_wins,

                    (
                        SELECT COUNT(*)
                        FROM series_sets AS ss
                        WHERE
                            ss.series_id = s.id
                            AND
                            ss.winner_side = 'team_b'
                    )
                        AS team_b_wins

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE
                    s.series_type =
                        '플레이오프'

                    AND
                    s.status <>
                        'cancelled'

                ORDER BY
                    s.id DESC
                """
            )


            series_rows = (
                cursor.fetchall()
            )


    # =========================
    # 단계별 최신 SERIES
    # =========================

    series_by_stage = {}


    for series_row in series_rows:

        playoff_stage = (
            series_row[
                "playoff_stage"
            ]
        )


        if (
            playoff_stage
            not in series_by_stage
        ):

            series_by_stage[
                playoff_stage
            ] = series_row


    # =========================
    # 3. 일정 + SERIES 결합
    # =========================

    playoffs = []


    for schedule_row in schedule_rows:

        database_stage = (
            schedule_row[
                "database_stage"
            ]
        )


        series_row = (
            series_by_stage.get(
                database_stage
            )
        )


        if database_stage in (
            "준플레이오프",
            "플레이오프",
        ):

            default_best_of = 5
            default_wins_required = 3

        elif (
            database_stage
            == "결승시리즈"
        ):

            default_best_of = 7
            default_wins_required = 4

        else:

            default_best_of = None
            default_wins_required = None


        # =====================
        # 아직 SERIES 없음
        # =====================

        if series_row is None:

            playoffs.append(
                {
                    "date":
                        schedule_row[
                            "date"
                        ],

                    "stage":
                        schedule_row[
                            "stage"
                        ],

                    "playoff_stage":
                        database_stage,

                    "series_id":
                        None,

                    "status":
                        "waiting",

                    "best_of":
                        default_best_of,

                    "wins_required":
                        default_wins_required,

                    "team_a":
                        (
                            schedule_row[
                                "default_team_a"
                            ]
                            or
                            "TBD"
                        ),

                    "team_b":
                        (
                            schedule_row[
                                "default_team_b"
                            ]
                            or
                            "TBD"
                        ),

                    "team_a_logo_path":
                        None,

                    "team_b_logo_path":
                        None,

                    "set_count":
                        0,

                    "team_a_wins":
                        0,

                    "team_b_wins":
                        0,
                }
            )

            continue


        # =====================
        # 실제 SERIES 존재
        # =====================

        team_a_wins = int(
            series_row[
                "team_a_wins"
            ]
        )

        team_b_wins = int(
            series_row[
                "team_b_wins"
            ]
        )


        winner = None


        if (
            series_row["status"]
            == "completed"
        ):

            if (
                team_a_wins
                >=
                int(
                    series_row[
                        "wins_required"
                    ]
                )
            ):

                winner = (
                    series_row[
                        "team_a_name"
                    ]
                )

            elif (
                team_b_wins
                >=
                int(
                    series_row[
                        "wins_required"
                    ]
                )
            ):

                winner = (
                    series_row[
                        "team_b_name"
                    ]
                )


        playoffs.append(
            {
                # 플레이오프 날짜는
                # 기존 일정표를 기준으로 사용
                "date":
                    schedule_row[
                        "date"
                    ],

                "stage":
                    schedule_row[
                        "stage"
                    ],

                "playoff_stage":
                    database_stage,

                "series_id":
                    series_row["id"],

                "status":
                    series_row[
                        "status"
                    ],

                "best_of":
                    series_row[
                        "best_of"
                    ],

                "wins_required":
                    series_row[
                        "wins_required"
                    ],

                "team_a":
                    series_row[
                        "team_a_name"
                    ],

                "team_b":
                    series_row[
                        "team_b_name"
                    ],

                "team_a_logo_path":
                    series_row[
                        "team_a_logo_path"
                    ],

                "team_b_logo_path":
                    series_row[
                        "team_b_logo_path"
                    ],

                "set_count":
                    int(
                        series_row[
                            "set_count"
                        ]
                    ),

                "team_a_wins":
                    team_a_wins,

                "team_b_wins":
                    team_b_wins,

                "winner":
                    winner,
            }
        )


    return playoffs


# =========================
# FC Online
# OUID 테스트
# =========================

@app.get(
    "/api/fconline/ouid/{nickname}"
)
def get_fconline_ouid(
    nickname: str,
):

    url = (
        f"{NEXON_API_BASE_URL}/id"
    )

    try:

        with httpx.Client(
            timeout=10.0
        ) as client:

            response = client.get(
                url,
                headers=get_nexon_headers(),
                params={
                    "nickname": nickname,
                },
            )

    except httpx.RequestError as error:

        raise HTTPException(
            status_code=502,
            detail=(
                "NEXON Open API 연결에 "
                "실패했습니다."
            ),
        ) from error


    if response.status_code != 200:

        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )


    return response.json()

# =========================
# FC Online 최근 매치 조회
# =========================

@app.get(
    "/api/fconline/matches/{nickname}"
)
def get_fconline_matches(
    nickname: str,
    matchtype: int,
    limit: int = 20,
):

    # 1. 닉네임 -> OUID
    ouid_response = httpx.get(
        f"{NEXON_API_BASE_URL}/id",
        headers=get_nexon_headers(),
        params={
            "nickname": nickname,
        },
        timeout=10.0,
    )

    if ouid_response.status_code != 200:
        raise HTTPException(
            status_code=ouid_response.status_code,
            detail=ouid_response.text,
        )

    ouid = ouid_response.json()["ouid"]


    # 2. OUID -> 최근 매치 ID
    match_response = httpx.get(
        f"{NEXON_API_BASE_URL}/user/match",
        headers=get_nexon_headers(),
        params={
            "ouid": ouid,
            "matchtype": matchtype,
            "offset": 0,
            "limit": limit,
        },
        timeout=10.0,
    )

    if match_response.status_code != 200:
        raise HTTPException(
            status_code=match_response.status_code,
            detail=match_response.text,
        )


    return {
        "nickname": nickname,
        "ouid": ouid,
        "matches": match_response.json(),
    }


# =========================
# FC Online 매치 종류
# =========================

@app.get("/api/fconline/match-types")
def get_fconline_match_types():

    url = (
        "https://open.api.nexon.com/"
        "static/fconline/meta/matchtype.json"
    )

    try:
        with httpx.Client(
            timeout=10.0
        ) as client:

            response = client.get(url)

    except httpx.RequestError as error:
        raise HTTPException(
            status_code=502,
            detail="매치 종류 정보를 불러오지 못했습니다.",
        ) from error


    if response.status_code != 200:
        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )


    return response.json()

# =========================
# FC Online 맞대결 탐색 테스트
# =========================

@app.get("/api/fconline/head-to-head-test")
def get_head_to_head_test():

    lee_nickname = "똭똭"
    seo_nickname = "붉은심장베컴"

    lee_ouid = get_ouid_by_nickname(
        lee_nickname
    )

    seo_ouid = get_ouid_by_nickname(
        seo_nickname
    )


    # 일반 1vs1에서 사용할 가능성이 있는
    # 매치 종류를 전부 검색
    match_types = {
        30: "리그 친선",
        40: "클래식 1on1",
        50: "공식경기",
        60: "공식 친선",
    }


    head_to_head_matches = []


    for match_type, description in (
        match_types.items()
    ):

        lee_matches = get_user_match_ids(
            lee_ouid,
            match_type,
        )

        seo_matches = get_user_match_ids(
            seo_ouid,
            match_type,
        )


        seo_match_set = set(
            seo_matches
        )


        # 똭똭 기록 순서를 유지하면서
        # 두 사람에게 공통으로 존재하는 matchId
        common_matches = [
            match_id
            for match_id in lee_matches
            if match_id in seo_match_set
        ]


        if common_matches:

            head_to_head_matches.append(
                {
                    "match_type": match_type,
                    "description": description,
                    "match_ids": common_matches,
                }
            )


    return {
        "team_a": {
            "fcl_name": "이준석",
            "nickname": lee_nickname,
        },

        "team_b": {
            "fcl_name": "서종원",
            "nickname": seo_nickname,
        },

        "matches": head_to_head_matches,
    }

# =========================
# FC Online 매치 상세 테스트
# =========================

@app.get(
    "/api/fconline/match-detail-test/{match_id}"
)
def get_match_detail_test(
    match_id: str,
):

    response = httpx.get(
        f"{NEXON_API_BASE_URL}/match-detail",
        headers=get_nexon_headers(),
        params={
            "matchid": match_id,
        },
        timeout=10.0,
    )


    if response.status_code != 200:

        raise HTTPException(
            status_code=response.status_code,
            detail=response.text,
        )


    return response.json()

# =========================
# FCL 3세트 분석 테스트
# =========================

@app.get("/api/fconline/series-test")
def get_series_test():

    match_ids = [
        "6a81cdffa962a502d85e1eaa",
        "6a81cb89bdd2ff3b3f6807a8",
        "6a81c8f6115d94f2dc8ca43c",
    ]

    series_matches = []


    for match_id in match_ids:

        response = httpx.get(
            f"{NEXON_API_BASE_URL}/match-detail",
            headers=get_nexon_headers(),
            params={
                "matchid": match_id,
            },
            timeout=10.0,
        )


        if response.status_code != 200:

            raise HTTPException(
                status_code=response.status_code,
                detail=response.text,
            )


        match_data = response.json()

        participants = []


        for match_info in match_data["matchInfo"]:

            player_data = []


            for player in match_info["player"]:

                status = player["status"]


                # 출전하지 않은 선수 제외
                if status["spRating"] <= 0:
                    continue


                player_data.append(
                    {
                        "sp_id": player["spId"],
                        "rating": status["spRating"],
                        "goal": status["goal"],
                        "assist": status["assist"],
                    }
                )


            participants.append(
                {
                    "nickname": match_info["nickname"],

                    "score": (
                        match_info["shoot"][
                            "goalTotal"
                        ]
                    ),

                    "players": player_data,
                }
            )


        series_matches.append(
            {
                "match_id": match_data["matchId"],
                "date": match_data["matchDate"],
                "match_type": match_data["matchType"],
                "participants": participants,
            }
        )


    # 실제 시간 순서
    series_matches.sort(
        key=lambda match: match["date"]
    )


    return series_matches


# =========================
# FCL SERIES START
# Neon PostgreSQL
# =========================

@app.post("/api/fconline/series/start")
def start_fcl_series(
    request: SeriesStartRequest,
):

    # =========================
    # 친선전 생성 전용
    # =========================

    if (
        request.series_type
        != "프리시즌"
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "정규리그 SERIES는 "
                "미리 등록된 일정만 "
                "사용할 수 있습니다."
            ),
        )


    # =========================
    # 기본 검증
    # =========================

    if request.team_a == request.team_b:

        raise HTTPException(
            status_code=400,
            detail=(
                "같은 참가자끼리는 "
                "SERIES를 시작할 수 없습니다."
            ),
        )


    # =========================
    # 경기 날짜
    # =========================

    if not request.scheduled_date:

        raise HTTPException(
            status_code=400,
            detail="경기 날짜를 선택해주세요.",
        )


    try:

        scheduled_date = datetime.strptime(
            request.scheduled_date,
            "%Y-%m-%d",
        ).date()

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail=(
                "경기 날짜 형식은 "
                "YYYY-MM-DD여야 합니다."
            ),
        )


    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    if scheduled_date < today:

        raise HTTPException(
            status_code=400,
            detail=(
                "지난 날짜로 친선전을 "
                "생성할 수 없습니다."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 참가자 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname,
                    ouid

                FROM participants

                WHERE fcl_name IN (
                    %s,
                    %s
                )
                """,
                (
                    request.team_a,
                    request.team_b,
                ),
            )


            participant_rows = (
                cursor.fetchall()
            )


            participant_map = {
                participant["fcl_name"]:
                    participant

                for participant
                in participant_rows
            }


            if (
                request.team_a
                not in participant_map
                or
                request.team_b
                not in participant_map
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "등록되지 않은 "
                        "FCL 참가자입니다."
                    ),
                )


            team_a = participant_map[
                request.team_a
            ]

            team_b = participant_map[
                request.team_b
            ]


            if (
                not team_a["fc_nickname"]
                or
                not team_b["fc_nickname"]
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "FC Online 닉네임이 "
                        "등록되지 않은 참가자가 있습니다."
                    ),
                )


            # =========================
            # 동일 친선전 예약 중복 방지
            # 같은 날짜 + 같은 대진
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    status

                FROM series

                WHERE
                    series_type = '프리시즌'

                    AND
                    scheduled_date = %s

                    AND
                    status IN (
                        'scheduled',
                        'active'
                    )

                    AND (
                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )

                        OR

                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )
                    )

                LIMIT 1
                """,
                (
                    scheduled_date,

                    team_a["id"],
                    team_b["id"],

                    team_b["id"],
                    team_a["id"],
                ),
            )


            existing_preseason_series = (
                cursor.fetchone()
            )


            if existing_preseason_series:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "해당 날짜에 동일한 "
                        "친선전이 이미 예약되어 있습니다."
                    ),
                )


            # =========================
            # 이미 ACTIVE인 같은 대진 확인
            # =========================

            cursor.execute(
                """
                SELECT id

                FROM series

                WHERE status = 'active'

                AND (
                    (
                        team_a_id = %s
                        AND
                        team_b_id = %s
                    )

                    OR

                    (
                        team_a_id = %s
                        AND
                        team_b_id = %s
                    )
                )

                LIMIT 1
                """,
                (
                    team_a["id"],
                    team_b["id"],

                    team_b["id"],
                    team_a["id"],
                ),
            )


            active_series = (
                cursor.fetchone()
            )


            if active_series:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "두 참가자 사이에 "
                        "이미 진행 중인 SERIES가 있습니다."
                    ),
                )


            # =========================
            # 친선전 예약 생성
            # =========================
            cursor.execute(
                """
                INSERT INTO series (
                    series_type,

                    team_a_id,
                    team_b_id,

                    match_type,

                    scheduled_date,
                    round_number,

                    status
                )

                VALUES (
                    '프리시즌',
                    %s,
                    %s,
                    40,
                    %s,
                    NULL,
                    'scheduled'
                )

                RETURNING id
                """,
                (
                    team_a["id"],
                    team_b["id"],

                    scheduled_date,
                ),
            )


            series_row = (
                cursor.fetchone()
            )


            series_id = (
                series_row["id"]
            )


        connection.commit()


    return {
        "series_id":
            series_id,

        "series_type":
            "프리시즌",

        "team_a":
            request.team_a,

        "team_b":
            request.team_b,

        "nickname_a":
            team_a["fc_nickname"],

        "nickname_b":
            team_b["fc_nickname"],

        "match_type":
            40,

        "started_at":
            None,

        "status":
            "scheduled",

        "set_count":
            0,

        "scheduled_date":
            request.scheduled_date,

        "round_number":
            None,
    }


# =========================
# FCL HISTORY SERIES IMPORT
# 과거 친선전 실제 경기 등록
# =========================

@app.post(
    "/api/fconline/history/import"
)
def import_history_series(
    request: HistorySeriesImportRequest,
):

    # =========================
    # 기본 검증
    # =========================

    if request.team_a == request.team_b:

        raise HTTPException(
            status_code=400,
            detail=(
                "서로 다른 참가자를 "
                "선택해주세요."
            ),
        )


    try:

        target_date = datetime.strptime(
            request.match_date,
            "%Y-%m-%d",
        ).date()

    except ValueError:

        raise HTTPException(
            status_code=400,
            detail=(
                "경기 날짜 형식은 "
                "YYYY-MM-DD여야 합니다."
            ),
        )


    today = datetime.now(
        ZoneInfo("Asia/Seoul")
    ).date()


    if target_date >= today:

        raise HTTPException(
            status_code=400,
            detail=(
                "과거 경기 등록은 "
                "지난 날짜만 가능합니다."
            ),
        )


    # =========================
    # 참가자 조회
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname,
                    ouid

                FROM participants

                WHERE
                    fcl_name = %s
                    OR
                    fcl_name = %s
                """,
                (
                    request.team_a,
                    request.team_b,
                ),
            )


            participant_rows = (
                cursor.fetchall()
            )


    participant_map = {
        participant["fcl_name"]:
            participant

        for participant
        in participant_rows
    }


    team_a = participant_map.get(
        request.team_a
    )

    team_b = participant_map.get(
        request.team_b
    )


    if not team_a or not team_b:

        raise HTTPException(
            status_code=404,
            detail=(
                "참가자 정보를 "
                "찾을 수 없습니다."
            ),
        )


    nickname_a = (
        team_a["fc_nickname"]
    )

    nickname_b = (
        team_b["fc_nickname"]
    )


    if not nickname_a or not nickname_b:

        raise HTTPException(
            status_code=400,
            detail=(
                "FC Online 닉네임 "
                "정보가 없습니다."
            ),
        )


    # =========================
    # 동일 과거 SERIES 중복 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT id

                FROM series

                WHERE
                    series_type = '프리시즌'

                    AND
                    scheduled_date = %s

                    AND
                    status <> 'cancelled'

                    AND (
                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )

                        OR

                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )
                    )

                LIMIT 1
                """,
                (
                    target_date,

                    team_a["id"],
                    team_b["id"],

                    team_b["id"],
                    team_a["id"],
                ),
            )


            existing_series = (
                cursor.fetchone()
            )


    if existing_series:

        raise HTTPException(
            status_code=400,
            detail=(
                "해당 날짜의 친선전이 "
                "이미 등록되어 있습니다."
            ),
        )


    # =========================
    # OUID 준비
    # =========================

    ouid_a = get_participant_ouid(
        team_a["id"],
        nickname_a,
        team_a["ouid"],
    )


    ouid_b = get_participant_ouid(
        team_b["id"],
        nickname_b,
        team_b["ouid"],
    )


    # =========================
    # 두 참가자의 최근 경기
    # =========================

    matches_a = get_user_match_ids(
        ouid_a,
        40,
        limit=50,
    )


    matches_b = get_user_match_ids(
        ouid_b,
        40,
        limit=50,
    )


    matches_b_set = set(
        matches_b
    )


    common_match_ids = [
        match_id

        for match_id
        in matches_a

        if match_id in matches_b_set
    ]


    detected_matches = []


    # =========================
    # 날짜 + 정확한 맞대결 필터
    # =========================

    for match_id in common_match_ids:

        match_data = get_match_detail(
            match_id
        )


        if (
            match_data["matchType"]
            != 40
        ):
            continue


        played_at = (
            parse_nexon_datetime(
                match_data["matchDate"]
            )
        )


        if (
            played_at.date()
            != target_date
        ):
            continue


        match_nicknames = {
            match_info["nickname"]

            for match_info
            in match_data["matchInfo"]
        }


        if match_nicknames != {
            nickname_a,
            nickname_b,
        }:
            continue


        detected_matches.append(
            {
                "data":
                    match_data,

                "played_at":
                    played_at,
            }
        )


    # =========================
    # 시간순
    # =========================

    detected_matches.sort(
        key=lambda match:
            match["played_at"]
    )


    # =========================
    # 반드시 정확히 3경기
    # =========================

    if len(detected_matches) < 3:

        raise HTTPException(
            status_code=400,
            detail=(
                f"{request.match_date} "
                "맞대결을 "
                f"{len(detected_matches)}경기만 "
                "찾았습니다. "
                "3경기가 필요합니다."
            ),
        )


    # 시간순 첫 3경기만 SERIES로 사용
    detected_matches = (
        detected_matches[:3]
    )


    match_ids = [
        detected_match[
            "data"
        ][
            "matchId"
        ]



        for detected_match
        in detected_matches
    ]


    # =========================
    # 이미 사용된 실제 경기 확인
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    nexon_match_id

                FROM series_sets

                WHERE nexon_match_id
                    IN (%s, %s, %s)

                LIMIT 1
                """,
                (
                    match_ids[0],
                    match_ids[1],
                    match_ids[2],
                ),
            )


            used_match = (
                cursor.fetchone()
            )


    if used_match:

        raise HTTPException(
            status_code=400,
            detail=(
                "해당 FC Online 경기가 "
                "이미 다른 SERIES에 "
                "등록되어 있습니다."
            ),
        )


    # =========================
    # MVP + 전체 선수 기록 계산
    # DB 저장 전에 먼저 계산
    # =========================

    match_data_list = [
        detected_match["data"]

        for detected_match
        in detected_matches
    ]


    (
        mvp,
        _,
        player_stats,
    ) = calculate_series_mvp_from_matches(
        match_data_list
    )


    started_at = (
        detected_matches[0][
            "played_at"
        ]
    )

    completed_at = (
        detected_matches[-1][
            "played_at"
        ]
    )


    # =========================
    # SERIES + 3 SET 저장
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO series (
                    series_type,

                    team_a_id,
                    team_b_id,

                    match_type,

                    scheduled_date,
                    round_number,

                    started_at,
                    completed_at,
                    finished_at,

                    status,
                    stats_sync_status
                )

                VALUES (
                    '프리시즌',

                    %s,
                    %s,

                    40,

                    %s,
                    NULL,

                    %s,
                    %s,
                    %s,

                    'completed',
                    'synced'
                )

                RETURNING id
                """,
                (
                    team_a["id"],
                    team_b["id"],

                    target_date,

                    started_at,
                    completed_at,
                    completed_at,
                ),
            )


            series_row = (
                cursor.fetchone()
            )

            series_id = (
                series_row["id"]
            )


            for (
                set_number,
                detected_match,
            ) in enumerate(
                detected_matches,
                start=1,
            ):

                match_data = (
                    detected_match["data"]
                )

                played_at = (
                    detected_match[
                        "played_at"
                    ]
                )


                participant_info_map = {
                    match_info["nickname"]:
                        match_info

                    for match_info
                    in match_data["matchInfo"]
                }


                team_a_info = (
                    participant_info_map[
                        nickname_a
                    ]
                )

                team_b_info = (
                    participant_info_map[
                        nickname_b
                    ]
                )


                team_a_score = (
                    team_a_info[
                        "shoot"
                    ][
                        "goalTotal"
                    ]
                )

                team_b_score = (
                    team_b_info[
                        "shoot"
                    ][
                        "goalTotal"
                    ]
                )


        cursor.execute(
            """
            INSERT INTO series_sets (
                series_id,

                set_number,

                nexon_match_id,
                played_at,

                team_a_score,
                team_b_score,

                score_source
            )

            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                'nexon'
            )
            """,
            (
                series_id,
                set_number,

                match_data[
                    "matchId"
                ],

                played_at,

                team_a_score,
                team_b_score,
            ),
        )


        connection.commit()

    # =========================
    # 세트별 스쿼드 Snapshot 저장
    # =========================

    save_series_set_squad_players(
        series_id,

        team_a["id"],
        nickname_a,

        team_b["id"],
        nickname_b,

        detected_matches,
    )


    # =========================
    # 전체 선수 기록 저장
    # =========================

    save_series_player_stats(
        series_id,

        team_a["id"],
        nickname_a,

        team_b["id"],
        nickname_b,

        player_stats,
    )


    # =========================
    # MVP 저장
    # =========================

    if mvp:

        if (
            mvp["nickname"]
            == nickname_a
        ):

            mvp_participant_id = (
                team_a["id"]
            )

        elif (
            mvp["nickname"]
            == nickname_b
        ):

            mvp_participant_id = (
                team_b["id"]
            )

        else:

            mvp_participant_id = None


        if mvp_participant_id:

            save_series_mvp(
                series_id,
                mvp_participant_id,
                mvp,
            )


    # =========================
    # 결과 응답
    # =========================

    return {
        "series_id":
            series_id,

        "series_type":
            "프리시즌",

        "status":
            "completed",

        "stats_sync_status":
            "synced",

        "match_date":
            target_date.isoformat(),

        "team_a":
            request.team_a,

        "team_b":
            request.team_b,

        "sets_found":
            3,

        "match_ids":
            match_ids,

        "mvp":
            mvp,
    }

# =========================
# FCL SERIES ACTIVATE
# 예약된 경기 실제 시작
# 프리시즌 + 정규리그
# =========================

@app.post(
    "/api/fconline/series/{series_id}/activate"
)
def activate_fcl_series(
    series_id: int,
):

    now = datetime.now(
        ZoneInfo("Asia/Seoul")
    )

    today = now.date()


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.scheduled_date,

                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,

                    s.status,

                    s.team_a_id,
                    s.team_b_id,

                    team_a.fcl_name
                        AS team_a,

                    team_a.current_team_name
                        AS team_a_current_team_name,

                    team_a.current_team_logo_path
                        AS team_a_current_team_logo_path,

                    team_b.fcl_name
                        AS team_b,

                    team_b.current_team_name
                        AS team_b_current_team_name,

                    team_b.current_team_logo_path
                        AS team_b_current_team_logo_path

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            # =========================
            # SERIES 종류 확인
            # =========================

            if (
                series["series_type"]
                not in (
                    "프리시즌",
                    "정규리그",
                    "플레이오프",
                )
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "시작할 수 없는 "
                        "SERIES 종류입니다."
                    ),
                )


            # =========================
            # 예약 상태 확인
            # =========================

            if (
                series["status"]
                != "scheduled"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "예약된 경기만 "
                        "시작할 수 있습니다."
                    ),
                )


            # =========================
            # 경기 당일 확인
            # =========================

            if (
                series["scheduled_date"]
                != today
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "경기는 예정된 "
                        "경기 당일에만 "
                        "시작할 수 있습니다."
                    ),
                )


            # =========================
            # 동일 대진 ACTIVE 확인
            # =========================

            cursor.execute(
                """
                SELECT id
                FROM series

                WHERE
                    status = 'active'

                    AND id <> %s

                    AND (
                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )

                        OR

                        (
                            team_a_id = %s
                            AND
                            team_b_id = %s
                        )
                    )

                LIMIT 1
                """,
                (
                    series_id,

                    series["team_a_id"],
                    series["team_b_id"],

                    series["team_b_id"],
                    series["team_a_id"],
                ),
            )


            active_series = (
                cursor.fetchone()
            )


            if active_series:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "두 참가자 사이에 "
                        "이미 진행 중인 "
                        "SERIES가 있습니다."
                    ),
                )


            # =========================
            # SERIES 시작
            #
            # 경기 시작 순간의 현재 팀을
            # 역사 보존용 Snapshot으로 고정
            # =========================

            cursor.execute(
                """
                UPDATE series

                SET
                    status = 'active',
                    started_at = %s,
                    completed_at = NULL,
                    finished_at = NULL,
                    stats_sync_status =
                        'pending',

                    team_a_snapshot_name = %s,
                    team_a_snapshot_logo_path = %s,

                    team_b_snapshot_name = %s,
                    team_b_snapshot_logo_path = %s

                WHERE id = %s
                """,
                (
                    now,

                    series[
                        "team_a_current_team_name"
                    ],

                    series[
                        "team_a_current_team_logo_path"
                    ],

                    series[
                        "team_b_current_team_name"
                    ],

                    series[
                        "team_b_current_team_logo_path"
                    ],

                    series_id,
                ),
            )


        connection.commit()


    return {
        "series_id":
            series_id,

        "series_type":
            series["series_type"],

        "playoff_stage":
            series["playoff_stage"],

        "best_of":
            series["best_of"],

        "wins_required":
            series["wins_required"],

        "team_a":
            series["team_a"],

        "team_b":
            series["team_b"],

        "scheduled_date":
            series[
                "scheduled_date"
            ].isoformat(),

        "started_at":
            now.isoformat(),

        "status":
            "active",
    }


# =========================
# FCL SERIES CANCEL
# 친선전 취소
# =========================

@app.post(
    "/api/fconline/series/{series_id}/cancel"
)
def cancel_fcl_series(
    series_id: int,
):

    cancelled_at = datetime.now(
        ZoneInfo("Asia/Seoul")
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 확인
            #
            # 취소와 완료 처리가 동시에
            # 발생하지 않도록 행 잠금
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    series_type,
                    status,
                    started_at

                FROM series

                WHERE id = %s

                FOR UPDATE
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            # =========================
            # 프리시즌만 취소 가능
            # =========================

            if (
                series["series_type"]
                != "프리시즌"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "친선전 SERIES만 "
                        "취소할 수 있습니다."
                    ),
                )


            # =========================
            # 예약 / 진행 중만 취소 가능
            # =========================

            if (
                series["status"]
                not in (
                    "scheduled",
                    "active",
                )
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "예약 또는 진행 중인 "
                        "친선전만 취소할 수 있습니다."
                    ),
                )


            previous_status = series["status"]


            # =========================
            # 세트 삭제
            #
            # nexon_match_id UNIQUE 연결도
            # 같이 제거됨
            # =========================

            cursor.execute(
                """
                DELETE FROM series_sets
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            deleted_set_count = cursor.rowcount


            # =========================
            # MVP 삭제
            # =========================

            cursor.execute(
                """
                DELETE FROM series_mvp
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            # =========================
            # 선수 기록 삭제
            # =========================

            cursor.execute(
                """
                DELETE FROM series_player_stats
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            # =========================
            # SERIES 취소
            #
            # started_at은 유지
            # → 시작 후 취소 여부 확인 가능
            # =========================

            cursor.execute(
                """
                UPDATE series

                SET
                    status = 'cancelled',

                    cancelled_at = %s,

                    completed_at = NULL,
                    finished_at = NULL,

                    stats_sync_status = 'pending',

                    team_a_snapshot_name = NULL,
                    team_a_snapshot_logo_path = NULL,

                    team_b_snapshot_name = NULL,
                    team_b_snapshot_logo_path = NULL

                WHERE id = %s

                RETURNING
                    id,
                    status,
                    started_at,
                    cancelled_at
                """,
                (
                    cancelled_at,
                    series_id,
                ),
            )


            cancelled_series = cursor.fetchone()


        connection.commit()


    return {
        "series_id":
            cancelled_series["id"],

        "previous_status":
            previous_status,

        "status":
            cancelled_series["status"],

        "started_at":
            (
                cancelled_series[
                    "started_at"
                ].isoformat()
                if cancelled_series[
                    "started_at"
                ]
                else None
            ),

        "cancelled_at":
            cancelled_series[
                "cancelled_at"
            ].isoformat(),

        "deleted_set_count":
            deleted_set_count,

        "message":
            "친선전이 취소되었습니다.",
    }

# =========================
# FCL SERIES MANUAL COMPLETE
# 친선전 수동 결과 입력
# =========================

@app.post(
    "/api/fconline/series/{series_id}/manual-complete"
)
def manual_complete_fcl_series(
    series_id: int,
    request: ManualSeriesCompleteRequest,
):

    finished_at = datetime.now(
        ZoneInfo("Asia/Seoul")
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.status,

                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,

                    team_a.fcl_name
                        AS team_a,

                    team_a.current_team_name
                        AS team_a_current_team_name,

                    team_a.current_team_logo_path
                        AS team_a_current_team_logo_path,

                    team_b.fcl_name
                        AS team_b,

                    team_b.current_team_name
                        AS team_b_current_team_name,

                    team_b.current_team_logo_path
                        AS team_b_current_team_logo_path

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s

                FOR UPDATE OF s
                """,
                (
                    series_id,
                ),
            )


            series = cursor.fetchone()


            if not series:
                raise HTTPException(
                    status_code=404,
                    detail="SERIES를 찾을 수 없습니다.",
                )


            # =========================
            # 지원 SERIES
            # =========================

            if series["series_type"] not in (
                "프리시즌",
                "정규리그",
                "플레이오프",
            ):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "수동 결과를 입력할 수 없는 "
                        "SERIES입니다."
                    ),
                )


            # =========================
            # 진행 중 경기만 가능
            # =========================

            if series["status"] != "active":
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "진행 중인 SERIES만 "
                        "수동 완료할 수 있습니다."
                    ),
                )


            # =========================
            # 완료 Snapshot 검증
            # =========================

            if (
                not series[
                    "team_a_current_team_name"
                ]
                or
                not series[
                    "team_a_current_team_logo_path"
                ]
                or
                not series[
                    "team_b_current_team_name"
                ]
                or
                not series[
                    "team_b_current_team_logo_path"
                ]
            ):
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "현재 팀 또는 로고 정보가 "
                        "등록되지 않았습니다."
                    ),
                )


            # =========================
            # 입력된 모든 세트
            # =========================

            requested_sets = [
                (
                    1,
                    request.set1_team_a,
                    request.set1_team_b,
                    request.set1_winner_side,
                ),
                (
                    2,
                    request.set2_team_a,
                    request.set2_team_b,
                    request.set2_winner_side,
                ),
                (
                    3,
                    request.set3_team_a,
                    request.set3_team_b,
                    request.set3_winner_side,
                ),
                (
                    4,
                    request.set4_team_a,
                    request.set4_team_b,
                    request.set4_winner_side,
                ),
                (
                    5,
                    request.set5_team_a,
                    request.set5_team_b,
                    request.set5_winner_side,
                ),
                (
                    6,
                    request.set6_team_a,
                    request.set6_team_b,
                    request.set6_winner_side,
                ),
                (
                    7,
                    request.set7_team_a,
                    request.set7_team_b,
                    request.set7_winner_side,
                ),
            ]


            is_playoff = (
                series["series_type"]
                == "플레이오프"
            )


            if is_playoff:

                if (
                    series["best_of"] is None
                    or
                    series["wins_required"] is None
                ):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "플레이오프 진행 정보가 "
                            "올바르지 않습니다."
                        ),
                    )


                max_sets = int(
                    series["best_of"]
                )

                wins_required = int(
                    series["wins_required"]
                )

            else:

                max_sets = 3
                wins_required = None


            # =========================
            # 허용 세트 초과 방어
            # =========================

            for (
                set_number,
                team_a_score,
                team_b_score,
                explicit_winner_side,
            ) in requested_sets[max_sets:]:

                if (
                    team_a_score is not None
                    or team_b_score is not None
                    or explicit_winner_side is not None
                ):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{max_sets}세트를 초과하여 "
                            "결과를 입력할 수 없습니다."
                        ),
                    )


            # =========================
            # 실제 저장할 세트 검증
            # =========================

            manual_sets = []

            gap_found = False

            team_a_wins = 0
            team_b_wins = 0

            series_winner_side = None
            winning_set_number = None


            for (
                set_number,
                team_a_score,
                team_b_score,
                explicit_winner_side,
            ) in requested_sets[:max_sets]:

                # 둘 다 비어 있으면
                # 여기서 실제 경기 종료
                if (
                    team_a_score is None
                    and team_b_score is None
                ):

                    if explicit_winner_side is not None:
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 점수 없이 "
                                "승자만 지정할 수 없습니다."
                            ),
                        )

                    gap_found = True
                    continue


                # 한쪽만 입력됨
                if (
                    team_a_score is None
                    or team_b_score is None
                ):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{set_number}세트의 "
                            "양쪽 점수를 모두 입력해주세요."
                        ),
                    )


                # 중간 빈 세트 이후 입력
                if gap_found:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "중간 세트를 비워둔 채 "
                            "다음 세트를 입력할 수 없습니다."
                        ),
                    )


                if (
                    team_a_score < 0
                    or team_b_score < 0
                ):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "점수는 0 이상의 "
                            "정수여야 합니다."
                        ),
                    )


                # =========================
                # 세트 승자
                # =========================

                if team_a_score > team_b_score:

                    winner_side = "team_a"

                    if (
                        explicit_winner_side
                        is not None
                        and
                        explicit_winner_side
                        != winner_side
                    ):
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 승자와 "
                                "입력 점수가 일치하지 않습니다."
                            ),
                        )


                elif team_b_score > team_a_score:

                    winner_side = "team_b"

                    if (
                        explicit_winner_side
                        is not None
                        and
                        explicit_winner_side
                        != winner_side
                    ):
                        raise HTTPException(
                            status_code=400,
                            detail=(
                                f"{set_number}세트 승자와 "
                                "입력 점수가 일치하지 않습니다."
                            ),
                        )


                else:

                    # 일반 경기에서는 무승부 허용
                    if not is_playoff:
                        winner_side = "draw"

                    # 플레이오프는
                    # 승부차기 등 실제 승자 필요
                    else:

                        if explicit_winner_side not in (
                            "team_a",
                            "team_b",
                        ):
                            raise HTTPException(
                                status_code=400,
                                detail=(
                                    f"{set_number}세트가 동점입니다. "
                                    "플레이오프에서는 실제 승자를 "
                                    "지정해야 합니다."
                                ),
                            )

                        winner_side = (
                            explicit_winner_side
                        )


                # =========================
                # 선승 이후 추가 세트 방어
                # =========================

                if (
                    is_playoff
                    and
                    series_winner_side is not None
                ):
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "선승 도달 이후의 "
                            "추가 세트가 입력되었습니다."
                        ),
                    )


                manual_sets.append(
                    (
                        set_number,
                        team_a_score,
                        team_b_score,
                        winner_side,
                    )
                )


                # =========================
                # 플레이오프 승수 계산
                # =========================

                if is_playoff:

                    if winner_side == "team_a":
                        team_a_wins += 1

                    elif winner_side == "team_b":
                        team_b_wins += 1


                    if (
                        team_a_wins
                        >= wins_required
                    ):
                        series_winner_side = (
                            "team_a"
                        )

                        winning_set_number = (
                            set_number
                        )


                    elif (
                        team_b_wins
                        >= wins_required
                    ):
                        series_winner_side = (
                            "team_b"
                        )

                        winning_set_number = (
                            set_number
                        )


            # =========================
            # 일반 SERIES는 정확히 3세트
            # =========================

            if not is_playoff:

                if len(manual_sets) != 3:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            "프리시즌과 정규리그는 "
                            "3세트를 모두 입력해야 합니다."
                        ),
                    )


            # =========================
            # 플레이오프는 선승 필수
            # =========================

            else:

                if series_winner_side is None:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"{wins_required}승에 도달한 "
                            "참가자가 없습니다."
                        ),
                    )


            # =========================
            # 검증 완료 후 기존 기록 제거
            # =========================

            cursor.execute(
                """
                DELETE FROM series_sets
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            cursor.execute(
                """
                DELETE FROM series_mvp
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            cursor.execute(
                """
                DELETE FROM series_player_stats
                WHERE series_id = %s
                """,
                (
                    series_id,
                ),
            )


            # =========================
            # 세트 저장
            # =========================

            for (
                set_number,
                team_a_score,
                team_b_score,
                winner_side,
            ) in manual_sets:

                cursor.execute(
                    """
                    INSERT INTO series_sets (
                        series_id,
                        set_number,
                        nexon_match_id,
                        played_at,
                        team_a_score,
                        team_b_score,
                        score_source,
                        winner_side
                    )

                    VALUES (
                        %s,
                        %s,
                        NULL,
                        %s,
                        %s,
                        %s,
                        'manual',
                        %s
                    )
                    """,
                    (
                        series_id,
                        set_number,
                        finished_at,
                        team_a_score,
                        team_b_score,
                        winner_side,
                    ),
                )


            # =========================
            # SERIES 완료 + Snapshot
            # =========================

            cursor.execute(
                """
                UPDATE series

                SET
                    status = 'completed',
                    completed_at = %s,
                    finished_at = %s,
                    stats_sync_status = 'pending',

                    team_a_snapshot_name =
                        COALESCE(
                            team_a_snapshot_name,
                            %s
                        ),

                    team_a_snapshot_logo_path =
                        COALESCE(
                            team_a_snapshot_logo_path,
                            %s
                        ),

                    team_b_snapshot_name =
                        COALESCE(
                            team_b_snapshot_name,
                            %s
                        ),

                    team_b_snapshot_logo_path =
                        COALESCE(
                            team_b_snapshot_logo_path,
                            %s
                        )

                WHERE id = %s
                """,
                (
                    finished_at,
                    finished_at,

                    series[
                        "team_a_current_team_name"
                    ],
                    series[
                        "team_a_current_team_logo_path"
                    ],

                    series[
                        "team_b_current_team_name"
                    ],
                    series[
                        "team_b_current_team_logo_path"
                    ],

                    series_id,
                ),
            )


        connection.commit()

    # =========================
    # 정규리그 종료 후
    # 준플레이오프 자동 생성
    # =========================

    if (
        series["series_type"]
        == "정규리그"
    ):
        create_initial_playoff_if_ready()

    if (
        series["series_type"]
        == "플레이오프"
    ):
        create_next_playoff_if_ready(
            series_id
        )


    team_a_total_score = sum(
        manual_set[1]
        for manual_set in manual_sets
    )

    team_b_total_score = sum(
        manual_set[2]
        for manual_set in manual_sets
    )


    return {
        "series_id":
            series_id,

        "status":
            "completed",

        "stats_sync_status":
            "pending",

        "series_type":
            series["series_type"],

        "playoff_stage":
            series["playoff_stage"],

        "best_of":
            series["best_of"],

        "wins_required":
            series["wins_required"],

        "team_a":
            series["team_a"],

        "team_b":
            series["team_b"],

        "team_a_score":
            team_a_total_score,

        "team_b_score":
            team_b_total_score,

        "team_a_wins":
            (
                team_a_wins
                if is_playoff
                else None
            ),

        "team_b_wins":
            (
                team_b_wins
                if is_playoff
                else None
            ),

        "winner_side":
            series_winner_side,

        "winning_set":
            winning_set_number,

        "finished_at":
            finished_at.isoformat(),

        "sets": [
            {
                "set":
                    set_number,

                "team_a_score":
                    team_a_score,

                "team_b_score":
                    team_b_score,

                "winner_side":
                    winner_side,
            }

            for (
                set_number,
                team_a_score,
                team_b_score,
                winner_side,
            ) in manual_sets
        ],
    }

# =========================
# FCL SERIES STATUS
# =========================

# =========================
# FCL SERIES STATUS
# DB 조회 전용
# NEXON API 호출 없음
# =========================

@app.get(
    "/api/fconline/series/{series_id}/status"
)
def get_fcl_series_status(
    series_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.match_type,
                    s.scheduled_date,

                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,

                    s.started_at,
                    s.completed_at,
                    s.finished_at,
                    s.stats_sync_status,
                    s.status,

                    team_a.fcl_name
                        AS team_a_name,

                    team_a.fc_nickname
                        AS nickname_a,

                    team_b.fcl_name
                        AS team_b_name,

                    team_b.fc_nickname
                        AS nickname_b

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s
                """,
                (
                    series_id,
                ),
            )

            series = cursor.fetchone()

            if not series:

                raise HTTPException(
                    status_code=404,
                    detail="SERIES를 찾을 수 없습니다.",
                )

            # =========================
            # 저장된 SET
            # =========================

            cursor.execute(
                """
                SELECT
                    set_number,
                    nexon_match_id,
                    played_at,
                    team_a_score,
                    team_b_score,
                    score_source
                    winner_side

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )

            saved_sets = cursor.fetchall()

            # =========================
            # 저장된 MVP
            # =========================

            cursor.execute(
                """
                SELECT
                    p.fcl_name,
                    p.fc_nickname
                        AS nickname,

                    sm.sp_id,
                    sm.player_name,
                    sm.sets_played,
                    sm.rating_total,
                    sm.average_rating,
                    sm.goals,
                    sm.assists,
                    sm.image_url

                FROM series_mvp AS sm

                JOIN participants AS p
                    ON p.id =
                        sm.participant_id

                WHERE sm.series_id = %s
                """,
                (
                    series_id,
                ),
            )

            mvp = cursor.fetchone()

    sets = []

    for saved_set in saved_sets:

        sets.append(
            {
                "set":
                    saved_set[
                        "set_number"
                    ],

                "match_id":
                    saved_set[
                        "nexon_match_id"
                    ],

                "match_date":
                    (
                        saved_set[
                            "played_at"
                        ].isoformat()

                        if saved_set[
                            "played_at"
                        ]

                        else None
                    ),

                "team_a_score":
                    saved_set[
                        "team_a_score"
                    ],

                "team_b_score":
                    saved_set[
                        "team_b_score"
                    ],

                "score_source":
                    saved_set[
                        "score_source"
                    ],

                "winner_side":
                    saved_set[
                        "winner_side"
                    ],
            }
        )

    if mvp:

        mvp["rating_total"] = float(
            mvp["rating_total"]
        )

        mvp["average_rating"] = float(
            mvp["average_rating"]
        )

    return {
        "series": {
            "series_id":
                series_id,

            "series_type":
                series["series_type"],

            "playoff_stage":
                series[
                    "playoff_stage"
                ],

            "best_of":
                series[
                    "best_of"
                ],

            "wins_required":
                series[
                    "wins_required"
                ],

            "team_a":
                series["team_a_name"],

            "team_b":
                series["team_b_name"],

            "nickname_a":
                series["nickname_a"],

            "nickname_b":
                series["nickname_b"],

            "match_type":
                series["match_type"],

            "scheduled_date":
                (
                    series[
                        "scheduled_date"
                    ].isoformat()

                    if series[
                        "scheduled_date"
                    ]

                    else None
                ),

                "started_at":
                    (
                        series[
                            "started_at"
                        ].isoformat()

                        if series[
                            "started_at"
                        ]

                        else None
                    ),

                "finished_at":
                    (
                        series[
                            "finished_at"
                        ].isoformat()

                        if series[
                            "finished_at"
                        ]

                        else None
                    ),

                "stats_sync_status":
                    series[
                        "stats_sync_status"
                    ],

                "status":
                    series[
                        "status"
                    ],

                "set_count":
                    len(sets),
        },

        "sets": sets,

        "mvp": mvp,
    }


# =========================
# FCL SERIES STATUS
# NEXON 탐색 + DB 저장
# =========================
@app.post(
    "/api/fconline/series/{series_id}/sync"
)
def sync_fcl_series_status(
    series_id: int,
):
    # =========================
    # SERIES + 참가자 조회
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.match_type,
                    s.scheduled_date,

                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,

                    s.started_at,
                    s.completed_at,
                    s.finished_at,
                    s.stats_sync_status,
                    s.status,

                    team_a.id
                        AS team_a_id,

                    team_a.fcl_name
                        AS team_a_name,

                    team_a.current_team_name
                        AS team_a_current_team_name,

                    team_a.current_team_logo_path
                        AS team_a_current_team_logo_path,

                    team_a.fc_nickname
                        AS nickname_a,

                    team_a.ouid
                        AS ouid_a,

                    team_b.id
                        AS team_b_id,

                    team_b.fcl_name
                        AS team_b_name,

                    team_b.current_team_name
                        AS team_b_current_team_name,

                    team_b.current_team_logo_path
                        AS team_b_current_team_logo_path,

                    team_b.fc_nickname
                        AS nickname_b,

                    team_b.ouid
                        AS ouid_b

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s
                """,
                (
                    series_id,
                ),
            )

            series = cursor.fetchone()


    if not series:

        raise HTTPException(
            status_code=404,
            detail="SERIES를 찾을 수 없습니다.",
        )


    # =========================
    # 현재 SERIES가
    # 사후 동기화 대상인지 확인
    # =========================

    is_pending_result_sync = (
        series["status"] == "completed"
        and
        series["stats_sync_status"]
        in (
            "pending",
            "conflict",
        )
    )


    # active 또는
    # completed + pending/conflict만
    # NEXON 동기화 가능
    if (
        series["status"] != "active"
        and
        not is_pending_result_sync
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "NEXON 기록을 동기화할 수 없는 "
                "SERIES입니다."
            ),
        )


    nickname_a = series[
        "nickname_a"
    ]

    nickname_b = series[
        "nickname_b"
    ]


    if (
        not nickname_a
        or
        not nickname_b
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "FC Online 닉네임 정보가 "
                "없습니다."
            ),
        )


    # =========================
    # OUID 준비
    # =========================

    ouid_a = get_participant_ouid(
        series["team_a_id"],
        nickname_a,
        series["ouid_a"],
    )


    ouid_b = get_participant_ouid(
        series["team_b_id"],
        nickname_b,
        series["ouid_b"],
    )


    # =========================
    # NEXON 최근 경기 조회
    # =========================

    matches_a = get_user_match_ids(
        ouid_a,
        series["match_type"],
        limit=10,
    )


    # 개발/서비스 키 모두
    # 순간 호출 몰림 방지
    time.sleep(0.3)


    matches_b = get_user_match_ids(
        ouid_b,
        series["match_type"],
        limit=10,
    )


    matches_b_set = set(
        matches_b
    )


    common_match_ids = [
        match_id

        for match_id
        in matches_a

        if match_id
        in matches_b_set
    ]


    # =========================
    # SERIES 시간 범위
    # =========================

    started_at = parse_kst_datetime(
        series["started_at"]
    )


    finished_at = None


    if series["finished_at"]:

        finished_at = parse_kst_datetime(
            series["finished_at"]
        )


    detected_matches = []
    debug_matches = []


    # =========================
    # 실제 맞대결 탐색
    # =========================

    for match_id in common_match_ids:

        time.sleep(0.3)

        match_data = get_match_detail(
            match_id
        )


        match_date = parse_nexon_datetime(
            match_data["matchDate"]
        )


        match_nicknames = {
            match_info["nickname"]
            for match_info
            in match_data["matchInfo"]
        }


        reject_reason = None


        if match_date < started_at:

            reject_reason = (
                "before_started_at"
            )


        elif (
            finished_at is not None
            and
            match_date > finished_at
        ):

            reject_reason = (
                "after_finished_at"
            )


        elif (
            match_data["matchType"]
            != series["match_type"]
        ):

            reject_reason = (
                "match_type_mismatch"
            )


        elif match_nicknames != {
            nickname_a,
            nickname_b,
        }:

            reject_reason = (
                "nickname_mismatch"
            )


        debug_matches.append(
            {
                "match_id":
                    match_id,

                "match_date":
                    match_date.isoformat(),

                "match_type":
                    match_data[
                        "matchType"
                    ],

                "nicknames":
                    sorted(
                        match_nicknames
                    ),

                "reject_reason":
                    reject_reason,
            }
        )


        if reject_reason is not None:
            continue


        detected_matches.append(
            {
                "data":
                    match_data,

                "played_at":
                    match_date,
            }
        )

    # =========================
    # 시간순 정렬
    # =========================

    detected_matches.sort(
        key=lambda match:
            match["played_at"]
    )


    # =========================
    # SERIES 종류별 경기 제한
    # =========================

    if (
        series["series_type"]
        == "플레이오프"
    ):

        best_of = int(
            series["best_of"]
        )

        wins_required = int(
            series["wins_required"]
        )


        playoff_matches = []

        detected_team_a_wins = 0
        detected_team_b_wins = 0


        for detected_match in (
            detected_matches[:best_of]
        ):

            winner_side = (
                get_match_winner_side(
                    detected_match["data"],
                    nickname_a,
                    nickname_b,
                    series["series_type"],
                )
            )


            if winner_side is None:

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "플레이오프 경기의 "
                        "승패 결과를 확인할 수 없습니다."
                    ),
                )


            if winner_side == "draw":

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "플레이오프 경기에서 "
                        "무승부 결과가 감지되었습니다."
                    ),
                )


            detected_match[
                "winner_side"
            ] = winner_side


            playoff_matches.append(
                detected_match
            )


            if winner_side == "team_a":

                detected_team_a_wins += 1

            elif winner_side == "team_b":

                detected_team_b_wins += 1


            # 선승 도달 즉시 SERIES 종료 지점
            if (
                detected_team_a_wins
                >= wins_required

                or

                detected_team_b_wins
                >= wins_required
            ):
                break


        detected_matches = playoff_matches


    else:

        # 프리시즌 / 정규리그
        # 항상 3세트
        detected_matches = (
            detected_matches[:3]
        )

        # =========================
        # NEXON 경기 데이터 정합성 검증
        # =========================

        integrity_conflict = None

        for (
            set_index,
            detected_match,
        ) in enumerate(
            detected_matches,
            start=1,
        ):
            conflict_reason = (
                get_match_integrity_conflict(
                    detected_match["data"],
                    nickname_a,
                    nickname_b,
                    series["series_type"],
                )
            )

            if conflict_reason:
                integrity_conflict = (
                    f"{set_index}세트: "
                    f"{conflict_reason}"
                )
                break

        if integrity_conflict:
            if is_pending_result_sync:
                with get_db_connection() as connection:
                    with connection.cursor() as cursor:
                        cursor.execute(
                            """
                            UPDATE series
                            SET
                                stats_sync_status =
                                    'conflict'
                            WHERE id = %s
                            """,
                            (
                                series_id,
                            ),
                        )

                    connection.commit()

            return {
                "series": {
                    "series_id":
                        series_id,

                    "series_type":
                        series["series_type"],

                    "team_a":
                        series["team_a_name"],

                    "team_b":
                        series["team_b_name"],

                    "status":
                        series["status"],

                    "set_count":
                        0,

                    "stats_sync_status":
                        (
                            "conflict"
                            if is_pending_result_sync
                            else
                            series[
                                "stats_sync_status"
                            ]
                        ),
                },

                "sets": [],

                "mvp": None,

                "sync_message": (
                    "NEXON 경기 데이터 정합성 "
                    "검증에 실패했습니다. "
                    f"{integrity_conflict}"
                ),
            }


        for detected_match in detected_matches:

            detected_match[
                "winner_side"
            ] = (
                get_match_winner_side(
                    detected_match["data"],
                    nickname_a,
                    nickname_b,
                    series["series_type"],
                )
            )

    # =========================
    # NEXON 경기 데이터 정합성 검증
    # =========================

    integrity_conflict = None

    for (
        set_index,
        detected_match,
    ) in enumerate(
        detected_matches,
        start=1,
    ):
        conflict_reason = (
            get_match_integrity_conflict(
                detected_match["data"],
                nickname_a,
                nickname_b,
                series["series_type"],
            )
        )

        if conflict_reason:
            integrity_conflict = (
                f"{set_index}세트: "
                f"{conflict_reason}"
            )
            break

    if integrity_conflict:
        if is_pending_result_sync:
            with get_db_connection() as connection:
                with connection.cursor() as cursor:
                    cursor.execute(
                        """
                        UPDATE series
                        SET
                            stats_sync_status =
                                'conflict'
                        WHERE id = %s
                        """,
                        (
                            series_id,
                        ),
                    )

                connection.commit()

        return {
            "series": {
                "series_id":
                    series_id,

                "series_type":
                    series["series_type"],

                "team_a":
                    series["team_a_name"],

                "team_b":
                    series["team_b_name"],

                "status":
                    series["status"],

                "set_count":
                    0,

                "stats_sync_status":
                    (
                        "conflict"
                        if is_pending_result_sync
                        else
                        series[
                            "stats_sync_status"
                        ]
                    ),
            },

            "sets": [],

            "mvp": None,

            "sync_message": (
                "NEXON 경기 데이터 정합성 "
                "검증에 실패했습니다. "
                f"{integrity_conflict}"
            ),
        }

    # ==================================================
    # A.
    # 수동 결과 입력 완료 후
    # NEXON 기록 사후 연결
    # ==================================================

    if is_pending_result_sync:

        # =========================
        # 기존 수동 결과 조회
        # =========================

        with get_db_connection() as connection:

            with connection.cursor() as cursor:

                cursor.execute(
                    """
                    SELECT
                        set_number,
                        nexon_match_id,
                        played_at,
                        team_a_score,
                        team_b_score,
                        score_source

                    FROM series_sets

                    WHERE series_id = %s

                    ORDER BY set_number
                    """,
                    (
                        series_id,
                    ),
                )

                manual_sets = (
                    cursor.fetchall()
                )


        if len(manual_sets) != 3:

            raise HTTPException(
                status_code=400,
                detail=(
                    "수동 입력된 3세트 결과가 "
                    "완전하지 않습니다."
                ),
            )


        # =========================
        # Nexon 데이터가 아직
        # 3경기 전부 올라오지 않음
        # =========================

        if len(detected_matches) < 3:

            return {
                "series": {
                    "series_id":
                        series_id,

                    "series_type":
                        series[
                            "series_type"
                        ],

                    "team_a":
                        series[
                            "team_a_name"
                        ],

                    "team_b":
                        series[
                            "team_b_name"
                        ],

                    "nickname_a":
                        nickname_a,

                    "nickname_b":
                        nickname_b,

                    "match_type":
                        series[
                            "match_type"
                        ],

                    "started_at":
                        started_at.isoformat(),

                    "status":
                        "completed",

                    "set_count":
                        3,

                    "stats_sync_status":
                        "pending",
                },

                "sets": [
                    {
                        "set":
                            saved_set[
                                "set_number"
                            ],

                        "match_id":
                            saved_set[
                                "nexon_match_id"
                            ],

                        "match_date":
                            (
                                saved_set[
                                    "played_at"
                                ].isoformat()

                                if saved_set[
                                    "played_at"
                                ]

                                else None
                            ),

                        "team_a_score":
                            saved_set[
                                "team_a_score"
                            ],

                        "team_b_score":
                            saved_set[
                                "team_b_score"
                            ],
                    }

                    for saved_set
                    in manual_sets
                ],

                "detected_set_count":
                    len(detected_matches),

                "common_match_count":
                    len(common_match_ids),

                "debug": {
                    "started_at":
                        started_at.isoformat(),

                    "finished_at":
                        (
                            finished_at.isoformat()
                            if finished_at
                            else None
                        ),

                    "matches":
                        debug_matches,
                },

                "mvp":
                    None,

                "sync_message":
                    (
                        "NEXON 경기 기록 "
                        f"{len(detected_matches)}/3경기 감지. "
                        "아직 3경기 기록이 모두 "
                        "반영되지 않았습니다."
                    ),
            }


        # =========================
        # Nexon 점수 추출
        # =========================

        nexon_scores = []


        for detected_match in (
            detected_matches
        ):

            match_data = (
                detected_match[
                    "data"
                ]
            )


            participant_map = {
                match_info["nickname"]:
                    match_info

                for match_info
                in match_data[
                    "matchInfo"
                ]
            }


            team_a_info = (
                participant_map[
                    nickname_a
                ]
            )


            team_b_info = (
                participant_map[
                    nickname_b
                ]
            )


            team_a_score = (
                team_a_info[
                    "shoot"
                ][
                    "goalTotal"
                ]
            )


            team_b_score = (
                team_b_info[
                    "shoot"
                ][
                    "goalTotal"
                ]
            )


            nexon_scores.append(
                (
                    team_a_score,
                    team_b_score,
                )
            )


        # =========================
        # 수동 점수와
        # Nexon 점수 비교
        # =========================

        has_score_conflict = any(
            (
                manual_set[
                    "team_a_score"
                ]
                !=
                nexon_score[0]
            )
            or
            (
                manual_set[
                    "team_b_score"
                ]
                !=
                nexon_score[1]
            )

            for (
                manual_set,
                nexon_score,
            )
            in zip(
                manual_sets,
                nexon_scores,
            )
        )


        # =========================
        # 점수 불일치
        # =========================

        if has_score_conflict:

            with get_db_connection() as connection:

                with connection.cursor() as cursor:

                    cursor.execute(
                        """
                        UPDATE series

                        SET
                            stats_sync_status =
                                'conflict'

                        WHERE id = %s
                        """,
                        (
                            series_id,
                        ),
                    )

                connection.commit()


            return {
                "series": {
                    "series_id":
                        series_id,

                    "series_type":
                        series[
                            "series_type"
                        ],

                    "team_a":
                        series[
                            "team_a_name"
                        ],

                    "team_b":
                        series[
                            "team_b_name"
                        ],

                    "status":
                        "completed",

                    "set_count":
                        3,

                    "stats_sync_status":
                        "conflict",
                },

                "sets": [
                    {
                        "set":
                            saved_set[
                                "set_number"
                            ],

                        "team_a_score":
                            saved_set[
                                "team_a_score"
                            ],

                        "team_b_score":
                            saved_set[
                                "team_b_score"
                            ],
                    }

                    for saved_set
                    in manual_sets
                ],

                "detected_sets": [
                    {
                        "set":
                            index,

                        "team_a_score":
                            score[0],

                        "team_b_score":
                            score[1],
                    }

                    for (
                        index,
                        score,
                    )
                    in enumerate(
                        nexon_scores,
                        start=1,
                    )
                ],

                "mvp":
                    None,

                "sync_message":
                    (
                        "수동 입력 점수와 "
                        "NEXON 기록이 "
                        "일치하지 않습니다."
                    ),
            }


        # =========================
        # 점수 일치
        # 실제 matchId 연결
        # =========================

        with get_db_connection() as connection:

            with connection.cursor() as cursor:

                for (
                    index,
                    detected_match,
                ) in enumerate(
                    detected_matches,
                    start=1,
                ):

                    cursor.execute(
                        """
                        UPDATE series_sets

                        SET
                            nexon_match_id = %s,
                            played_at = %s,
                            score_source =
                                'nexon'

                        WHERE
                            series_id = %s

                            AND
                            set_number = %s
                        """,
                        (
                            detected_match[
                                "data"
                            ][
                                "matchId"
                            ],

                            detected_match[
                                "played_at"
                            ],

                            series_id,
                            index,
                        ),
                    )


                # 혹시 이전 통계가 있었다면
                # 깨끗하게 다시 계산
                cursor.execute(
                    """
                    DELETE FROM series_mvp
                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


                cursor.execute(
                    """
                    DELETE FROM
                        series_player_stats

                    WHERE series_id = %s
                    """,
                    (
                        series_id,
                    ),
                )


            connection.commit()

        # =========================
        # 세트별 스쿼드 Snapshot 저장
        # =========================

        save_series_set_squad_players(
            series_id,

            series[
                "team_a_id"
            ],
            nickname_a,

            series[
                "team_b_id"
            ],
            nickname_b,

            detected_matches,
        )


        # =========================
        # 이미 받은 match detail로
        # MVP / 선수 기록 계산
        #
        # 추가 Nexon 호출 없음
        # =========================

        mvp_matches = [
            detected_match[
                "data"
            ]

            for detected_match
            in detected_matches
        ]


        (
            mvp,
            _,
            player_stats,
        ) = calculate_series_mvp_from_matches(
            mvp_matches
        )


        save_series_player_stats(
            series_id,

            series[
                "team_a_id"
            ],
            nickname_a,

            series[
                "team_b_id"
            ],
            nickname_b,

            player_stats,
        )


        if mvp:

            if (
                mvp["nickname"]
                == nickname_a
            ):

                mvp_participant_id = (
                    series[
                        "team_a_id"
                    ]
                )


            elif (
                mvp["nickname"]
                == nickname_b
            ):

                mvp_participant_id = (
                    series[
                        "team_b_id"
                    ]
                )


            else:

                mvp_participant_id = None


            if mvp_participant_id:

                save_series_mvp(
                    series_id,
                    mvp_participant_id,
                    mvp,
                )


        # =========================
        # 사후 동기화 완료
        # =========================

        with get_db_connection() as connection:

            with connection.cursor() as cursor:

                cursor.execute(
                    """
                    UPDATE series

                    SET
                        stats_sync_status =
                            'synced'

                    WHERE id = %s
                    """,
                    (
                        series_id,
                    ),
                )

                settlement_result = (
                    settle_predictions_for_series(
                        cursor,
                        series_id,
                    )
                )

            connection.commit()


        sets = []


        for (
            index,
            detected_match,
        ) in enumerate(
            detected_matches,
            start=1,
        ):

            manual_set = (
                manual_sets[
                    index - 1
                ]
            )


            sets.append(
                {
                    "set":
                        index,

                    "match_id":
                        detected_match[
                            "data"
                        ][
                            "matchId"
                        ],

                    "match_date":
                        detected_match[
                            "played_at"
                        ].isoformat(),

                    "team_a_score":
                        manual_set[
                            "team_a_score"
                        ],

                    "team_b_score":
                        manual_set[
                            "team_b_score"
                        ],
                }
            )

    # =========================
    # 정규리그 종료 후
    # 준플레이오프 자동 생성
    # =========================

        if (
            series["series_type"]
            == "정규리그"
            and
            status == "completed"
        ):
            create_initial_playoff_if_ready()


        if (
            series["series_type"]
            == "플레이오프"
            and
            status == "completed"
        ):
            create_next_playoff_if_ready(
                series_id
            )


        return {
            "series": {
                "series_id":
                    series_id,

                "series_type":
                    series[
                        "series_type"
                    ],

                "team_a":
                    series[
                        "team_a_name"
                    ],

                "team_b":
                    series[
                        "team_b_name"
                    ],

                "nickname_a":
                    nickname_a,

                "nickname_b":
                    nickname_b,

                "match_type":
                    series[
                        "match_type"
                    ],

                "started_at":
                    started_at.isoformat(),

                "status":
                    "completed",

                "set_count":
                    3,

                "stats_sync_status":
                    "synced",
            },

            "sets":
                sets,

            "mvp":
                mvp,

            "sync_message":
                (
                    "NEXON 기록 동기화가 "
                    "완료되었습니다."
                ),
        }

    # =========================
    # 완료 Snapshot 기준 정보 확인
    #
    # 이미 completed 상태에서 하는
    # 사후 sync에는 적용하지 않음
    # =========================

    if (
        not series[
            "team_a_current_team_name"
        ]
        or
        not series[
            "team_a_current_team_logo_path"
        ]
        or
        not series[
            "team_b_current_team_name"
        ]
        or
        not series[
            "team_b_current_team_logo_path"
        ]
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "현재 팀 또는 로고 정보가 "
                "등록되지 않았습니다."
            ),
        )




    # ==================================================
    # B.
    # 아직 active 상태에서
    # 직접 NEXON 기록 확인
    # ==================================================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            for (
                index,
                detected_match,
            ) in enumerate(
                detected_matches,
                start=1,
            ):

                match_data = (
                    detected_match[
                        "data"
                    ]
                )


                played_at = (
                    detected_match[
                        "played_at"
                    ]
                )


                participant_map = {
                    match_info["nickname"]:
                        match_info

                    for match_info
                    in match_data[
                        "matchInfo"
                    ]
                }


                team_a_score = (
                    participant_map[
                        nickname_a
                    ][
                        "shoot"
                    ][
                        "goalTotal"
                    ]
                )


                team_b_score = (
                    participant_map[
                        nickname_b
                    ][
                        "shoot"
                    ][
                        "goalTotal"
                    ]
                )


                cursor.execute(
                    """
                    INSERT INTO series_sets (
                        series_id,
                        set_number,
                        nexon_match_id,
                        played_at,
                        team_a_score,
                        team_b_score,
                        score_source,
                        winner_side
                    )

                    VALUES (
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        'nexon',
                        %s
                    )

                    ON CONFLICT (
                        series_id,
                        set_number
                    )

                    DO UPDATE SET
                        nexon_match_id =
                            EXCLUDED.nexon_match_id,

                        played_at =
                            EXCLUDED.played_at,

                        team_a_score =
                            EXCLUDED.team_a_score,

                        team_b_score =
                            EXCLUDED.team_b_score,

                        score_source =
                            'nexon',

                        winner_side =
                            EXCLUDED.winner_side
                    """,
                    (
                        series_id,
                        index,

                        match_data[
                            "matchId"
                        ],

                        played_at,

                        team_a_score,
                        team_b_score,

                        detected_match[
                            "winner_side"
                        ],
                    ),
                )


            # =========================
            # 저장된 세트 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    set_number,
                    nexon_match_id,
                    played_at,
                    team_a_score,
                    team_b_score,
                    score_source,
                    winner_side

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    series_id,
                ),
            )


            saved_sets = (
                cursor.fetchall()
            )


            set_count = len(
                saved_sets
            )


            # =========================
            # SERIES 완료 여부
            # =========================

            team_a_wins = 0
            team_b_wins = 0


            if (
                series["series_type"]
                == "플레이오프"
            ):

                for saved_set in saved_sets:

                    if (
                        saved_set[
                            "winner_side"
                        ]
                        == "team_a"
                    ):

                        team_a_wins += 1


                    elif (
                        saved_set[
                            "winner_side"
                        ]
                        == "team_b"
                    ):

                        team_b_wins += 1


                series_completed = (
                    team_a_wins
                    >= int(
                        series[
                            "wins_required"
                        ]
                    )

                    or

                    team_b_wins
                    >= int(
                        series[
                            "wins_required"
                        ]
                    )
                )


            else:

                # 프리시즌 / 정규리그
                # 항상 3세트
                series_completed = (
                    set_count >= 3
                )


            # =========================
            # SERIES 종료
            # =========================

            if series_completed:

                completed_at = (
                    saved_sets[-1][
                        "played_at"
                    ]
                )


                cursor.execute(
            """
            UPDATE series

            SET
                status =
                    'completed',

                completed_at = %s,

                finished_at =
                    COALESCE(
                        finished_at,
                        %s
                    ),

                stats_sync_status =
                    'pending',

                team_a_snapshot_name =
                    COALESCE(
                        team_a_snapshot_name,
                        %s
                    ),

                team_a_snapshot_logo_path =
                    COALESCE(
                        team_a_snapshot_logo_path,
                        %s
                    ),

                team_b_snapshot_name =
                    COALESCE(
                        team_b_snapshot_name,
                        %s
                    ),

                team_b_snapshot_logo_path =
                    COALESCE(
                        team_b_snapshot_logo_path,
                        %s
                    )

            WHERE id = %s
            """,
            (
                completed_at,
                completed_at,

                series[
                    "team_a_current_team_name"
                ],

                series[
                    "team_a_current_team_logo_path"
                ],

                series[
                    "team_b_current_team_name"
                ],

                series[
                    "team_b_current_team_logo_path"
                ],

                series_id,
            ),
        )


                status = "completed"


            else:

                status = "active"


        connection.commit()


    # =========================
    # 세트별 스쿼드 Snapshot 저장
    # =========================

    save_series_set_squad_players(
        series_id,

        series[
            "team_a_id"
        ],
        nickname_a,

        series[
            "team_b_id"
        ],
        nickname_b,

        detected_matches,
    )




    # =========================
    # MVP
    # =========================

    mvp = None


    final_stats_sync_status = (
        series[
            "stats_sync_status"
        ]
    )


    if (
        status == "completed"
        and
        len(saved_sets) > 0
        and
        len(detected_matches)
            >= len(saved_sets)
    ):

        mvp_matches = [
            detected_match[
                "data"
            ]

            for detected_match
            in detected_matches[
                :len(saved_sets)
            ]
        ]


        (
            mvp,
            _,
            player_stats,
        ) = calculate_series_mvp_from_matches(
            mvp_matches
        )


        save_series_player_stats(
            series_id,

            series[
                "team_a_id"
            ],
            nickname_a,

            series[
                "team_b_id"
            ],
            nickname_b,

            player_stats,
        )


        if mvp:

            if (
                mvp["nickname"]
                == nickname_a
            ):

                mvp_participant_id = (
                    series[
                        "team_a_id"
                    ]
                )


            elif (
                mvp["nickname"]
                == nickname_b
            ):

                mvp_participant_id = (
                    series[
                        "team_b_id"
                    ]
                )


            else:

                mvp_participant_id = None


            if mvp_participant_id:

                save_series_mvp(
                    series_id,
                    mvp_participant_id,
                    mvp,
                )


        with get_db_connection() as connection:

            with connection.cursor() as cursor:

                cursor.execute(
                    """
                    UPDATE series

                    SET
                        stats_sync_status =
                            'synced'

                    WHERE id = %s
                    """,
                    (
                        series_id,
                    ),
                )

                settlement_result = (
                    settle_predictions_for_series(
                        cursor,
                        series_id,
                    )
                )

            connection.commit()


        final_stats_sync_status = (
            "synced"
        )


    # =========================
    # 응답용 세트
    # =========================

    sets = []


    for saved_set in saved_sets:

        sets.append(
            {
                "set":
                    saved_set[
                        "set_number"
                    ],

                "match_id":
                    saved_set[
                        "nexon_match_id"
                    ],

                "match_date":
                    (
                        saved_set[
                            "played_at"
                        ].isoformat()

                        if saved_set[
                            "played_at"
                        ]

                        else None
                    ),

                "team_a_score":
                    saved_set[
                        "team_a_score"
                    ],

                "team_b_score":
                    saved_set[
                        "team_b_score"
                    ],

                "winner_side":
                    saved_set[
                        "winner_side"
                    ],
            }
        )
    if (
        series["series_type"]
        == "정규리그"
        and
        status == "completed"
    ):
        create_initial_playoff_if_ready()

    if (
        series["series_type"]
        == "플레이오프"
        and
        status == "completed"
    ):
        create_next_playoff_if_ready(
            series_id
        )

    return {
        "series": {
            "series_id":
                series_id,

            "series_type":
                series[
                    "series_type"
                ],

            "team_a":
                series[
                    "team_a_name"
                ],

            "team_b":
                series[
                    "team_b_name"
                ],

            "nickname_a":
                nickname_a,

            "nickname_b":
                nickname_b,

            "match_type":
                series[
                    "match_type"
                ],

            "started_at":
                started_at.isoformat(),

            "status":
                status,

            "set_count":
                len(sets),

            "stats_sync_status":
                final_stats_sync_status,
        },

        "sets":
            sets,

        "mvp":
            mvp,
    }


# =========================
# SERIES MVP DB 조회
# =========================

@app.get(
    "/api/fconline/series/{series_id}/mvp"
)
def get_saved_series_mvp(
    series_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    sm.series_id,

                    p.fcl_name,
                    p.fc_nickname
                        AS nickname,

                    sm.sp_id,
                    sm.player_name,

                    sm.sets_played,

                    sm.rating_total,
                    sm.average_rating,

                    sm.goals,
                    sm.assists,

                    sm.image_url

                FROM series_mvp AS sm

                JOIN participants AS p
                    ON p.id =
                        sm.participant_id

                WHERE sm.series_id = %s
                """,
                (
                    series_id,
                ),
            )


            mvp = cursor.fetchone()


    if not mvp:

        raise HTTPException(
            status_code=404,
            detail=(
                "저장된 MVP가 없습니다."
            ),
        )


    # PostgreSQL NUMERIC -> JSON 숫자
    mvp["rating_total"] = float(
        mvp["rating_total"]
    )

    mvp["average_rating"] = float(
        mvp["average_rating"]
    )


    return mvp


# =========================
# SEASON CHAMPION
# 시즌 우승자 + FINAL MVP
# =========================

@app.get("/api/season/champion")
def get_season_champion():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 현재 시즌 결승 SERIES
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.status,
                    s.stats_sync_status,
                    s.best_of,
                    s.wins_required,

                    s.team_a_id,
                    s.team_b_id,

                    s.team_a_snapshot_name,
                    s.team_a_snapshot_logo_path,

                    s.team_b_snapshot_name,
                    s.team_b_snapshot_logo_path,

                    team_a.fcl_name
                        AS team_a_name,

                    team_a.fc_nickname
                        AS team_a_nickname,

                    team_b.fcl_name
                        AS team_b_name,

                    team_b.fc_nickname
                        AS team_b_nickname

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE
                    s.series_type = '플레이오프'

                    AND
                    s.playoff_stage = '결승시리즈'

                    AND
                    s.status <> 'cancelled'

                ORDER BY s.id DESC

                LIMIT 1
                """
            )


            final_series = cursor.fetchone()


            # 결승 자체가 아직 없음
            if not final_series:
                return {
                    "season": 1,
                    "completed": False,
                    "champion": None,
                    "final": None,
                    "final_mvp": None,
                }


            # 결승이 아직 진행 중
            if (
                final_series["status"]
                != "completed"
            ):
                return {
                    "season": 1,
                    "completed": False,
                    "champion": None,
                    "final": {
                        "series_id":
                            final_series["id"],

                        "status":
                            final_series["status"],
                    },
                    "final_mvp": None,
                }


            # =========================
            # 결승 세트 결과
            # =========================

            cursor.execute(
                """
                SELECT
                    set_number,
                    winner_side

                FROM series_sets

                WHERE series_id = %s

                ORDER BY set_number
                """,
                (
                    final_series["id"],
                ),
            )


            final_sets = cursor.fetchall()


            team_a_wins = 0
            team_b_wins = 0

            champion_side = None
            winning_set = None

            invalid_final = False


            for final_set in final_sets:

                winner_side = final_set[
                    "winner_side"
                ]


                if winner_side not in (
                    "team_a",
                    "team_b",
                ):
                    invalid_final = True
                    break


                # 이미 우승 확정 뒤
                # 추가 세트가 존재하면 비정상
                if champion_side is not None:
                    invalid_final = True
                    break


                if winner_side == "team_a":
                    team_a_wins += 1

                else:
                    team_b_wins += 1


                if (
                    team_a_wins
                    >= int(
                        final_series[
                            "wins_required"
                        ]
                    )
                ):
                    champion_side = "team_a"

                    winning_set = final_set[
                        "set_number"
                    ]


                elif (
                    team_b_wins
                    >= int(
                        final_series[
                            "wins_required"
                        ]
                    )
                ):
                    champion_side = "team_b"

                    winning_set = final_set[
                        "set_number"
                    ]


            # =========================
            # 실제 4승 미도달
            # 또는 비정상 데이터
            # =========================

            if (
                invalid_final
                or
                champion_side is None
            ):
                return {
                    "season": 1,
                    "completed": False,
                    "champion": None,

                    "final": {
                        "series_id":
                            final_series["id"],

                        "status":
                            final_series["status"],

                        "team_a_wins":
                            team_a_wins,

                        "team_b_wins":
                            team_b_wins,
                    },

                    "final_mvp": None,
                }


            # =========================
            # 우승자
            # =========================

            if champion_side == "team_a":

                champion = {
                    "participant_id":
                        final_series[
                            "team_a_id"
                        ],

                    "fcl_name":
                        final_series[
                            "team_a_name"
                        ],

                    "nickname":
                        final_series[
                            "team_a_nickname"
                        ],

                    "team_name":
                        final_series[
                            "team_a_snapshot_name"
                        ],

                    "team_logo_path":
                        final_series[
                            "team_a_snapshot_logo_path"
                        ],
                }

            else:

                champion = {
                    "participant_id":
                        final_series[
                            "team_b_id"
                        ],

                    "fcl_name":
                        final_series[
                            "team_b_name"
                        ],

                    "nickname":
                        final_series[
                            "team_b_nickname"
                        ],

                    "team_name":
                        final_series[
                            "team_b_snapshot_name"
                        ],

                    "team_logo_path":
                        final_series[
                            "team_b_snapshot_logo_path"
                        ],
                }


            # =========================
            # FINAL MVP
            # =========================

            cursor.execute(
                """
                SELECT
                    p.fcl_name,
                    p.fc_nickname
                        AS nickname,

                    sm.sp_id,
                    sm.player_name,
                    sm.sets_played,
                    sm.rating_total,
                    sm.average_rating,
                    sm.goals,
                    sm.assists,
                    sm.image_url

                FROM series_mvp AS sm

                JOIN participants AS p
                    ON p.id =
                        sm.participant_id

                WHERE sm.series_id = %s
                """,
                (
                    final_series["id"],
                ),
            )


            final_mvp_row = cursor.fetchone()


    final_mvp = None


    if final_mvp_row:

        final_mvp = {
            "fcl_name":
                final_mvp_row[
                    "fcl_name"
                ],

            "nickname":
                final_mvp_row[
                    "nickname"
                ],

            "sp_id":
                final_mvp_row[
                    "sp_id"
                ],

            "player_name":
                final_mvp_row[
                    "player_name"
                ],

            "sets_played":
                final_mvp_row[
                    "sets_played"
                ],

            "rating_total":
                float(
                    final_mvp_row[
                        "rating_total"
                    ]
                ),

            "average_rating":
                float(
                    final_mvp_row[
                        "average_rating"
                    ]
                ),

            "goals":
                final_mvp_row[
                    "goals"
                ],

            "assists":
                final_mvp_row[
                    "assists"
                ],

            "image_url":
                final_mvp_row[
                    "image_url"
                ],
        }


    return {
        "season": 1,

        "completed": True,

        "champion":
            champion,

        "final": {
            "series_id":
                final_series["id"],

            "best_of":
                final_series["best_of"],

            "wins_required":
                final_series[
                    "wins_required"
                ],

            "team_a_wins":
                team_a_wins,

            "team_b_wins":
                team_b_wins,

            "winning_set":
                winning_set,

            "stats_sync_status":
                final_series[
                    "stats_sync_status"
                ],
        },

        "final_mvp":
            final_mvp,
    }


# =========================
# 완료된 SERIES 결과 조회
# =========================

@app.get(
    "/api/fconline/series/completed-results"
)
def get_completed_series_results():

    results = []


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 완료 SERIES 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id AS series_id,
                    s.series_type,
                    s.round_number,

                    s.playoff_stage,
                    s.best_of,
                    s.wins_required,

                    s.scheduled_date,
                    s.started_at,
                    s.completed_at,
                    s.stats_sync_status,

                    s.team_a_snapshot_name,
                    s.team_a_snapshot_logo_path,

                    s.team_b_snapshot_name,
                    s.team_b_snapshot_logo_path,

                    team_a.fcl_name
                        AS team_a,

                    team_b.fcl_name
                        AS team_b,

                    mvp_owner.fcl_name
                        AS mvp_fcl_name,

                    mvp_owner.fc_nickname
                        AS mvp_nickname,

                    sm.sp_id
                        AS mvp_sp_id,

                    sm.player_name
                        AS mvp_player_name,

                    sm.sets_played
                        AS mvp_sets_played,

                    sm.rating_total
                        AS mvp_rating_total,

                    sm.average_rating
                        AS mvp_average_rating,

                    sm.goals
                        AS mvp_goals,

                    sm.assists
                        AS mvp_assists,

                    sm.image_url
                        AS mvp_image_url

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                LEFT JOIN series_mvp AS sm
                    ON sm.series_id =
                        s.id

                LEFT JOIN participants AS mvp_owner
                    ON mvp_owner.id =
                        sm.participant_id

                WHERE s.status = 'completed'

                ORDER BY s.completed_at DESC
                """
            )


            series_rows = (
                cursor.fetchall()
            )


            # =========================
            # SERIES별 세트 조회
            # =========================

            for series_row in series_rows:

                cursor.execute(
                    """
                    SELECT
                        set_number,
                        team_a_score,
                        team_b_score,
                        winner_side
                    FROM series_sets

                    WHERE series_id = %s

                    ORDER BY set_number
                    """,
                    (
                        series_row[
                            "series_id"
                        ],
                    ),
                )


                set_rows = (
                    cursor.fetchall()
                )


                # 3세트 미완성 데이터 보호
                if len(set_rows) < 3:
                    continue


                sets = []

                team_a_total_score = 0
                team_b_total_score = 0


                for set_row in set_rows:

                    team_a_score = (
                        set_row[
                            "team_a_score"
                        ]
                    )

                    team_b_score = (
                        set_row[
                            "team_b_score"
                        ]
                    )


                    team_a_total_score += (
                        team_a_score
                    )

                    team_b_total_score += (
                        team_b_score
                    )


                    sets.append(
                        {
                            "set":
                                set_row[
                                    "set_number"
                                ],

                            "team_a_score":
                                team_a_score,

                            "team_b_score":
                                team_b_score,

                            "winner_side":
                                set_row[
                                    "winner_side"
                                ],
                        }
                    )


                # =========================
                # MVP
                # =========================

                mvp = None


                if (
                    series_row[
                        "mvp_player_name"
                    ]
                    is not None
                ):

                    mvp = {
                        "fcl_name":
                            series_row[
                                "mvp_fcl_name"
                            ],

                        "nickname":
                            series_row[
                                "mvp_nickname"
                            ],

                        "sp_id":
                            series_row[
                                "mvp_sp_id"
                            ],

                        "player_name":
                            series_row[
                                "mvp_player_name"
                            ],

                        "sets_played":
                            series_row[
                                "mvp_sets_played"
                            ],

                        "rating_total":
                            float(
                                series_row[
                                    "mvp_rating_total"
                                ]
                            ),

                        "average_rating":
                            float(
                                series_row[
                                    "mvp_average_rating"
                                ]
                            ),

                        "goals":
                            series_row[
                                "mvp_goals"
                            ],

                        "assists":
                            series_row[
                                "mvp_assists"
                            ],

                        "image_url":
                            series_row[
                                "mvp_image_url"
                            ],
                    }


                started_at = parse_kst_datetime(
                    series_row[
                        "started_at"
                    ]
                )

                result_date = (
                    series_row[
                        "scheduled_date"
                    ]
                )


                if result_date is None:

                    result_date = (
                        started_at.date()
                    )


                results.append(
                    {
                        "series_id":
                            series_row[
                                "series_id"
                            ],

                        "source":
                            "database",

                        "stats_sync_status":
                            series_row[
                                "stats_sync_status"
                            ],

                        "date":
                            result_date.strftime(
                                "%Y-%m-%d"
                            ),

                        "completed_at":
                        (
                            series_row[
                                "completed_at"
                            ].isoformat()

                            if series_row[
                                "completed_at"
                            ]

                            else None
                        ),

                        "round":
                            series_row[
                                "round_number"
                            ],
                        "match_type":
                            series_row[
                                "series_type"
                            ],

                        "playoff_stage":
                            series_row[
                                "playoff_stage"
                            ],

                        "best_of":
                            series_row[
                                "best_of"
                            ],

                        "wins_required":
                            series_row[
                                "wins_required"
                            ],

                        "team_a":
                            series_row[
                                "team_a"
                            ],

                        "team_b":
                            series_row[
                                "team_b"
                            ],

                        "team_a_snapshot_name":
                            series_row[
                                "team_a_snapshot_name"
                            ],

                        "team_a_snapshot_logo_path":
                            series_row[
                                "team_a_snapshot_logo_path"
                            ],

                        "team_b_snapshot_name":
                            series_row[
                                "team_b_snapshot_name"
                            ],

                        "team_b_snapshot_logo_path":
                            series_row[
                                "team_b_snapshot_logo_path"
                            ],

                        "team_a_score":
                            team_a_total_score,

                        "team_b_score":
                            team_b_total_score,

                        "sets":
                            sets,

                        "mvp":
                            mvp,
                    }
                )


    return results

# =========================
# SERIES 전체 선수 기록 조회
# =========================

@app.get(
    "/api/fconline/series/{series_id}/player-stats"
)
def get_series_player_stats(
    series_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    sps.series_id,

                    p.fcl_name,
                    p.fc_nickname
                        AS nickname,

                    sps.sp_id,
                    sps.player_name,

                    sps.sets_played,

                    sps.rating_total,
                    sps.average_rating,

                    sps.goals,
                    sps.assists,

                    sps.image_url

                FROM series_player_stats
                    AS sps

                JOIN participants AS p
                    ON p.id =
                        sps.participant_id

                WHERE sps.series_id = %s

                ORDER BY
                    sps.goals DESC,
                    sps.assists DESC,
                    sps.rating_total DESC
                """,
                (
                    series_id,
                ),
            )


            players = cursor.fetchall()


    for player in players:

        player["rating_total"] = float(
            player["rating_total"]
        )

        player["average_rating"] = float(
            player["average_rating"]
        )


    return players

# =========================
# SERIES 세트별 스쿼드 조회
# =========================

@app.get(
    "/api/fconline/series/{series_id}/squads"
)
def get_series_squads(
    series_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # SERIES 정보
            # =========================

            cursor.execute(
                """
                SELECT
                    s.id,
                    s.series_type,
                    s.playoff_stage,
                    s.status,

                    team_a.id
                        AS team_a_id,

                    team_a.fcl_name
                        AS team_a_fcl_name,

                    COALESCE(
                        s.team_a_snapshot_name,
                        team_a.current_team_name
                    ) AS team_a_name,

                    COALESCE(
                        s.team_a_snapshot_logo_path,
                        team_a.current_team_logo_path
                    ) AS team_a_logo_path,

                    team_b.id
                        AS team_b_id,

                    team_b.fcl_name
                        AS team_b_fcl_name,

                    COALESCE(
                        s.team_b_snapshot_name,
                        team_b.current_team_name
                    ) AS team_b_name,

                    COALESCE(
                        s.team_b_snapshot_logo_path,
                        team_b.current_team_logo_path
                    ) AS team_b_logo_path

                FROM series AS s

                JOIN participants AS team_a
                    ON team_a.id =
                        s.team_a_id

                JOIN participants AS team_b
                    ON team_b.id =
                        s.team_b_id

                WHERE s.id = %s
                """,
                (
                    series_id,
                ),
            )

            series = cursor.fetchone()


            if not series:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "SERIES를 찾을 수 없습니다."
                    ),
                )


            # =========================
            # SET + Snapshot 선수
            # =========================

            cursor.execute(
                """
                SELECT
                    ss.set_number,
                    ss.played_at,

                    ss.team_a_score,
                    ss.team_b_score,
                    ss.winner_side,

                    sssp.side,
                    sssp.source_order,

                    sssp.sp_id,
                    sssp.player_name,

                    sssp.sp_position,
                    sssp.sp_grade,

                    sssp.rating,
                    sssp.goals,
                    sssp.assists,

                    sssp.image_url

                FROM series_sets AS ss

                LEFT JOIN
                    series_set_squad_players
                        AS sssp
                    ON sssp.series_set_id =
                        ss.id

                WHERE ss.series_id = %s

                ORDER BY
                    ss.set_number,
                    sssp.side,
                    sssp.source_order
                """,
                (
                    series_id,
                ),
            )

            rows = cursor.fetchall()


    # =========================
    # SET별 응답 구성
    # =========================

    set_map = {}


    for row in rows:

        set_number = row[
            "set_number"
        ]


        if set_number not in set_map:

            set_map[set_number] = {
                "set":
                    set_number,

                "played_at":
                    (
                        row[
                            "played_at"
                        ].isoformat()

                        if row[
                            "played_at"
                        ]

                        else None
                    ),

                "team_a_score":
                    row[
                        "team_a_score"
                    ],

                "team_b_score":
                    row[
                        "team_b_score"
                    ],

                "winner_side":
                    row[
                        "winner_side"
                    ],

                "team_a_squad": [],
                "team_b_squad": [],
            }


        # Snapshot이 없는 SET도
        # LEFT JOIN으로 반환 가능
        if row["side"] is None:
            continue


        player = {
            "source_order":
                row[
                    "source_order"
                ],

            "sp_id":
                row[
                    "sp_id"
                ],

            "player_name":
                row[
                    "player_name"
                ],

            "sp_position":
                row[
                    "sp_position"
                ],

            "sp_grade":
                row[
                    "sp_grade"
                ],

            "rating":
                float(
                    row[
                        "rating"
                    ]
                ),

            "goals":
                row[
                    "goals"
                ],

            "assists":
                row[
                    "assists"
                ],

            "image_url":
                row[
                    "image_url"
                ],
        }


        if row["side"] == "team_a":

            set_map[
                set_number
            ][
                "team_a_squad"
            ].append(
                player
            )


        elif row["side"] == "team_b":

            set_map[
                set_number
            ][
                "team_b_squad"
            ].append(
                player
            )


    return {
        "series_id":
            series["id"],

        "series_type":
            series[
                "series_type"
            ],

        "playoff_stage":
            series[
                "playoff_stage"
            ],

        "status":
            series[
                "status"
            ],

        "team_a": {
            "participant_id":
                series[
                    "team_a_id"
                ],

            "fcl_name":
                series[
                    "team_a_fcl_name"
                ],

            "team_name":
                series[
                    "team_a_name"
                ],

            "logo_path":
                series[
                    "team_a_logo_path"
                ],
        },

        "team_b": {
            "participant_id":
                series[
                    "team_b_id"
                ],

            "fcl_name":
                series[
                    "team_b_fcl_name"
                ],

            "team_name":
                series[
                    "team_b_name"
                ],

            "logo_path":
                series[
                    "team_b_logo_path"
                ],
        },

        "sets":
            list(
                set_map.values()
            ),
    }

# =========================
# DB 초기화
# =========================

@app.post("/api/database/init")
def init_database():

    initialize_database()


    # =========================
    # COMMUNITY NOTICE MIGRATION
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # PREDICTIONS
            # 승부예측
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    predictions (

                    id BIGSERIAL PRIMARY KEY,

                    user_id BIGINT
                        NOT NULL
                        REFERENCES users(id)
                        ON DELETE CASCADE,

                    series_id BIGINT
                        NOT NULL
                        REFERENCES series(id)
                        ON DELETE CASCADE,

                    predicted_participant_id BIGINT
                        REFERENCES participants(id),

                    prediction_type VARCHAR(20)
                        NOT NULL
                        DEFAULT 'participant',

                    stake_points INTEGER
                        NOT NULL,

                    odds NUMERIC(8, 2)
                        NOT NULL
                        DEFAULT 2.00,

                    status VARCHAR(20)
                        NOT NULL
                        DEFAULT 'pending',

                    payout_points INTEGER
                        NOT NULL
                        DEFAULT 0,

                    settled_at TIMESTAMPTZ,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    updated_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT
                        predictions_stake_check

                    CHECK (
                        prediction_type IN (
                            'participant',
                            'draw'
                        )
                    ),

                    CHECK (
                        stake_points > 0
                    ),

                    CONSTRAINT
                        predictions_odds_check

                    CHECK (
                        odds >= 1.00
                    ),

                    CONSTRAINT
                        predictions_payout_check

                    CHECK (
                        payout_points >= 0
                    ),

                    CONSTRAINT
                        predictions_status_check

                    CHECK (
                        status IN (
                            'pending',
                            'win',
                            'loss',
                            'refunded'
                        )
                    ),

                    CONSTRAINT
                        predictions_user_series_unique

                    UNIQUE (
                        user_id,
                        series_id
                    )
                )
                """
            )

            # =========================
            # PREDICTIONS
            # 무승부 예측 지원
            # =========================

            cursor.execute(
                """
                ALTER TABLE predictions

                ADD COLUMN IF NOT EXISTS
                    prediction_type VARCHAR(20)
                """
            )


            cursor.execute(
                """
                UPDATE predictions

                SET prediction_type =
                    'participant'

                WHERE prediction_type
                    IS NULL
                """
            )


            cursor.execute(
                """
                ALTER TABLE predictions

                ALTER COLUMN
                    predicted_participant_id

                DROP NOT NULL
                """
            )

            # =========================
            # PREDICTIONS
            # 세트별 승부예측 지원
            # =========================

            cursor.execute(
                """
                ALTER TABLE predictions

                ADD COLUMN IF NOT EXISTS
                    set_number INTEGER
                """
            )


            # 기존 테스트 예측 데이터는
            # 우선 1세트 예측으로 변환
            cursor.execute(
                """
                UPDATE predictions

                SET set_number = 1

                WHERE set_number IS NULL
                """
            )


            cursor.execute(
                """
                ALTER TABLE predictions

                ALTER COLUMN set_number
                SET NOT NULL
                """
            )


            # 기존 경기당 1회 UNIQUE 제거
            cursor.execute(
                """
                ALTER TABLE predictions

                DROP CONSTRAINT IF EXISTS
                    predictions_user_series_unique
                """
            )


            # 세트 번호 검증
            cursor.execute(
                """
                DO $$
                BEGIN

                    IF NOT EXISTS (
                        SELECT 1
                        FROM pg_constraint
                        WHERE conname =
                            'predictions_set_number_check'
                    ) THEN

                        ALTER TABLE predictions

                        ADD CONSTRAINT
                            predictions_set_number_check

                        CHECK (
                            set_number
                            BETWEEN 1 AND 3
                        );

                    END IF;

                END
                $$;
                """
            )


            # 회원 + 경기 + 세트별 1회
            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS
                    ux_predictions_user_series_set

                ON predictions (
                    user_id,
                    series_id,
                    set_number
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_predictions_series_set

                ON predictions (
                    series_id,
                    set_number,
                    status
                )
                """
            )



            # =========================
            # POINT TRANSACTION
            # IDEMPOTENCY
            # 중복 지급 / 환불 방지
            # =========================

            cursor.execute(
                """
                ALTER TABLE point_transactions

                ADD COLUMN IF NOT EXISTS
                    idempotency_key VARCHAR(150)
                """
            )


            cursor.execute(
                """
                CREATE UNIQUE INDEX IF NOT EXISTS
                    idx_point_transactions_idempotency_key

                ON point_transactions (
                    idempotency_key
                )

                WHERE
                    idempotency_key
                    IS NOT NULL
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_predictions_series_id

                ON predictions (
                    series_id
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_predictions_user_id

                ON predictions (
                    user_id,
                    created_at DESC
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_predictions_pending

                ON predictions (
                    series_id,
                    status
                )
                """
            )

            cursor.execute(
                """
                ALTER TABLE community_posts

                ADD COLUMN IF NOT EXISTS
                    is_notice BOOLEAN
                    NOT NULL
                    DEFAULT FALSE
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_community_posts_notice

                ON community_posts (
                    is_notice,
                    created_at DESC
                )
                """
            )

            # =========================
            # COMMUNITY COMMENTS
            # =========================

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS
                    community_comments (

                    id BIGSERIAL PRIMARY KEY,

                    post_id BIGINT NOT NULL
                        REFERENCES community_posts(id)
                        ON DELETE CASCADE,

                    user_id BIGINT NOT NULL
                        REFERENCES users(id)
                        ON DELETE CASCADE,

                    content TEXT NOT NULL,

                    created_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    updated_at TIMESTAMPTZ
                        NOT NULL
                        DEFAULT NOW(),

                    CONSTRAINT
                        community_comments_content_check

                    CHECK (
                        LENGTH(
                            BTRIM(content)
                        ) > 0
                    )
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_community_comments_post_id

                ON community_comments (
                    post_id,
                    created_at ASC,
                    id ASC
                )
                """
            )


            cursor.execute(
                """
                CREATE INDEX IF NOT EXISTS
                    idx_community_comments_user_id

                ON community_comments (
                    user_id
                )
                """
            )


        connection.commit()


    return {
        "status":
            "success",

        "message":
            "FCL 데이터베이스가 초기화되었습니다.",
    }

# =========================
# 참가자 DB 확인
# =========================

@app.get("/api/database/participants")
def get_database_participants():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    fcl_name,
                    fc_nickname,
                    ouid

                FROM participants

                ORDER BY id
                """
            )

            participants = (
                cursor.fetchall()
            )


    return participants

# 메타데이터 API
@app.get(
    "/api/fconline/metadata/seasons"
)
def get_fconline_season_metadata():

    seasons = {}


    # =========================
    # Nexon 메타데이터
    # 아직 서비스 중일 때 사용
    # =========================

    try:

        live_metadata = get_season_metadata()


        seasons.update(
            live_metadata
        )

    except Exception as error:

        print(
            "Nexon 시즌 메타데이터 "
            "조회 실패:",
            error,
        )


    # =========================
    # 우리 Snapshot
    # 항상 Snapshot 우선
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    season_id,
                    class_name

                FROM
                    fconline_season_snapshots

                ORDER BY
                    season_id
                """
            )


            snapshots = cursor.fetchall()


    for snapshot in snapshots:

        season_id = snapshot[
                "season_id"
            ]


        seasons[
            season_id
        ] = {
            "season_id":
                season_id,

            "class_name":
                snapshot[
                    "class_name"
                ],

            "season_image_url":
                (
                    "/api/fconline/"
                    "metadata/seasons/"
                    f"{season_id}/image"
                ),
        }


    return {
        "seasons":
            list(
                seasons.values()
            )
    }

@app.get(
    "/api/fconline/"
    "metadata/seasons/"
    "{season_id}/image"
)
def get_fconline_season_snapshot_image(
    season_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    image_data,
                    image_content_type

                FROM
                    fconline_season_snapshots

                WHERE
                    season_id = %s
                """,
                (
                    season_id,
                ),
            )


            snapshot = cursor.fetchone()


    if not snapshot:

        raise HTTPException(
            status_code=404,
            detail="저장된 시즌 이미지가 없습니다.",
        )


    return Response(
        content=
            bytes(
                snapshot[
                    "image_data"
                ]
            ),

        media_type=
            snapshot[
                "image_content_type"
            ],
    )

@app.post(
    "/api/admin/fconline/season-snapshots/backfill"
)
def admin_backfill_fconline_season_snapshots(
    admin_token: str =
        Depends(
            require_admin
        ),
):

    # =========================
    # 과거 기록에 등장한 모든 sp_id
    # =========================

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT sp_id
                FROM series_set_squad_players

                UNION

                SELECT sp_id
                FROM series_mvp

                UNION

                SELECT sp_id
                FROM series_player_stats
                """
            )


            rows = cursor.fetchall()


    # =========================
    # 시즌별 대표 sp_id 하나만 사용
    # =========================

    season_sp_ids = {}


    for row in rows:

        sp_id = row["sp_id"]


        if sp_id is None:
            continue


        season_id = (
            int(sp_id)
            // 1_000_000
        )


        if season_id not in season_sp_ids:

            season_sp_ids[
                season_id
            ] = int(sp_id)


    # =========================
    # Snapshot 저장
    # =========================

    saved = []
    failed = []


    for (
        season_id,
        sp_id,
    ) in sorted(
        season_sp_ids.items()
    ):

        try:

            snapshot = ensure_fconline_season_snapshot(
                    sp_id
                )


            saved.append(
                {
                    "season_id":
                        snapshot[
                            "season_id"
                        ],

                    "class_name":
                        snapshot[
                            "class_name"
                        ],
                }
            )


        except Exception as error:

            failed.append(
                {
                    "season_id":
                        season_id,

                    "error":
                        (
                            error.detail
                            if isinstance(
                                error,
                                HTTPException,
                            )
                            else str(error)
                        ),
                }
            )


    return {
        "target_season_count":
            len(
                season_sp_ids
            ),

        "snapshot_count":
            len(saved),

        "failed_count":
            len(failed),

        "snapshots":
            saved,

        "failed":
            failed,
    }

# =========================
# COMMUNITY
# 자유게시판
# =========================

def hash_community_password(
    password: str,
):

    iterations = 200_000

    salt = secrets.token_bytes(
        16
    )

    password_hash = (
        hashlib.pbkdf2_hmac(
            "sha256",
            password.encode(
                "utf-8"
            ),
            salt,
            iterations,
        )
    )

    return (
        f"pbkdf2_sha256$"
        f"{iterations}$"
        f"{salt.hex()}$"
        f"{password_hash.hex()}"
    )


def verify_community_password(
    password: str,
    stored_hash: str,
):

    if not stored_hash:
        return False


    try:

        (
            algorithm,
            iterations_text,
            salt_hex,
            expected_hash_hex,
        ) = stored_hash.split(
            "$"
        )


        if (
            algorithm
            != "pbkdf2_sha256"
        ):
            return False


        iterations = int(
            iterations_text
        )

        salt = bytes.fromhex(
            salt_hex
        )

        expected_hash = (
            bytes.fromhex(
                expected_hash_hex
            )
        )


        actual_hash = (
            hashlib.pbkdf2_hmac(
                "sha256",
                password.encode(
                    "utf-8"
                ),
                salt,
                iterations,
            )
        )


        return hmac.compare_digest(
            actual_hash,
            expected_hash,
        )

    except (
        ValueError,
        TypeError,
    ):

        return False

def get_community_post_attachments(
    cursor,
    post_id: int,
):

    cursor.execute(
        """
        SELECT
            id,
            original_file_name,
            content_type,
            sort_order

        FROM community_attachments

        WHERE post_id = %s

        ORDER BY
            sort_order ASC,
            id ASC
        """,
        (
            post_id,
        ),
    )

    rows = cursor.fetchall()

    return [
        {
            "id": row["id"],
            "original_file_name": row["original_file_name"],
            "content_type": row["content_type"],
            "sort_order": row["sort_order"],
            "image_url": (
                "/api/community/"
                f"attachments/{row['id']}"
            ),
        }
        for row in rows
    ]

@app.get("/api/community/posts")
def get_community_posts(
    board_type: str = "free",
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    cp.id,
                    cp.board_type,
                    cp.title,
                    cp.content,
                    cp.author_name,
                    cp.is_notice,
                    cp.created_at,
                    cp.updated_at,
                    COUNT(ca.id) AS attachment_count

                FROM community_posts AS cp

                LEFT JOIN community_attachments AS ca
                    ON ca.post_id = cp.id

                WHERE cp.board_type = %s

                GROUP BY
                    cp.id

                ORDER BY
                    cp.is_notice DESC,
                    cp.created_at DESC,
                    cp.id DESC
                """,
                (
                    board_type,
                ),
            )


            rows = cursor.fetchall()


    posts = []


    for row in rows:

        posts.append(
            {
                "id":
                    row["id"],

                "board_type":
                    row["board_type"],

                "is_notice":
                    bool(
                        row["is_notice"]
                    ),

                "title":
                    row["title"],

                "content":
                    row["content"],

                "author_name":
                    row["author_name"],

                "created_at":
                    row["created_at"].isoformat(),

                "updated_at":
                    row["updated_at"].isoformat(),

                "attachment_count":
                    int(row["attachment_count"]),
            }
        )


    return posts

@app.post("/api/community/posts")
def create_community_post(
    payload: dict,
):

    board_type = str(
            payload.get(
                "board_type",
                ""
            )
        ).strip()


    title = str(
            payload.get(
                "title",
                ""
            )
        ).strip()


    content = str(
            payload.get(
                "content",
                ""
            )
        ).strip()


    author_name = str(
            payload.get(
                "author_name",
                ""
            )
        ).strip()

    password = str(
            payload.get(
                "password",
                ""
            )
        ).strip()


    if board_type not in (
        "free",
        "player_photo_request",
    ):

        raise HTTPException(
            status_code=400,
            detail="올바르지 않은 게시판 유형입니다.",
        )


    if not title:

        raise HTTPException(
            status_code=400,
            detail="제목을 입력해주세요.",
        )


    if not content:

        raise HTTPException(
            status_code=400,
            detail="내용을 입력해주세요.",
        )


    if not author_name:

        raise HTTPException(
            status_code=400,
            detail="작성자를 입력해주세요.",
        )

    if len(password) < 4:

        raise HTTPException(
            status_code=400,
            detail="비밀번호는 4자 이상 입력해주세요.",
        )


    if len(password) > 50:

        raise HTTPException(
            status_code=400,
            detail="비밀번호는 50자 이하로 입력해주세요.",
        )


    password_hash = (
        hash_community_password(
            password
        )
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO community_posts (
                    board_type,
                    title,
                    content,
                    author_name,
                    password_hash
                )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    %s
                )

                RETURNING
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    created_at,
                    updated_at
                """,
                (
                    board_type,
                    title,
                    content,
                    author_name,
                    password_hash,
                ),
            )


            row = cursor.fetchone()


        connection.commit()


    return {
        "id":
            row["id"],

        "board_type":
            row["board_type"],

        "title":
            row["title"],

        "content":
            row["content"],

        "author_name":
            row["author_name"],

        "created_at":
            row["created_at"].isoformat(),

        "updated_at":
            row["updated_at"].isoformat(),
    }



@app.get("/api/community/posts/{post_id}")
def get_community_post(
    post_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    is_notice,
                    created_at,
                    updated_at

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )

            row = cursor.fetchone()

            if not row:

                raise HTTPException(
                    status_code=404,
                    detail="게시글을 찾을 수 없습니다.",
                )

            attachments = (
                get_community_post_attachments(
                    cursor,
                    post_id,
                )
            )


    return {
        "id":
            row["id"],

        "board_type":
            row["board_type"],

        "is_notice":
            bool(
                row["is_notice"]
            ),

        "title":
            row["title"],

        "content":
            row["content"],

        "author_name":
            row["author_name"],

        "created_at":
            row["created_at"].isoformat(),

        "updated_at":
            row["updated_at"].isoformat(),

        "attachments":
            attachments,
    }

# =========================
# COMMUNITY COMMENTS
# 댓글 조회
# =========================

@app.get(
    "/api/community/posts/{post_id}/comments"
)
def get_community_comments(
    post_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 게시글 존재 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    board_type
                FROM community_posts
                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


            post = cursor.fetchone()


            if not post:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "게시글을 찾을 수 없습니다."
                    ),
                )


            # 자유게시판에만 댓글 사용
            if (
                post["board_type"]
                != "free"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "자유게시판 게시글의 "
                        "댓글만 조회할 수 있습니다."
                    ),
                )


            # =========================
            # 댓글 조회
            # =========================

            cursor.execute(
                """
                SELECT
                    cc.id,
                    cc.post_id,
                    cc.user_id,
                    cc.content,
                    cc.created_at,
                    cc.updated_at,

                    u.nickname
                        AS author_nickname,

                    u.is_admin
                        AS author_is_admin

                FROM community_comments
                    AS cc

                INNER JOIN users AS u
                    ON u.id = cc.user_id

                WHERE
                    cc.post_id = %s

                ORDER BY
                    cc.created_at ASC,
                    cc.id ASC
                """,
                (
                    post_id,
                ),
            )


            rows = cursor.fetchall()


    return [
        {
            "id":
                row["id"],

            "post_id":
                row["post_id"],

            "user_id":
                row["user_id"],

            "author_nickname":
                row[
                    "author_nickname"
                ],

            "author_is_admin":
                bool(
                    row[
                        "author_is_admin"
                    ]
                ),

            "content":
                row["content"],

            "created_at":
                row[
                    "created_at"
                ].isoformat(),

            "updated_at":
                row[
                    "updated_at"
                ].isoformat(),
        }

        for row in rows
    ]


# =========================
# COMMUNITY COMMENTS
# 댓글 작성
# =========================

@app.post(
    "/api/community/posts/{post_id}/comments"
)
def create_community_comment(
    post_id: int,

    request:
        CommunityCommentCreateRequest,

    user = Depends(
        require_user
    ),
):

    content = (
        request.content
        .strip()
    )


    # =========================
    # 내용 검증
    # =========================

    if not content:

        raise HTTPException(
            status_code=400,
            detail=(
                "댓글 내용을 입력해주세요."
            ),
        )


    if len(content) > 1000:

        raise HTTPException(
            status_code=400,
            detail=(
                "댓글은 1000자 이하로 "
                "입력해주세요."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            # =========================
            # 게시글 존재 확인
            # =========================

            cursor.execute(
                """
                SELECT
                    id,
                    board_type
                FROM community_posts
                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


            post = cursor.fetchone()


            if not post:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "게시글을 찾을 수 없습니다."
                    ),
                )


            if (
                post["board_type"]
                != "free"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "자유게시판 게시글에만 "
                        "댓글을 작성할 수 있습니다."
                    ),
                )


            # =========================
            # 댓글 저장
            # =========================

            cursor.execute(
                """
                INSERT INTO
                    community_comments (
                        post_id,
                        user_id,
                        content
                    )

                VALUES (
                    %s,
                    %s,
                    %s
                )

                RETURNING
                    id,
                    post_id,
                    user_id,
                    content,
                    created_at,
                    updated_at
                """,
                (
                    post_id,
                    user["id"],
                    content,
                ),
            )


            comment = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "message":
            "댓글이 등록되었습니다.",

        "comment": {
            "id":
                comment["id"],

            "post_id":
                comment["post_id"],

            "user_id":
                comment["user_id"],

            "author_nickname":
                user["nickname"],

            "author_is_admin":
                bool(
                    user["is_admin"]
                ),

            "content":
                comment["content"],

            "created_at":
                comment[
                    "created_at"
                ].isoformat(),

            "updated_at":
                comment[
                    "updated_at"
                ].isoformat(),
        },
    }


@app.patch(
    "/api/community/posts/{post_id}"
)
def update_community_post(
    post_id: int,
    payload: dict,
):

    password = str(
        payload.get(
            "password",
            ""
        )
    ).strip()

    title = str(
        payload.get(
            "title",
            ""
        )
    ).strip()

    author_name = str(
        payload.get(
            "author_name",
            ""
        )
    ).strip()

    content = str(
        payload.get(
            "content",
            ""
        )
    ).strip()


    if not password:

        raise HTTPException(
            status_code=400,
            detail="비밀번호를 입력해주세요.",
        )


    if not title:

        raise HTTPException(
            status_code=400,
            detail="제목을 입력해주세요.",
        )


    if not author_name:

        raise HTTPException(
            status_code=400,
            detail="작성자를 입력해주세요.",
        )


    if not content:

        raise HTTPException(
            status_code=400,
            detail="내용을 입력해주세요.",
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    password_hash,
                    is_notice

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )

            post_row = cursor.fetchone()


            if not post_row:

                raise HTTPException(
                    status_code=404,
                    detail="게시글을 찾을 수 없습니다.",
                )

            if post_row["is_notice"]:

                raise HTTPException(
                    status_code=403,
                    detail=(
                        "공지사항은 관리자만 "
                        "수정할 수 있습니다."
                    ),
                )


            if not verify_community_password(
                password,
                post_row[
                    "password_hash"
                ],
            ):

                raise HTTPException(
                    status_code=403,
                    detail="비밀번호가 일치하지 않습니다.",
                )


            cursor.execute(
                """
                UPDATE community_posts

                SET
                    title = %s,
                    author_name = %s,
                    content = %s,
                    updated_at = NOW()

                WHERE id = %s

                RETURNING
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    created_at,
                    updated_at
                """,
                (
                    title,
                    author_name,
                    content,
                    post_id,
                ),
            )

            row = cursor.fetchone()


        connection.commit()


    return {
        "id": row["id"],
        "board_type": row["board_type"],
        "title": row["title"],
        "content": row["content"],
        "author_name": row["author_name"],
        "created_at": row["created_at"].isoformat(),
        "updated_at": row["updated_at"].isoformat(),
    }

@app.delete(
    "/api/community/posts/{post_id}"
)
def delete_community_post(
    post_id: int,
    payload: dict,
):

    password = str(
        payload.get(
            "password",
            ""
        )
    ).strip()


    if not password:

        raise HTTPException(
            status_code=400,
            detail="비밀번호를 입력해주세요.",
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    password_hash,
                    is_notice

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )

            post_row = cursor.fetchone()


            if not post_row:

                raise HTTPException(
                    status_code=404,
                    detail="게시글을 찾을 수 없습니다.",
                )

            if post_row["is_notice"]:

                raise HTTPException(
                    status_code=403,
                    detail=(
                        "공지사항은 관리자만 "
                        "삭제할 수 있습니다."
                    ),
                )


            if not verify_community_password(
                password,
                post_row[
                    "password_hash"
                ],
            ):

                raise HTTPException(
                    status_code=403,
                    detail="비밀번호가 일치하지 않습니다.",
                )


            cursor.execute(
                """
                DELETE FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


        connection.commit()


    return {
        "deleted": True,
        "post_id": post_id,
    }

# =========================
# COMMUNITY COMMENTS
# 댓글 수정
# =========================

@app.patch(
    "/api/community/comments/{comment_id}"
)
def update_community_comment(
    comment_id: int,

    request:
        CommunityCommentUpdateRequest,

    user = Depends(
        require_user
    ),
):

    content = (
        request.content
        .strip()
    )


    if not content:

        raise HTTPException(
            status_code=400,
            detail="댓글 내용을 입력해주세요.",
        )


    if len(content) > 1000:

        raise HTTPException(
            status_code=400,
            detail=(
                "댓글은 1000자 이하로 "
                "입력해주세요."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    post_id,
                    user_id
                FROM community_comments
                WHERE id = %s
                """,
                (
                    comment_id,
                ),
            )


            comment = cursor.fetchone()


            if not comment:

                raise HTTPException(
                    status_code=404,
                    detail="댓글을 찾을 수 없습니다.",
                )


            # 작성자 본인만 수정 가능
            if (
                comment["user_id"]
                != user["id"]
            ):

                raise HTTPException(
                    status_code=403,
                    detail=(
                        "본인이 작성한 댓글만 "
                        "수정할 수 있습니다."
                    ),
                )


            cursor.execute(
                """
                UPDATE community_comments

                SET
                    content = %s,
                    updated_at = NOW()

                WHERE id = %s

                RETURNING
                    id,
                    post_id,
                    user_id,
                    content,
                    created_at,
                    updated_at
                """,
                (
                    content,
                    comment_id,
                ),
            )


            updated_comment = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "message":
            "댓글이 수정되었습니다.",

        "comment": {
            "id":
                updated_comment["id"],

            "post_id":
                updated_comment["post_id"],

            "user_id":
                updated_comment["user_id"],

            "author_nickname":
                user["nickname"],

            "author_is_admin":
                bool(
                    user["is_admin"]
                ),

            "content":
                updated_comment["content"],

            "created_at":
                updated_comment[
                    "created_at"
                ].isoformat(),

            "updated_at":
                updated_comment[
                    "updated_at"
                ].isoformat(),
        },
    }

# =========================
# COMMUNITY COMMENTS
# 댓글 삭제
# =========================

@app.delete(
    "/api/community/comments/{comment_id}"
)
def delete_community_comment(
    comment_id: int,

    user = Depends(
        require_user
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    post_id,
                    user_id

                FROM community_comments

                WHERE id = %s
                """,
                (
                    comment_id,
                ),
            )


            comment = cursor.fetchone()


            if not comment:

                raise HTTPException(
                    status_code=404,
                    detail="댓글을 찾을 수 없습니다.",
                )


            is_owner = (
                comment["user_id"]
                == user["id"]
            )


            is_admin = bool(
                user["is_admin"]
            )


            # 작성자 또는 관리자만 삭제 가능
            if (
                not is_owner
                and
                not is_admin
            ):

                raise HTTPException(
                    status_code=403,
                    detail=(
                        "댓글을 삭제할 "
                        "권한이 없습니다."
                    ),
                )


            cursor.execute(
                """
                DELETE FROM
                    community_comments

                WHERE id = %s
                """,
                (
                    comment_id,
                ),
            )


        connection.commit()


    return {
        "deleted":
            True,

        "comment_id":
            comment_id,

        "post_id":
            comment["post_id"],
    }

@app.post(
    "/api/community/admin/notices"
)
def create_community_notice(
    request:
        AdminCommunityNoticeRequest,

    admin_user = Depends(
        require_user_admin
    ),
):

    title = (
        request.title
        .strip()
    )


    content = (
        request.content
        .strip()
    )


    if not title:

        raise HTTPException(
            status_code=400,
            detail=(
                "제목을 입력해주세요."
            ),
        )


    if not content:

        raise HTTPException(
            status_code=400,
            detail=(
                "내용을 입력해주세요."
            ),
        )


    if len(title) > 200:

        raise HTTPException(
            status_code=400,
            detail=(
                "제목은 200자 이하로 "
                "입력해주세요."
            ),
        )


    # 일반 게시글 비밀번호 수정 기능으로
    # 공지를 수정할 수 없도록
    # 아무도 알 수 없는 임의 비밀번호 생성
    random_password = (
        secrets.token_urlsafe(
            32
        )
    )


    password_hash = (
        hash_community_password(
            random_password
        )
    )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO community_posts (
                    board_type,
                    title,
                    content,
                    author_name,
                    password_hash,
                    is_notice
                )

                VALUES (
                    'free',
                    %s,
                    %s,
                    %s,
                    %s,
                    TRUE
                )

                RETURNING
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    is_notice,
                    created_at,
                    updated_at
                """,
                (
                    title,
                    content,
                    admin_user[
                        "nickname"
                    ],
                    password_hash,
                ),
            )


            notice = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "message":
            "공지사항이 등록되었습니다.",

        "notice": {
            "id":
                notice["id"],

            "board_type":
                notice[
                    "board_type"
                ],

            "title":
                notice["title"],

            "content":
                notice["content"],

            "author_name":
                notice[
                    "author_name"
                ],

            "is_notice":
                notice[
                    "is_notice"
                ],

            "created_at":
                notice[
                    "created_at"
                ].isoformat(),

            "updated_at":
                notice[
                    "updated_at"
                ].isoformat(),
        },
    }

# =========================
# COMMUNITY ADMIN
# =========================

@app.patch(
    "/api/community/admin/posts/{post_id}"
)
def admin_update_community_post(
    post_id: int,

    request:
        AdminCommunityPostUpdateRequest,

    admin_user = Depends(
        require_user_admin
    ),
):

    title = (
        request.title
        .strip()
    )

    content = (
        request.content
        .strip()
    )


    if not title:

        raise HTTPException(
            status_code=400,
            detail=(
                "제목을 입력해주세요."
            ),
        )


    if not content:

        raise HTTPException(
            status_code=400,
            detail=(
                "내용을 입력해주세요."
            ),
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    board_type

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


            post = (
                cursor.fetchone()
            )


            if not post:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "게시글을 찾을 수 없습니다."
                    ),
                )


            if (
                post["board_type"]
                != "free"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "자유게시판 게시글만 "
                        "관리할 수 있습니다."
                    ),
                )


            cursor.execute(
                """
                UPDATE community_posts

                SET
                    title = %s,
                    content = %s,
                    updated_at = NOW()

                WHERE id = %s

                RETURNING
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    created_at,
                    updated_at
                """,
                (
                    title,
                    content,
                    post_id,
                ),
            )


            updated_post = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "message":
            "게시글이 관리자에 의해 수정되었습니다.",

        "post": {
            "id":
                updated_post["id"],

            "board_type":
                updated_post["board_type"],

            "title":
                updated_post["title"],

            "content":
                updated_post["content"],

            "author_name":
                updated_post["author_name"],

            "created_at":
                updated_post[
                    "created_at"
                ].isoformat(),

            "updated_at":
                updated_post[
                    "updated_at"
                ].isoformat(),
        },
    }


@app.delete(
    "/api/community/admin/posts/{post_id}"
)
def admin_delete_community_post(
    post_id: int,

    admin_user = Depends(
        require_user_admin
    ),
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    id,
                    board_type,
                    title

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


            post = (
                cursor.fetchone()
            )


            if not post:

                raise HTTPException(
                    status_code=404,
                    detail=(
                        "게시글을 찾을 수 없습니다."
                    ),
                )


            if (
                post["board_type"]
                != "free"
            ):

                raise HTTPException(
                    status_code=400,
                    detail=(
                        "자유게시판 게시글만 "
                        "관리할 수 있습니다."
                    ),
                )


            # 첨부파일 레코드 먼저 제거
            cursor.execute(
                """
                DELETE FROM community_attachments

                WHERE post_id = %s
                """,
                (
                    post_id,
                ),
            )


            cursor.execute(
                """
                DELETE FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


        connection.commit()


    return {
        "deleted":
            True,

        "post_id":
            post_id,

        "message":
            "게시글이 관리자에 의해 삭제되었습니다.",
    }

# =========================
# PLAYER PHOTO REQUEST
# =========================

@app.post(
    "/api/community/player-photo-requests"
)
def create_player_photo_request(
    payload: dict,
):

    player_name = str(
        payload.get(
            "player_name",
            ""
        )
    ).strip()

    season_name = str(
        payload.get(
            "season_name",
            ""
        )
    ).strip()

    sp_id_value = payload.get(
        "sp_id"
    )

    content = str(
        payload.get(
            "content",
            ""
        )
    ).strip()

    author_name = str(
        payload.get(
            "author_name",
            ""
        )
    ).strip()

    password = str(
        payload.get(
            "password",
            ""
        )
    ).strip()


    if not player_name:

        raise HTTPException(
            status_code=400,
            detail="선수명을 입력해주세요.",
        )


    if not content:

        raise HTTPException(
            status_code=400,
            detail="요청 내용을 입력해주세요.",
        )


    if not author_name:

        raise HTTPException(
            status_code=400,
            detail="작성자 닉네임을 입력해주세요.",
        )


    if len(password) < 4:

        raise HTTPException(
            status_code=400,
            detail="비밀번호는 4자 이상 입력해주세요.",
        )


    if len(password) > 50:

        raise HTTPException(
            status_code=400,
            detail="비밀번호는 50자 이하로 입력해주세요.",
        )


    sp_id = None


    if (
        sp_id_value is not None
        and
        str(sp_id_value).strip()
    ):

        try:

            sp_id = int(
                sp_id_value
            )

        except (
            TypeError,
            ValueError,
        ):

            raise HTTPException(
                status_code=400,
                detail="sp_id가 올바르지 않습니다.",
            )


    password_hash = (
        hash_community_password(
            password
        )
    )


    if season_name:

        title = (
            f"{player_name} "
            f"({season_name}) "
            "선수 사진 요청"
        )

    else:

        title = (
            f"{player_name} "
            "선수 사진 요청"
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO community_posts (
                    board_type,
                    title,
                    content,
                    author_name,
                    password_hash
                )

                VALUES (
                    'player_photo_request',
                    %s,
                    %s,
                    %s,
                    %s
                )

                RETURNING
                    id,
                    board_type,
                    title,
                    content,
                    author_name,
                    created_at,
                    updated_at
                """,
                (
                    title,
                    content,
                    author_name,
                    password_hash,
                ),
            )

            post_row = cursor.fetchone()


            cursor.execute(
                """
                INSERT INTO player_photo_requests (
                    post_id,
                    player_name,
                    season_name,
                    sp_id,
                    request_status
                )

                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    'pending'
                )

                RETURNING
                    player_name,
                    season_name,
                    sp_id,
                    request_status,
                    admin_note,
                    completed_at
                """,
                (
                    post_row["id"],
                    player_name,
                    (
                        season_name
                        if season_name
                        else None
                    ),
                    sp_id,
                ),
            )

            request_row = (
                cursor.fetchone()
            )


        connection.commit()


    return {
        "id":
            post_row["id"],

        "board_type":
            post_row["board_type"],

        "title":
            post_row["title"],

        "content":
            post_row["content"],

        "author_name":
            post_row["author_name"],

        "created_at":
            post_row[
                "created_at"
            ].isoformat(),

        "updated_at":
            post_row[
                "updated_at"
            ].isoformat(),

        "player_name":
            request_row[
                "player_name"
            ],

        "season_name":
            request_row[
                "season_name"
            ],

        "sp_id":
            request_row[
                "sp_id"
            ],

        "request_status":
            request_row[
                "request_status"
            ],

        "admin_note":
            request_row[
                "admin_note"
            ],

        "completed_at":
            (
                request_row[
                    "completed_at"
                ].isoformat()

                if request_row[
                    "completed_at"
                ]

                else None
            ),
    }

@app.get(
    "/api/community/player-photo-requests"
)
def get_player_photo_requests():

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    cp.id,
                    cp.title,
                    cp.content,
                    cp.author_name,
                    cp.created_at,
                    cp.updated_at,

                    ppr.player_name,
                    ppr.season_name,
                    ppr.sp_id,
                    ppr.request_status,
                    ppr.admin_note,
                    ppr.completed_at,

                    COUNT(ca.id) AS attachment_count

                FROM community_posts cp

                INNER JOIN player_photo_requests ppr
                    ON ppr.post_id = cp.id

                LEFT JOIN community_attachments ca
                    ON ca.post_id = cp.id

                WHERE
                    cp.board_type = 'player_photo_request'

                GROUP BY
                    cp.id,
                    ppr.post_id

                ORDER BY
                    cp.id DESC
                """
            )

            rows = cursor.fetchall()


    return [
        {
            "id":
                row["id"],

            "title":
                row["title"],

            "content":
                row["content"],

            "author_name":
                row["author_name"],

            "created_at":
                row[
                    "created_at"
                ].isoformat(),

            "updated_at":
                row[
                    "updated_at"
                ].isoformat(),

            "player_name":
                row["player_name"],

            "season_name":
                row["season_name"],

            "sp_id":
                row["sp_id"],

            "request_status":
                row[
                    "request_status"
                ],

            "admin_note":
                row["admin_note"],

            "completed_at":
                (
                    row[
                        "completed_at"
                    ].isoformat()

                    if row[
                        "completed_at"
                    ]

                    else None
                ),

            "attachment_count":
                row[
                    "attachment_count"
                ],
        }

        for row in rows
    ]

@app.get(
    "/api/community/player-photo-requests/{post_id}"
)
def get_player_photo_request(
    post_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    cp.id,
                    cp.board_type,
                    cp.title,
                    cp.content,
                    cp.author_name,
                    cp.created_at,
                    cp.updated_at,

                    ppr.player_name,
                    ppr.season_name,
                    ppr.sp_id,
                    ppr.request_status,
                    ppr.admin_note,
                    ppr.completed_at

                FROM community_posts cp

                INNER JOIN player_photo_requests ppr
                    ON ppr.post_id = cp.id

                WHERE
                    cp.id = %s
                    AND
                    cp.board_type = 'player_photo_request'
                """,
                (
                    post_id,
                ),
            )

            row = cursor.fetchone()


            if not row:

                raise HTTPException(
                    status_code=404,
                    detail="선수 사진 요청을 찾을 수 없습니다.",
                )


            attachments = (
                get_community_post_attachments(
                    cursor,
                    post_id,
                )
            )


    return {
        "id":
            row["id"],

        "board_type":
            row["board_type"],

        "title":
            row["title"],

        "content":
            row["content"],

        "author_name":
            row["author_name"],

        "created_at":
            row[
                "created_at"
            ].isoformat(),

        "updated_at":
            row[
                "updated_at"
            ].isoformat(),

        "player_name":
            row["player_name"],

        "season_name":
            row["season_name"],

        "sp_id":
            row["sp_id"],

        "request_status":
            row[
                "request_status"
            ],

        "admin_note":
            row["admin_note"],

        "completed_at":
            (
                row[
                    "completed_at"
                ].isoformat()

                if row[
                    "completed_at"
                ]

                else None
            ),

        "attachments":
            attachments,
    }

@app.patch(
    "/api/admin/community/player-photo-requests/{post_id}"
)
def update_admin_player_photo_request(
    post_id: int,
    payload: dict,
    _admin=Depends(require_admin),
):

    request_status = str(
        payload.get(
            "request_status",
            ""
        )
    ).strip()

    admin_note = str(
        payload.get(
            "admin_note",
            ""
        )
    ).strip()


    allowed_statuses = {
        "pending",
        "in_progress",
        "completed",
        "rejected",
    }


    if (
        request_status
        not in allowed_statuses
    ):

        raise HTTPException(
            status_code=400,
            detail="올바르지 않은 처리 상태입니다.",
        )


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE player_photo_requests

                SET
                    request_status = %s,
                    admin_note = %s,
                    completed_at =
                        CASE
                            WHEN %s = 'completed'
                            THEN COALESCE(
                                completed_at,
                                NOW()
                            )
                            ELSE NULL
                        END

                WHERE post_id = %s

                RETURNING
                    post_id,
                    player_name,
                    season_name,
                    sp_id,
                    request_status,
                    admin_note,
                    completed_at
                """,
                (
                    request_status,
                    (
                        admin_note
                        if admin_note
                        else None
                    ),
                    request_status,
                    post_id,
                ),
            )

            row = cursor.fetchone()


            if not row:

                raise HTTPException(
                    status_code=404,
                    detail="선수 사진 요청을 찾을 수 없습니다.",
                )


        connection.commit()


    return {
        "post_id":
            row["post_id"],

        "player_name":
            row["player_name"],

        "season_name":
            row["season_name"],

        "sp_id":
            row["sp_id"],

        "request_status":
            row[
                "request_status"
            ],

        "admin_note":
            row["admin_note"],

        "completed_at":
            (
                row[
                    "completed_at"
                ].isoformat()

                if row[
                    "completed_at"
                ]

                else None
            ),
    }

# =========================
# COMMUNITY ATTACHMENTS
# =========================

@app.post(
    "/api/community/posts/{post_id}/attachments"
)
async def upload_community_attachments(
    post_id: int,
    files: list[UploadFile] = File(...),
):

    if len(files) > 5:

        raise HTTPException(
            status_code=400,
            detail="이미지는 최대 5장까지 첨부할 수 있습니다.",
        )


    allowed_content_types = {
        "image/jpeg",
        "image/png",
        "image/webp",
        "image/gif",
    }


    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT id

                FROM community_posts

                WHERE id = %s
                """,
                (
                    post_id,
                ),
            )


            post_row = cursor.fetchone()


            if not post_row:

                raise HTTPException(
                    status_code=404,
                    detail="게시글을 찾을 수 없습니다.",
                )


            cursor.execute(
                """
                SELECT COUNT(*) AS count

                FROM community_attachments

                WHERE post_id = %s
                """,
                (
                    post_id,
                ),
            )


            attachment_count = int(
                    cursor.fetchone()["count"]
                )


            if (
                attachment_count
                + len(files)
                > 5
            ):

                raise HTTPException(
                    status_code=400,
                    detail="게시글당 이미지는 최대 5장까지 첨부할 수 있습니다.",
                )


            uploaded_attachments = []


            for index, upload_file in enumerate(
                files,
                start=attachment_count,
            ):

                if (
                    upload_file.content_type
                    not in allowed_content_types
                ):

                    raise HTTPException(
                        status_code=400,
                        detail="지원하지 않는 이미지 형식입니다.",
                    )


                image_data = await upload_file.read()


                if len(image_data) > 5 * 1024 * 1024:

                    raise HTTPException(
                        status_code=400,
                        detail="이미지 한 장의 최대 크기는 5MB입니다.",
                    )


                cursor.execute(
                    """
                    INSERT INTO community_attachments (
                        post_id,
                        original_file_name,
                        content_type,
                        image_data,
                        sort_order
                    )

                    VALUES (
                        %s,
                        %s,
                        %s,
                        %s,
                        %s
                    )

                    RETURNING
                        id,
                        original_file_name,
                        content_type,
                        sort_order
                    """,
                    (
                        post_id,
                        upload_file.filename
                        or "image",
                        upload_file.content_type,
                        image_data,
                        index,
                    ),
                )


                attachment_row = cursor.fetchone()


                uploaded_attachments.append(
                    {
                        "id":
                            attachment_row["id"],

                        "original_file_name":
                            attachment_row[
                                "original_file_name"
                            ],

                        "content_type":
                            attachment_row[
                                "content_type"
                            ],

                        "sort_order":
                            attachment_row[
                                "sort_order"
                            ],

                        "image_url":
                            (
                                "/api/community/"
                                "attachments/"
                                f"{attachment_row['id']}"
                            ),
                    }
                )


        connection.commit()


    return {
        "post_id":
            post_id,

        "attachments":
            uploaded_attachments,
    }

@app.get(
    "/api/community/attachments/{attachment_id}"
)
def get_community_attachment(
    attachment_id: int,
):

    with get_db_connection() as connection:

        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT
                    content_type,
                    image_data

                FROM community_attachments

                WHERE id = %s
                """,
                (
                    attachment_id,
                ),
            )


            row = cursor.fetchone()


    if not row:

        raise HTTPException(
            status_code=404,
            detail="첨부 이미지를 찾을 수 없습니다.",
        )


    return Response(
        content=bytes(
            row["image_data"]
        ),
        media_type=row["content_type"],
    )


# =========================
# 프론트엔드
#
# 반드시 API 선언보다 아래에 위치
# =========================

app.mount(
    "/",
    StaticFiles(
        directory=FRONTEND_DIR,
        html=True,
    ),
    name="frontend",
)